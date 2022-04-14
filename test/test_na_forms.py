"""Test Lexedata handling of NA values

In https://github.com/cldf/cldf/issues/108, we started a discussion concerning
different types of NA values in lexical data.

CLDF formalizes the following distinction:

> Data creators often want to distinguish two kinds of missing data, in
> particular when the data is extracted from sources:
>
> 1. data that is missing/unknown because it was never extracted from the source,
> 2. data that is indicated in the source as unknown.
>
> The CSVW data model can be used to express this difference as follows: Case 1
> can be modeled by not including the relevant data as row at all. Case 2 can
> be modeled using the null property of the relevant column specification as
> value in a data row.

There is, however, a third case: Data where the source states that the
parameter value is not applicable to the language. We encode this by
"-". We call these “NA forms”.

Some source datasets use "?" to indicate cases 1 and 2, this needs to be
handled upon import.

Special handling of these three different NA/Missing forms alongside valid forms is
tested by this module. It affects multiple different components of Lexedata.

"""
import io
from pathlib import Path
import tempfile
from collections import OrderedDict

import pytest
import pycldf
import openpyxl as op

from lexedata.importer.excel_matrix import load_dataset
from lexedata.importer.excel_interleaved import import_interleaved
from lexedata.importer.excel_long_format import add_single_languages

from lexedata.exporter.edictor import forms_to_tsv

from lexedata.edit.add_segments import add_segments_to_dataset
from lexedata.edit.add_concepticon import create_concepticon_for_concepts
from lexedata.edit.detect_cognates import cognate_code_to_file
from lexedata.edit.add_singleton_cognatesets import create_singletons

from lexedata.report.homophones import list_homophones
from lexedata.report.coverage import coverage_report, Missing

from lexedata import util, cli
from lexedata.types import WorldSet
from lexedata.util.fs import copy_dataset
from helper_functions import copy_metadata, copy_to_temp


# Test the importers
@pytest.mark.skip()
def test_matrix_import_skips_question_marks():
    data = [  # noqa
        ["Concept", "L1", "L2"],
        ["C1", "?", "Form1"],
        ["C2", "Form2", "Form3"],
    ]
    # create excel with data
    wb = op.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    dirname = Path(tempfile.mkdtemp(prefix="lexedata-test"))
    target = dirname / "test.xlsx"
    wb.save(target)

    # TODO: struggeling to make a matrix importer run on such a simple structure in a way whitout too much coding ... this might even be a bad sign?
    # create simple metadata
    dataset = pycldf.Wordlist.in_dir(dirname)
    dataset.write(FormTable=[{"Concept": "", "L1": "", "L2": "", "value": ""}])  # noqa
    # metadata = dataset.tablegroup._fname
    metadata = Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    # load the dataset
    load_dataset(
        metadata=metadata,
        lexicon=target,
    )
    dataset = pycldf.Dataset.from_metadata(metadata)
    forms = {f for f in dataset["FormTable"]}
    print(forms)
    assert True
    # TODO: Put data into an Excel sheet
    # TODO: Run a matrix importer
    # TODO: Check that the resulting dataset has no entry for C1 in L1


def test_interleaved_import_skips_na():
    data = [
        ["", "Duala", "Ntomba"],
        ["all", "ɓɛ́sɛ̃(nk)", "umá"],
        ["", "1", "?"],
        ["arm", "?", "lobɔ́kɔ"],
        ["", "7", "1"],
        ["missing", "?", "?"],
        ["", "?"],  # Yes, this is purposefully a line with only two entries.
    ]

    # create excel with data
    wb = op.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    # import excel
    forms = [tuple(r) for r in import_interleaved(ws)]

    assert forms == [
        ("duala_all", "Duala", "all", "ɓɛ́sɛ̃(nk)", None, "1"),
        ("duala_arm", "Duala", "arm", None, None, "7"),
        ("ntomba_all", "Ntomba", "all", "umá", None, None),
        ("ntomba_arm", "Ntomba", "arm", "lobɔ́kɔ", None, "1"),
    ]


def test_single_excel_import_skips_na():
    data = [  # noqa
        ["phonetic", "Form", "English"],
        ["aa", "e.ta.'kɾã", "one"],
        ["bb", "mĩ.'ɾõ1", "two"],
        ["?", "?", "one"],
        ["cc", "?", "three"],
    ]
    # create excel with data
    wb = op.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    sheets = [sheet for sheet in wb]
    metadata = copy_metadata(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    _ = add_single_languages(
        metadata=metadata,
        sheets=sheets,
        match_form=[],
        concept_name="English",
        language_name=None,
        ignore_superfluous=True,
        ignore_missing=True,
        status_update=None,
        logger=cli.logger,
    )
    dataset = pycldf.Dataset.from_metadata(metadata)
    forms = [f for f in dataset["FormTable"]]
    assert forms == [
        OrderedDict(
            [
                ("ID", "sheet_one"),
                ("Language_ID", "Sheet"),
                ("Concept_ID", "one"),
                ("Form", "e.ta.'kɾã"),
                ("Segments", []),
                ("Value", "aa\te.ta.'kɾã\tone"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
        OrderedDict(
            [
                ("ID", "sheet_two"),
                ("Language_ID", "Sheet"),
                ("Concept_ID", "two"),
                ("Form", "mĩ.'ɾõ1"),
                ("Segments", []),
                ("Value", "bb\tmĩ.'ɾõ1\ttwo"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
    ]


@pytest.mark.skip()
def test_matrix_importer_dash():
    pass


def test_interleaved_import_dash():
    data = [
        ["", "Duala", "Ntomba"],
        ["all", "ɓɛ́sɛ̃(nk)", "umá"],
        ["", "1", "-"],
        ["arm", "-", "lobɔ́kɔ"],
        ["", "7", "1"],
    ]

    # create excel with data
    wb = op.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    # import excel
    forms = [tuple(r) for r in import_interleaved(ws)]
    expected_forms = [
        ("duala_all", "Duala", "all", "ɓɛ́sɛ̃(nk)", None, "1"),
        ("duala_arm", "Duala", "arm", "-", None, "7"),
        ("ntomba_all", "Ntomba", "all", "umá", None, "-"),
        ("ntomba_arm", "Ntomba", "arm", "lobɔ́kɔ", None, "1"),
    ]
    assert forms == expected_forms


def test_single_excel_import_dash():
    data = [  # noqa
        ["phonetic", "Form", "English"],
        ["aa", "e.ta.'kɾã", "one"],
        ["bb", "-", "two"],
    ]
    # create excel with data
    wb = op.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    sheets = [sheet for sheet in wb]
    metadata = copy_metadata(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    _ = add_single_languages(
        metadata=metadata,
        sheets=sheets,
        match_form=[],
        concept_name="English",
        language_name=None,
        ignore_superfluous=True,
        ignore_missing=True,
        status_update=None,
        logger=cli.logger,
    )
    dataset = pycldf.Dataset.from_metadata(metadata)
    forms = [f for f in dataset["FormTable"]]
    assert forms == [
        OrderedDict(
            [
                ("ID", "sheet_one"),
                ("Language_ID", "Sheet"),
                ("Concept_ID", "one"),
                ("Form", "e.ta.'kɾã"),
                ("Segments", []),
                ("Value", "aa\te.ta.'kɾã\tone"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
        OrderedDict(
            [
                ("ID", "sheet_two"),
                ("Language_ID", "Sheet"),
                ("Concept_ID", "two"),
                ("Form", "-"),
                ("Segments", []),
                ("Value", "bb\t-\ttwo"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
    ]


# Test exporters
# TODO: @Gereon Could you take care of the phylogenetics exporter?
def test_phylogenetics_exporter_dash_is_absence():
    forms = [  # noqa
        {"ID": "L1C1", "Language_ID": "L1", "Concept_ID": "C1", "Form": "L1C1"},
        {"ID": "L2C1", "Language_ID": "L2", "Concept_ID": "C1", "Form": "L2C1"},
        {"ID": "L1C2", "Language_ID": "L1", "Concept_ID": "C2", "Form": "L1C2"},
        {"ID": "L2C2", "Language_ID": "L2", "Concept_ID": "C2", "Form": "-"},
    ]
    cognates = [  # noqa
        {"Form_ID": "L1C1", "Cognateset_ID": "1"},
        {"Form_ID": "L2C1", "Cognateset_ID": "1"},
        {"Form_ID": "L1C2", "Cognateset_ID": "2"},
    ]
    # TODO: Run the phylogenetics exporter
    assert {"L1": "011", "L2": "010"}  # TODO == alignment


# TODO: @Gereon Could you take care of the phylogenetics exporter?
def test_phylogenetics_exporter_unknown():
    forms = [  # noqa
        {"ID": "L1C1", "Language_ID": "L1", "Concept_ID": "C1", "Form": "L1C1"},
        {"ID": "L2C1", "Language_ID": "L2", "Concept_ID": "C1", "Form": "L2C1"},
        {"ID": "L1C2", "Language_ID": "L1", "Concept_ID": "C2", "Form": "L1C2"},
        {"ID": "L2C2", "Language_ID": "L2", "Concept_ID": "C2", "Form": ""},
        {"ID": "L2C3", "Language_ID": "L2", "Concept_ID": "C3", "Form": "L1C3"},
    ]
    cognates = [  # noqa
        {"Form_ID": "L1C1", "Cognateset_ID": "1"},
        {"Form_ID": "L2C1", "Cognateset_ID": "1"},
        {"Form_ID": "L1C2", "Cognateset_ID": "2"},
        {"Form_ID": "L1C3", "Cognateset_ID": "3"},
    ]
    # TODO: Run the phylogenetics exporter
    assert {"L1": "011?", "L2": "01?1"}  # TODO == alignment


def test_edictor_exporter_no_na_forms():
    forms = [  # noqa
        {
            "ID": "L1C1",
            "Language_ID": "L1",
            "Parameter_ID": ["C1"],
            "Form": "",
            "Value": " ",
        },
        {
            "ID": "L2C1",
            "Language_ID": "L2",
            "Parameter_ID": ["C1"],
            "Form": "L2C1",
            "Value": " ",
        },
        {
            "ID": "L1C2",
            "Language_ID": "L1",
            "Parameter_ID": ["C2"],
            "Form": "L1C2",
            "Value": " ",
        },
        {
            "ID": "L2C2",
            "Language_ID": "L2",
            "Parameter_ID": ["C2"],
            "Form": "-",
            "Value": " ",
        },
    ]
    cognates = [  # noqa
        {"Form_ID": "L2C1", "Cognateset_ID": "1"},
        {"Form_ID": "L1C2", "Cognateset_ID": "2"},
    ]
    dirname = Path(tempfile.mkdtemp(prefix="lexedata-test"))
    target = dirname / "cldf-metadata.json"
    dataset = copy_dataset(
        original=Path(__file__).parent
        / "data/cldf/smallmawetiguarani/cldf-metadata.json",
        target=target,
    )

    dataset.write(FormTable=forms)
    forms, judgements_about_form, cognateset_mapping = forms_to_tsv(
        dataset=dataset, languages=WorldSet(), concepts=WorldSet(), cognatesets=cognates
    )
    assert forms == {
        "L2C1": {
            "id": "L2C1",
            "languageReference": "L2",
            "parameterReference": "C1",
            "form": "L2C1",
            "orthographic": None,
            "phonemic": None,
            "phonetic": None,
            "variants": "",
            "segments": [],
            "comment": None,
            "procedural_comment": None,
            "value": " ",
            "source": "",
            "_parameterReference": "C1",
        },
        "L1C2": {
            "id": "L1C2",
            "languageReference": "L1",
            "parameterReference": "C2",
            "form": "L1C2",
            "orthographic": None,
            "phonemic": None,
            "phonetic": None,
            "variants": "",
            "segments": [],
            "comment": None,
            "procedural_comment": None,
            "value": " ",
            "source": "",
            "_parameterReference": "C2",
        },
    }


# Test other scripts
# TODO: detect cognates creates a bunch of .tsv files. Is that supposed to be the case?
# TODO: I'm unsure about the created cognatesets.csv. Could you, Gereon, check that it is alright? SO far no changes performed on detect_cognates to block na forms...
@pytest.mark.skip()
def test_detect_cognates_ignores_na_forms():
    forms = [  # noqa
        {
            "ID": "L2C1",
            "Language_ID": "L2",
            "Concept_ID": "C1",
            "Form": "L2C1",
            "Value": "L2C1",
        },
        {
            "ID": "L1C1",
            "Language_ID": "L1",
            "Concept_ID": "C1",
            "Form": "",
            "Value": "?",
        },
        {
            "ID": "L1C2",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "L1C2",
            "Value": "L1C2",
        },
        {
            "ID": "L2C2",
            "Language_ID": "L2",
            "Concept_ID": "C2",
            "Form": "-",
            "Value": "-",
        },
    ]
    cognates = [  # noqa
        {"ID": "1", "Form_ID": "L2C1", "Cognateset": "1"},
        {"ID": "2", "Form_ID": "L1C2", "Cognateset": "2"},
    ]
    languages = [{"ID": "L1", "Name": "L1"}, {"ID": "L2", "Name": "L2"}]
    concepts = [{"ID": "C1", "Name": "C1"}, {"ID": "C2", "Name": "C2"}]
    # load dataset and write content and segment
    dataset = pycldf.Dataset.from_metadata(
        copy_metadata(Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json")
    )
    dataset.write(
        FormTable=forms,
        CognateTable=cognates,
        LanguageTable=languages,
        ParameterTable=concepts,
    )
    _ = add_segments_to_dataset(
        dataset=dataset,
        transcription="Form",
        overwrite_existing=False,
        replace_form=False,
    )
    # call detect cognates
    cognate_code_to_file(
        metadata=dataset.tablegroup._fname,
        ratio=1.5,
        soundclass="sca",
        cluster_method="infomap",
        gop=-2,
        mode="overlap",
        threshold=0.55,
        initial_threshold=0.7,
        output_file=dataset.tablegroup._fname.parent / "output",
    )
    # TODO: run ACD
    # TODO: Check that the cognatesets contain only L2C1 and L1C2
    cogsets = [c for c in dataset["CognatesetTable"]]
    print(cogsets)
    assert False


def test_add_segments_skips_na_forms():
    forms = [
        {
            "ID": "L2C2",
            "Language_ID": "L2",
            "Concept_ID": "C2",
            "Form": "-",
            "Value": "-",
        },
        {
            "ID": "L1C1",
            "Language_ID": "L1",
            "Concept_ID": "C1",
            "Form": "",
            "Value": "?",
        },
        {
            "ID": "L2C1",
            "Language_ID": "L2",
            "Concept_ID": "C1",
            "Form": "L2C1",
            "Value": "L2C1",
        },
        {
            "ID": "L1C2",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "L1C2",
            "Value": "L1C2",
        },
    ]
    dataset = pycldf.Dataset.from_metadata(
        copy_metadata(Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json")
    )
    dataset.write(FormTable=forms)
    _ = add_segments_to_dataset(
        dataset=dataset,
        transcription="Form",
        overwrite_existing=False,
        replace_form=False,
    )
    segmented_forms = [f for f in dataset["FormTable"]]
    assert segmented_forms == [
        OrderedDict(
            [
                ("ID", "L2C2"),
                ("Language_ID", "L2"),
                ("Concept_ID", "C2"),
                ("Form", "-"),
                ("Segments", []),
                ("Value", "-"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
        OrderedDict(
            [
                ("ID", "L1C1"),
                ("Language_ID", "L1"),
                ("Concept_ID", "C1"),
                ("Form", None),
                ("Segments", []),
                ("Value", "?"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
        OrderedDict(
            [
                ("ID", "L2C1"),
                ("Language_ID", "L2"),
                ("Concept_ID", "C1"),
                ("Form", "L2C1"),
                ("Segments", ["L", "²", "C", "¹"]),
                ("Value", "L2C1"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
        OrderedDict(
            [
                ("ID", "L1C2"),
                ("Language_ID", "L1"),
                ("Concept_ID", "C2"),
                ("Form", "L1C2"),
                ("Segments", ["L", "¹", "C", "²"]),
                ("Value", "L1C2"),
                ("Comment", None),
                ("Source", []),
            ]
        ),
    ]


def test_add_singlestons():
    forms = [  # noqa
        {
            "ID": "L2C1",
            "Language_ID": "L2",
            "Concept_ID": "C1",
            "Form": "L2C1",
            "Value": "L2C1",
            "Segments": ["f"],
        },
        {
            "ID": "L1C1",
            "Language_ID": "L1",
            "Concept_ID": "C1",
            "Form": "",
            "Value": "?",
            "Segments": [],
        },
        {
            "ID": "L1C2",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "L1C2",
            "Value": "L1C2",
            "Segments": ["f"],
        },
        {
            "ID": "L2C2",
            "Language_ID": "L2",
            "Concept_ID": "C2",
            "Form": "-",
            "Value": "-",
            "Segments": [],
        },
    ]
    cognates = [{"ID": "1", "Form_ID": "L2C1", "Cognateset": "1"}]
    concepts = [{"ID": "C1", "Name": "C1"}, {"ID": "C2", "Name": "C1"}]
    cogsets = [{"ID": "1", "Name": "1"}]
    dataset = pycldf.Dataset.from_metadata(
        copy_metadata(Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json")
    )
    dataset.write(
        FormTable=forms,
        ParameterTable=concepts,
        CognateTable=cognates,
        CognatesetTable=cogsets,
    )
    all_cogsets, judgements = create_singletons(dataset=dataset)
    assert [dict(x) for x in all_cogsets] == [
        {"ID": "1", "Name": "1", "Comment": None},
        {"ID": "X_L1C2_1", "Name": "C2", "Comment": None},
    ]
    assert [dict(j) for j in judgements] == [
        {
            "ID": "1",
            "Form_ID": "L2C1",
            "Cognateset": "1",
            "Segment_Slice": [],
            "Alignment": [],
            "Comment": None,
        },
        {
            "ID": "X_L1C2_1",
            "Form_ID": "L1C2",
            "Cognateset": "X_L1C2_1",
            "Segment_Slice": ["1:1"],
            "Alignment": ["f"],
            "Comment": None,
        },
    ]


def test_homohpones_skips_na_forms(capsys):
    forms = [  # noqa
        {
            "ID": "L1C1",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "form",
            "Value": "-",
        },
        {
            "ID": "L1C2",
            "Language_ID": "L1",
            "Concept_ID": "C1",
            "Form": "",
            "Value": "-",
        },
        {
            "ID": "L1C3",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "form",
            "Value": "-",
        },
        {
            "ID": "L1C4",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "",
            "Value": "-",
        },
        {
            "ID": "L1C5",
            "Language_ID": "L1",
            "Concept_ID": "C2",
            "Form": "-",
            "Value": "-",
        },
        {
            "ID": "L1C6",
            "Language_ID": "L2",
            "Concept_ID": "C2",
            "Form": "-",
            "Value": "-",
        },
    ]
    dataset, target = copy_to_temp(
        Path(__file__).parent / r"data/cldf/minimal/cldf-metadata.json"
    )
    target = target.parent
    dataset.write(
        FormTable=forms,
        ParameterTable=[{"ID": "C1", "Name": "one"}, {"ID": "C2", "Name": "two"}],
    )
    # add concepticon reference
    create_concepticon_for_concepts(
        dataset=dataset,
        language=[],
        concepticon_glosses=False,
        concepticon_definition=False,
        overwrite=False,
        status_update=None,
    )
    output = io.StringIO()
    list_homophones(dataset=dataset, out=output)
    assert set(output.getvalue().split("\n")) == {
        "L1, 'form': Unknown (but at least one concept not found):",
        "\t L1C1 (C2)",
        "\t L1C3 (C2)",
        "",
    }


def test_extended_cldf_validate():
    # TODO: What tests should this run? What kind of data should fail?
    # - A dataset where an NA form is in cognatesets should fail
    # - A dataset where the segments is "" and the form is "-" should pass
    ...


def test_coverage_report_missing_and_na_default():
    ds = util.fs.new_wordlist(
        FormTable=[
            {"ID": "f1", "Language_ID": "l1", "Parameter_ID": "c1", "Form": ""},
            {"ID": "f2", "Language_ID": "l1", "Parameter_ID": "c2", "Form": "-"},
            {"ID": "f3", "Language_ID": "l1", "Parameter_ID": "c3", "Form": "form"},
            {"ID": "f4", "Language_ID": "l2", "Parameter_ID": "c1", "Form": "form"},
            {"ID": "f5", "Language_ID": "l2", "Parameter_ID": "c2", "Form": "form"},
            {"ID": "f6", "Language_ID": "l2", "Parameter_ID": "c3", "Form": "form"},
        ]
    )
    ds["FormTable", "Form"].required = False
    data = coverage_report(ds, only_coded=False)
    assert data == [["l1", "l1", 2, 2.0 / 3.0, 1.0], ["l2", "l2", 3, 1.0, 1.0]]


def test_coverage_report_missing_and_na_ignore():
    ds = util.fs.new_wordlist(
        FormTable=[
            {"ID": "f1", "Language_ID": "l1", "Parameter_ID": "c1", "Form": ""},
            {"ID": "f2", "Language_ID": "l1", "Parameter_ID": "c2", "Form": "-"},
            {"ID": "f3", "Language_ID": "l1", "Parameter_ID": "c3", "Form": "form"},
            {"ID": "f4", "Language_ID": "l2", "Parameter_ID": "c1", "Form": "form"},
            {"ID": "f5", "Language_ID": "l2", "Parameter_ID": "c2", "Form": "form"},
            {"ID": "f6", "Language_ID": "l2", "Parameter_ID": "c3", "Form": "form"},
        ]
    )
    ds["FormTable", "Form"].required = False
    data = coverage_report(ds, only_coded=False, missing=Missing.IGNORE)
    assert data == [["l1", "l1", 1, 1.0 / 3.0, 1.0], ["l2", "l2", 3, 1.0, 1.0]]
