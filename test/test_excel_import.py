import pytest
from pathlib import Path
import argparse
import re

import pycldf
import openpyxl

from helper_functions import copy_metadata, copy_to_temp
import lexedata.importer.excel_matrix as f


@pytest.fixture
def empty_excel():
    return Path(__file__).parent / "data/excel/empty_excel.xlsx"


# TODO: have a look at this test. just trying to pass codecov
def test_db_chache():
    copy = copy_metadata(Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json")
    res = dict()
    dataset = pycldf.Dataset.from_metadata(copy)
    db = f.DB(output_dataset=dataset)
    db.cache_dataset()
    for table in dataset.tables:
        table_type = (
            table.common_props.get("dc:conformsTo", "").rsplit("#", 1)[1] or table.url
        )
        res[table_type] = {}
    assert db.cache == res


def test_no_wordlist_and_no_cogsets(tmp_path):
    # mock empty json file
    path = tmp_path / "invented_path"
    path.open("w").write("{}")
    with pytest.raises(
        argparse.ArgumentError,
        match="At least one of WORDLIST and COGNATESETS excel files must be specified.*",
    ):
        f.load_dataset(
            metadata=path,
            lexicon=None,
            cognate_lexicon=None,
        )


def test_no_dialect_excel_parser(tmp_path, caplog, empty_excel):
    # ExcelParser
    path = tmp_path / "invented_path"
    path.open("w").write("{}")
    with pytest.raises(ValueError):
        # mock empty json file
        f.load_dataset(
            metadata=path,
            lexicon=empty_excel,
        )
    assert re.search("User-defined format specification .* missing", caplog.text)
    assert re.search("default parser", caplog.text)


def test_no_dialect_excel_cognate_parser(tmp_path, caplog, empty_excel):
    # ExcelCognateParser
    path = tmp_path / "invented_path"
    path.open("w").write("{}")
    with pytest.raises(ValueError):
        # mock empty json file
        f.load_dataset(metadata=path, lexicon=None, cognate_lexicon=empty_excel)
    assert re.search("User-defined format specification .* missing", caplog.text)
    assert re.search("default parser", caplog.text)


def test_dialect_missing_key_excel_parser(tmp_path, caplog, empty_excel):
    # ExcelParser
    path = tmp_path / "invented_path"
    path.open("w").write("""{"special:fromexcel": {}}""")
    with pytest.raises(ValueError):
        f.load_dataset(path, lexicon=empty_excel)
    assert re.search(
        "User-defined format specification in the json-file was missing the key lang_cell_regexes, "
        "falling back to default parser",
        caplog.text,
    )


def test_dialect_missing_key_excel_cognate_parser(tmp_path, caplog, empty_excel):
    path = tmp_path / "invented_path"
    path.open("w").write("""{"special:fromexcel": {}}""")
    # CognateExcelParser
    with pytest.raises(ValueError):
        f.load_dataset(path, lexicon=None, cognate_lexicon=empty_excel)
    assert re.search(
        r"User-defined format specification in the json-file was missing the key .*falling back to default parser.*",
        caplog.text,
    )


def test_no_first_row_in_excel(empty_excel):
    original = Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    copy = copy_metadata(original=original)
    with pytest.raises(
        AssertionError,
        match="Your first data row didn't have a name. Please check your format specification or ensure the "
        "first row has a name.",
    ):
        f.load_dataset(metadata=copy, lexicon=empty_excel)


def test_language_regex_error():
    excel = Path(__file__).parent / "data/excel/small_defective_no_regexes.xlsx"
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.lang_cell_regexes = [r"(?P<Name>\[.*)", "(?P<Curator>.*)"]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)

    with pytest.raises(
        ValueError,
        match=r"In cell G1: Expected to encounter match for .*, but found no_language",
    ):
        EP.parse_cells(lexicon_wb)


def test_language_comment_regex_error():
    excel = Path(__file__).parent / "data/excel/small_defective_no_regexes.xlsx"
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.lang_comment_regexes = [r"(\[.*)"]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)
    with pytest.raises(
        ValueError,
        match="In cell G1: Expected to encounter match for .*, but found no_lan_comment.*",
    ):
        EP.parse_cells(lexicon_wb)


def test_properties_regex_error():
    excel = Path(__file__).parent / "data/excel/small_defective_no_regexes.xlsx"
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.row_cell_regexes = [
        "(?P<set>.*)",
        # wrong regex
        r"(?P<Name>\[.*)",
        "(?P<English>.*)",
        "(?P<Spanish>.*)",
        "(?P<Portuguese>.*)",
        "(?P<French>.*)",
    ]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)

    with pytest.raises(
        ValueError,
        match=r"In cell B3: Expected to encounter match for .*, but found no_concept_name",
    ):
        EP.parse_cells(lexicon_wb)


def test_properties_comment_regex_error():
    excel = Path(__file__).parent / "data/excel/small_defective_no_regexes.xlsx"
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.row_comment_regexes = [".*", r"\[.*", ".*", ".*", ".*", ".*"]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)
    with pytest.raises(
        ValueError,
        match=r"In cell B3: Expected to encounter match for .*, but found no_concept_comment",
    ):
        EP.parse_cells(lexicon_wb)


def test_cognate_parser_language_not_found():
    excel = Path(__file__).parent / "data/excel/minimal_cog.xlsx"
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    dataset, copy = copy_to_temp(original)
    lexicon_wb = openpyxl.load_workbook(excel).active
    EP = f.ExcelCognateParser(output_dataset=dataset)
    print(dataset["LanguageTable"])
    with pytest.raises(
        ValueError,
        match="Failed to find object {'ID': 'autaa', 'Name': 'Autaa', 'Comment': "
        "'fictitious!'} in the database. In cell: D1",
    ):
        EP.db.cache_dataset()
        EP.parse_cells(lexicon_wb)
