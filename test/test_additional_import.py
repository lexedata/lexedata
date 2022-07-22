import logging
import re
from pathlib import Path

import openpyxl
import pycldf
import pytest

from lexedata.cli import logger
from lexedata.importer.excel_long_format import (
    ImportLanguageReport,
    add_single_languages,
    read_single_excel_sheet,
)

from helper_functions import copy_metadata, copy_to_temp_no_bib
from mock_excel import MockSingleExcelSheet


@pytest.fixture(
    params=[
        (
            "data/cldf/smallmawetiguarani/cldf-metadata.json",
            "data/excel/test_single_excel_maweti.xlsx",
            "English",
        )
    ]
)
def single_import_parameters(request):
    original = Path(__file__).parent / request.param[0]
    dataset, target = copy_to_temp_no_bib(original)
    excel = Path(__file__).parent / request.param[1]
    concept_name = request.param[2]
    return dataset, target, excel, concept_name


def test_concept_file_not_found(caplog):
    copy = copy_metadata(Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json")
    add_single_languages(
        dataset=pycldf.Dataset.from_metadata(copy),
        sheets=[],
        match_form=None,
        concept_name=None,
        language_name=None,
        ignore_missing=True,
        ignore_superfluous=True,
        status_update=None,
        logger=logger,
    )
    assert re.search(
        r"Did not find concepts\.csv\. Importing all forms independent of concept",
        caplog.text,
    )


def test_import_multi_languages_in_one_sheet_by_id(caplog, single_import_parameters):
    dataset, original, excel, concept_name = single_import_parameters
    excel = MockSingleExcelSheet(
        [
            [
                "Language_ID",
                "orthographic",
                "phonemic",
                "phonetic",
                "Form",
                "English",
                "Comment",
                "procedural_comment",
                "Source",
                "Segments",
                "variants",
            ],
            [
                "ache",
                "",
                "",
                "e.ta.'kɾã",
                "e.ta.'kɾã",
                "one",
                "",
                "",
                "",
                "",
                "",
            ],
            [
                "ache",
                "",
                "",
                "mĩ.'ɾõ1",
                "mĩ.'ɾõ1",
                "polysemy_concept",
                "",
                "",
                "",
                "",
                "",
            ],
            [
                "ache",
                "dáhui",
                "",
                "dáhui",
                "dáhui",
                "one@",
                "",
                "",
                "Meléndez2011",
                "",
                "",
            ],
            [
                "ache",
                "salíiri",
                "",
                "salí:ɻi",
                "salí:ɻi",
                "ADULT",
                "",
                "",
                "Meléndez2011",
                "",
                "",
            ],
            [
                "ache",
                "káaɻu",
                "",
                "ká:ɻu",
                "ká:ɻu",
                "AFRAID",
                "",
                "ZJO: /ka:ru/ 'miedo' (Melendez2011)",
                "Ramirez2001a",
                "",
                "",
            ],
            [
                "ache",
                "bánio",
                "",
                "bánio",
                "bánio",
                "again",
                "",
                "",
                "Meléndez2011",
                "",
                "",
            ],
            [
                "kaiwa",
                "teʔiowaa",
                "",
                "",
                "teʔiowaa",
                "five",
                "",
                "",
                "",
                "",
                "[teʔiowaa]",
            ],
            [
                "Canamari",
                "nusu-chüa",
                "",
                "",
                "nusu-c͡çɨa",
                "ANKLE",
                "",
                "",
                "Martius1867",
                "",
                "",
            ],
            [
                "Canamari",
                "nutzûma",
                "",
                "",
                "nut͡suma",
                "ANUS",
                "",
                "",
                "Martius1867",
                "",
                "",
            ],
            [
                "Canamari",
                "nutanachy",
                "",
                "",
                "nutanac͡çi",
                "ARMPIT",
                "",
                "",
                "Martius1867",
                "",
                "",
            ],
        ]
    )
    dataset.write(
        ParameterTable=list(dataset["ParameterTable"])
        + [
            {"ID": x, "Name": x}
            for x in [
                "polysemy_concept",
                "ADULT",
                "AFRAID",
                "again",
                "ANKLE",
                "ANUS",
                "ARMPIT",
            ]
        ]
    )
    report = add_single_languages(
        dataset=dataset,
        sheets=[excel],
        match_form=None,
        concept_name="English",
        language_name=None,
        ignore_missing=True,
        ignore_superfluous=True,
        status_update=None,
        logger=logger,
    )
    assert report == {
        "ache": ImportLanguageReport(
            is_new_language=False, new=4, existing=1, skipped=0, concepts=1
        ),
        "kaiwa": ImportLanguageReport(
            is_new_language=False, new=0, existing=1, skipped=0, concepts=0
        ),
        "Canamari": ImportLanguageReport(
            is_new_language=True, new=3, existing=0, skipped=0, concepts=0
        ),
    }


def test_import_new_language_warn_once(caplog, single_import_parameters):
    dataset, original, excel, concept_name = single_import_parameters
    excel = MockSingleExcelSheet(
        [
            [
                "Language",
                "orthographic",
                "phonemic",
                "phonetic",
                "Form",
                "English",
                "Comment",
                "procedural_comment",
                "Source",
                "Segments",
                "variants",
            ],
            [
                "A",
                "",
                "",
                "e",
                "e",
                "one",
                "",
                "",
                "",
                "",
                "",
            ],
            [
                "A",
                "",
                "",
                "m",
                "m",
                "two",
                "",
                "",
                "",
                "",
                "",
            ],
        ]
    )
    with caplog.at_level(logging.WARNING):
        add_single_languages(
            dataset=dataset,
            sheets=[excel],
            match_form=None,
            concept_name="English",
            language_name="Language",
            ignore_missing=True,
            ignore_superfluous=True,
            status_update=None,
            logger=logger,
        )
    assert caplog.text.count("new language A,") == 1


def test_import_multi_languages_in_one_sheet_by_name(caplog, single_import_parameters):
    dataset, original, excel, concept_name = single_import_parameters
    excel = MockSingleExcelSheet(
        [
            [
                "Language",
                "orthographic",
                "phonemic",
                "phonetic",
                "Form",
                "English",
                "Comment",
                "procedural_comment",
                "Source",
                "Segments",
                "variants",
            ],
            [
                "Aché",
                "",
                "",
                "e.ta.'kɾã",
                "e.ta.'kɾã",
                "one",
                "",
                "",
                "",
                "",
                "",
            ],
            [
                "Aché",
                "",
                "",
                "mĩ.'ɾõ1",
                "mĩ.'ɾõ1",
                "polysemy_concept",
                "",
                "",
                "",
                "",
                "",
            ],
            [
                "Aché",
                "dáhui",
                "",
                "dáhui",
                "dáhui",
                "one@",
                "",
                "",
                "Meléndez2011",
                "",
                "",
            ],
            [
                "Aché",
                "salíiri",
                "",
                "salí:ɻi",
                "salí:ɻi",
                "ADULT",
                "",
                "",
                "Meléndez2011",
                "",
                "",
            ],
            [
                "Aché",
                "káaɻu",
                "",
                "ká:ɻu",
                "ká:ɻu",
                "AFRAID",
                "",
                "ZJO: /ka:ru/ 'miedo' (Melendez2011)",
                "Ramirez2001a",
                "",
                "",
            ],
            [
                "Aché",
                "bánio",
                "",
                "bánio",
                "bánio",
                "again",
                "",
                "",
                "Meléndez2011",
                "",
                "",
            ],
            [
                "Kaiwá",
                "teʔiowaa",
                "",
                "",
                "teʔiowaa",
                "five",
                "",
                "",
                "",
                "",
                "[teʔiowaa]",
            ],
            [
                "Canamari",
                "nusu-chüa",
                "",
                "",
                "nusu-c͡çɨa",
                "ANKLE",
                "",
                "",
                "Martius1867",
                "",
                "",
            ],
            [
                "Canamari",
                "nutzûma",
                "",
                "",
                "nut͡suma",
                "ANUS",
                "",
                "",
                "Martius1867",
                "",
                "",
            ],
            [
                "Canamari",
                "nutanachy",
                "",
                "",
                "nutanac͡çi",
                "ARMPIT",
                "",
                "",
                "Martius1867",
                "",
                "",
            ],
        ]
    )
    dataset.write(
        ParameterTable=list(dataset["ParameterTable"])
        + [
            {"ID": x, "Name": x}
            for x in [
                "polysemy_concept",
                "ADULT",
                "AFRAID",
                "again",
                "ANKLE",
                "ANUS",
                "ARMPIT",
            ]
        ]
    )
    report = add_single_languages(
        dataset=dataset,
        sheets=[excel],
        match_form=None,
        concept_name="English",
        language_name="Language",
        ignore_missing=True,
        ignore_superfluous=True,
        status_update=None,
        logger=logger,
    )
    assert report == {
        "ache": ImportLanguageReport(
            is_new_language=False, new=4, existing=1, skipped=0, concepts=1
        ),
        "kaiwa": ImportLanguageReport(
            is_new_language=False, new=0, existing=1, skipped=0, concepts=0
        ),
        "Canamari": ImportLanguageReport(
            is_new_language=True, new=3, existing=0, skipped=0, concepts=0
        ),
    }


def test_multi_sheet_import(single_import_parameters, caplog):
    dataset, original, excel, concept_name = single_import_parameters
    excel = openpyxl.load_workbook(excel)
    dataset.write(
        ParameterTable=list(dataset["ParameterTable"])
        + [
            {"ID": x, "Name": x}
            for x in [
                "polysemy_concept",
                "ADULT",
                "AFRAID",
                "again",
                "ANKLE",
                "ANUS",
                "ARMPIT",
            ]
        ]
    )
    report = add_single_languages(
        dataset=dataset,
        sheets=[excel[sheet] for sheet in excel.sheetnames],
        match_form=None,
        concept_name="English",
        language_name=None,
        ignore_missing=True,
        ignore_superfluous=True,
        status_update=None,
        logger=logger,
    )
    assert report == {
        "ache": ImportLanguageReport(
            is_new_language=False, new=4, existing=1, skipped=0, concepts=1
        ),
        "Canamari": ImportLanguageReport(
            is_new_language=True, new=3, existing=0, skipped=0, concepts=0
        ),
    }
    assert [dict(f) for f in dataset["FormTable"]][-1] == {
        "ID": "canamari_armpit",
        "Language_ID": "Canamari",
        "Parameter_ID": ["ARMPIT"],
        "Form": "nutanac͡çi",
        "orthographic": "nutanachy",
        "phonemic": None,
        "phonetic": None,
        "variants": [],
        "Segments": [],
        "Comment": None,
        "procedural_comment": None,
        "Value": "nutanachy\t\t\tnutanac͡çi\tARMPIT\t\t\tMartius1867\t\t",
        "Source": ["Martius1867"],
    }


def test_add_new_forms_maweti(single_import_parameters):
    dataset, original, excel, concept_name = single_import_parameters
    excel = openpyxl.load_workbook(excel)
    c_f_id = dataset["FormTable", "id"].name
    c_f_concept = dataset["FormTable", "parameterReference"].name
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    old_form_ids = {row[c_f_id] for row in dataset["FormTable"]}
    sheet = [sheet for sheet in excel.sheetnames]
    for sheet in sheet:
        read_single_excel_sheet(
            dataset=dataset,
            sheet=excel[sheet],
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )
    new_form_ids = {row[c_f_id] for row in dataset["FormTable"]}
    new_forms = new_form_ids - old_form_ids
    assert len(new_forms) == 1
    for f in new_forms:
        assert f.startswith("ache_one")
    (new_form,) = [row for row in dataset["FormTable"] if row[c_f_id] in new_forms]
    assert new_form[c_f_concept] == ["one_1"]


def test_import_error_missing_parameter_column(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "variants",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
            ],
            [],
        ]
    )
    with pytest.raises(
        AssertionError, match=f"Could not find concept column {concept_name} in .*"
    ):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )


def test_missing_columns1(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "variants",
                "Form",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
            ],
            [],
        ]
    )
    with pytest.raises(
        ValueError,
        match=".*sheet MockSingleExcelSheet.*missing col.*{[^a-z]*orthographic[^a-z]*}.*--ignore-missing-columns",
    ):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )


# TODO: Discuss with Gereon, Test has multiple asserts
def test_missing_columns2(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "variants",
                "Form",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
                "undescribed",
                "superfluous2",
            ],
            [],
        ]
    )
    with pytest.raises(
        ValueError, match=".*sheet MockSingleExcelSheet.*unexpected col.*"
    ) as ex_info:
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column="English",
            ignore_missing=True,
        )
    assert "undescribed" in ex_info.value.args[0]
    assert "superfluous2" in ex_info.value.args[0]
    assert "--ignore-superfluous-columns" in ex_info.value.args[0]


def test_superfluous_columns1(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "variants",
                "Form",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
                "orthographic",
                "superfluous",
            ],
            [],
        ]
    )
    with pytest.raises(
        ValueError,
        match=".* Excel sheet MockSingleExcelSheet contained unexpected columns {'superfluous'}.*"
        ".* use --ignore-superfluous-columns .*",
    ):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column="English",
        )


def test_missing_value_column(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    dataset.remove_columns("FormTable", "Value")
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "English",
                "variants",
                "Form",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
                "orthographic",
                "superfluous",
            ],
            [],
        ]
    )
    # Check that we are warned about a missing value column
    with caplog.at_level(logging.WARNING):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column="English",
            ignore_superfluous=True,
        )
    assert re.search(
        "does not specify .* #value .* Value",
        caplog.text,
    )


def test_missing_concept(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "variants",
                "Form",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
                "orthographic",
                "superfluous",
            ],
            [],
        ]
    )
    concept_column_name = "Concept_Column_Name"
    # AssertionError on concept column not in excel header
    with pytest.raises(AssertionError, match=f".*{concept_column_name}.*"):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_column_name,
            ignore_superfluous=True,
        )


def test_superfluous_columns2(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    concepts = {"Concept": "c_id_1"}
    sheet = MockSingleExcelSheet(
        [
            [
                "Form",
                "Segments",
                "English",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
                "variants",
                "orthographic",
                "superfluous",
            ],
            ["test form", "t e s t", "Concept"],
        ]
    )
    # AssertionError on concept column not in excel header
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column="English",
            ignore_superfluous=True,
        )
    assert re.search(
        r"Excel sheet MockSingleExcelSheet contained unexpected columns {'superfluous'}. These columns will be ignored",
        caplog.text,
    )


def test_no_concept_separator(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    dataset["FormTable", "parameterReference"].separator = None
    dataset.write_metadata()

    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    sheet.title = "new_language"

    # Import this single form in a new language
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=sheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            is_new_language=True, new=1, existing=0, skipped=0, concepts=0
        )
    }

    # Import it again, with a new concept
    sheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "three",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    sheet.title = "new_language"

    # Test new concept was added as new form
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=sheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            is_new_language=True, new=1, existing=0, skipped=0, concepts=0
        )
    }
    # Test messages mention the solutions
    print(caplog.text)
    assert re.search(
        r"not.* polysemous forms.*separator.*FormTable.*parameterReference.*json.*lexedata\.report\.list_homophones",
        caplog.text,
    )


def test_list_missing_concepts(single_import_parameters, capsys):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    del dataset["FormTable", "value"]
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    dataset.write(FormTable=[])
    sheet = MockSingleExcelSheet(
        [
            [
                "Language_ID",
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "ache",
                "missing1",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "missing2",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "missing2",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    missing = set()
    read_single_excel_sheet(
        dataset=dataset,
        sheet=sheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
        missing_concepts=missing,
    )
    assert missing == {"missing1", "missing2"}


def test_list_multi_concepts(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    dataset["FormTable", "parameterReference"].separator = "; "
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    del dataset["FormTable", "value"]
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    dataset.write(FormTable=[])
    sheet = MockSingleExcelSheet(
        [
            [
                "Language_ID",
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "ache",
                "missing1; one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "one; two",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "missing1; missing2",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "one; missing3; two",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    missing = set()
    with caplog.at_level(logging.WARNING):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
            missing_concepts=missing,
        )
    assert missing == {"missing1", "missing2", "missing3"}
    assert re.search(r"concept.*missing1.*not.*found.*links", caplog.text)
    assert re.search(
        r"concept.*missing1[^a-zA-z]*missing2.*not.*found.*skipped", caplog.text
    )
    assert re.search(r"concept.*missing3.*not.*found.*links", caplog.text)


def test_duplicate_forms_no_value(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    del dataset["FormTable", "value"]
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    dataset.write(FormTable=[])
    sheet = MockSingleExcelSheet(
        [
            [
                "Language_ID",
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "ache",
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
            [
                "ache",
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )
    assert "Form form 'one' in Language ache was already in dataset"


def test_language_id(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    dataset.write(FormTable=[])
    sheet = MockSingleExcelSheet(
        [
            [
                "Language_ID",
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "ache",
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )
    assert {"ache"} == {f["Language_ID"] for f in dataset["FormTable"]}


def test_no_language_table(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    dataset.remove_table("LanguageTable")
    dataset.write(FormTable=[])
    sheet = MockSingleExcelSheet(
        [
            [
                "Language_ID",
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "ache",
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )
    assert "no LanguageTable" in caplog.text


def test_language_name(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_f_language = dataset["FormTable", "languageReference"].name
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    dataset.write(FormTable=[])
    sheet = MockSingleExcelSheet(
        [
            [
                "Language",
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "Aché",
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    read_single_excel_sheet(
        dataset=dataset,
        sheet=sheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
        language_name_column="Language",
    )
    assert {"ache"} == {f[c_f_language] for f in dataset["FormTable"]}


def test_concept_separator(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_f_concept = dataset["FormTable", "parameterReference"].name
    match_form = [c_f_concept]
    concepts = dict()
    sheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "phonemic",
                "orthographic",
                "variants",
            ],
            [
                "three",
                "form",
                "f o r m",
                "auto-generated",
                "",
                "source[10]",
                "phonetic",
                "phonemic",
                "orthographic",
                "",
            ],
        ]
    )
    # ValueError on missing column
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            entries_to_concepts=concepts,
            match_form=match_form,
            concept_column="English",
        )
    assert re.search(
        r"[mM]atch.*concept.*lexedata\.report\.list_homophones", caplog.text
    )


def test_concept_not_found(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "FAKE",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    mocksheet.title = "new_language"
    read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    )
    assert re.search(r"concept.*FAKE.*not.*found", caplog.text)


def test_form_exists(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "two",  # existing concept, but new association
                "e.ta.'kɾã",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ],
        ]
    )
    mocksheet.title = "ache"
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=mocksheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )
    assert re.search(
        r"two.*e\.ta\.'kɾã.*was already in dataset",
        caplog.text,
    )


def test_new_concept_association(single_import_parameters, caplog):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "two",  # existing concept, but new association
                "e.ta.'kɾã",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ],
        ]
    )
    mocksheet.title = "ache"
    with caplog.at_level(logging.INFO):
        read_single_excel_sheet(
            dataset=dataset,
            sheet=mocksheet,
            entries_to_concepts=concepts,
            concept_column=concept_name,
        )
    # Test new concept association
    assert re.search(
        r"two.*added to.*ache_one",
        caplog.text,
    )


#############################
# Test report functionality #
#############################


def test_import_report_new_language(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    mocksheet.title = "new_language"
    # Import this single form in a new language
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            is_new_language=True, new=1, existing=0, skipped=0, concepts=0
        )
    }


def test_import_report_existing_form(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    mocksheet.title = "new_language"
    # Import this single form in a new language
    read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    )
    # Import it again, now both form and language should be existing
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            # TODO: Actually, this isn't a new language. The difference between
            # adding forms for a language that is not in the LanguageTable yet,
            # but already has forms in the FormTable, and adding something
            # completely new, is washed out by read_single_language. The
            # interpretation of “Does this language still need to be added to
            # the LanguageTable?” for is_new_language is consistent.
            is_new_language=True,
            new=0,
            existing=1,
            skipped=0,
            concepts=0,
        )
    }


def test_import_report_skipped(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "FAKE",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    mocksheet.title = "new_language"
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            is_new_language=True,
            new=0,
            existing=0,
            skipped=1,
            concepts=0,
        )
    }


# TODO: Multiply asserts
def test_import_report_add_concept(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    sheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    sheet.title = "new_language"

    # Import this single form in a new language
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=sheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            is_new_language=True, new=1, existing=0, skipped=0, concepts=0
        )
    }

    # Import it again, with a new concept
    sheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "three",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    sheet.title = "new_language"

    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=sheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            # TODO: Actually, this isn't a new language. The difference between
            # adding forms for a language that is not in the LanguageTable yet,
            # but already has forms in the FormTable, and adding something
            # completely new, is washed out by read_single_language. The
            # interpretation of “Does this language still need to be added to
            # the LanguageTable?” for is_new_language is consistent.
            is_new_language=True,
            new=0,
            existing=0,
            skipped=0,
            concepts=1,
        )
    }


def test_add_concept_to_existing_form(single_import_parameters):
    dataset, target, excel, concept_name = single_import_parameters
    c_c_id = dataset["ParameterTable", "id"].name
    c_c_name = dataset["ParameterTable", "name"].name
    concepts = {c[c_c_name]: c[c_c_id] for c in dataset["ParameterTable"]}
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "one",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    mocksheet.title = "new_language"
    # Import this single form in a new language
    read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    )
    mocksheet = MockSingleExcelSheet(
        [
            [
                "English",
                "Form",
                "phonemic",
                "orthographic",
                "Segments",
                "procedural_comment",
                "Comment",
                "Source",
                "phonetic",
                "variants",
            ],
            [
                "two",
                "form",
                "phonemic",
                "orthographic",
                "f o r m",
                "-",
                "None",
                "source[10]",
                "phonetic",
                "",
            ],
        ]
    )
    mocksheet.title = "new_language"
    # Import it again, now both form and language should be existing, but the form has a new concept
    assert read_single_excel_sheet(
        dataset=dataset,
        sheet=mocksheet,
        entries_to_concepts=concepts,
        concept_column=concept_name,
    ) == {
        "new_language": ImportLanguageReport(
            # TODO: Actually, this isn't a new language. The difference between
            # adding forms for a language that is not in the LanguageTable yet,
            # but already has forms in the FormTable, and adding something
            # completely new, is washed out by read_single_language. The
            # interpretation of “Does this language still need to be added to
            # the LanguageTable?” for is_new_language is consistent.
            is_new_language=True,
            new=0,
            existing=0,
            skipped=0,
            concepts=1,
        )
    }
