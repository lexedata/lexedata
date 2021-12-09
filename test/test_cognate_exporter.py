import re
import logging
from pathlib import Path

import pytest
import tempfile
import openpyxl as op

from lexedata import util
from helper_functions import empty_copy_of_cldf_wordlist, copy_to_temp
from lexedata.util.fs import get_dataset
from lexedata.exporter.cognates import ExcelWriter

try:
    from pycldf.dataset import SchemaError
except ImportError:
    # SchemaError was introduced in pycldf 1.24.0
    SchemaError = KeyError


def test_adding_singleton_cognatesets(caplog):
    dataset = get_dataset(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dirname = Path(tempfile.mkdtemp(prefix="lexedata-test"))
    with caplog.at_level(logging.WARNING):
        excel_writer = ExcelWriter(
            dataset=dataset,
            singleton_cognate=True,
        )
        output = dirname / "out.xlsx"
        excel_writer.create_excel(out=output)
    assert re.search("no Status_Column to write", caplog.text)

    # load central concepts from output
    ws = op.load_workbook(output).active
    cogset_index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == "CogSet":
                cogset_index = cell.column - 1
    # when accessing the row as a tuple the index is not 1-based as for excel sheets
    cogset_ids = [row[cogset_index].value for row in ws.iter_rows(min_row=2)]
    assert cogset_ids == [
        "one1",
        "one1",
        "one2",
        "one6",
        "two1",
        "three1",
        "two8",
        "three9",
        "four1",
        "four8",
        "five5",
        "X1_old_paraguayan_guarani",
        "X2_paraguayan_guarani",
        "X3_ache",
    ]


def test_adding_singleton_cognatesets_with_status(caplog):
    dataset = get_dataset(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dataset.add_columns("CognatesetTable", "Status_Column")
    dirname = Path(tempfile.mkdtemp(prefix="lexedata-test"))
    with caplog.at_level(logging.WARNING):
        excel_writer = ExcelWriter(
            dataset=dataset, singleton_cognate=True, singleton_status="NEW"
        )
        output = dirname / "out.xlsx"
        excel_writer.create_excel(out=output)
    assert re.search("no Status_Column to write", caplog.text) is None

    # load central concepts from output
    ws = op.load_workbook(output).active
    cogset_index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == "Status_Column":
                cogset_index = cell.column - 1
    # when accessing the row as a tuple the index is not 1-based as for excel sheets
    status = [row[cogset_index].value for row in ws.iter_rows(min_row=2)]
    assert status == [
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "NEW",
        "NEW",
        "NEW",
    ]


def test_no_cognateset_table(caplog):
    dataset, _ = empty_copy_of_cldf_wordlist(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dataset.remove_table("CognatesetTable")
    # TODO: SystemExit or dataset error?
    with pytest.raises((SystemExit, SchemaError)) as exc_info:
        ExcelWriter(
            dataset=dataset,
        )
    if exc_info.type == SystemExit:
        assert "presupposes a separate CognatesetTable" in caplog.text
        assert "lexedata.edit.add_table" in caplog.text


def test_no_cognate_table(caplog):
    dataset, _ = empty_copy_of_cldf_wordlist(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dataset.remove_table("CognateTable")
    with pytest.raises(SystemExit):
        ExcelWriter(
            dataset=dataset,
        )
    assert "presupposes a separate CognateTable" in caplog.text
    assert "lexedata.edit.add_cognate_table" in caplog.text


def test_no_segment_column(caplog):
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dataset.remove_columns("FormTable", "Segments")
    writer = ExcelWriter(
        dataset=dataset,
    )
    forms = util.cache_table(dataset).values()
    for form in forms:
        assert writer.get_segments(form) is form["form"] and re.search(
            r".*No segments column found. Falling back to cldf form.*", caplog.text
        )
        caplog.clear()


def test_no_comment_column():
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dataset.remove_columns("FormTable", "comment")
    writer = ExcelWriter(
        dataset=dataset,
    )
    forms = util.cache_table(dataset).values()
    for form in forms:
        assert writer.form_to_cell_value(form).strip() == "‘one, one’"
        break


def test_missing_required_column():
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    )
    dataset.remove_columns("FormTable", "ID")
    # TODO: switch to pycldf.dataset.SchemaError
    with pytest.raises(KeyError):
        excel_writer = ExcelWriter(
            dataset=dataset, singleton_cognate=True, singleton_status="NEW"
        )
        output = dataset.tablegroup._fname.parent / "out.xlsx"
        excel_writer.create_excel(out=output)
