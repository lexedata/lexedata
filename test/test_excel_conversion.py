import re
import pytest
import logging
import tempfile
import itertools
from pathlib import Path

import pycldf
import openpyxl

from helper_functions import (
    copy_to_temp,
    copy_to_temp_no_bib,
    copy_to_temp_bad_bib,
    empty_copy_of_cldf_wordlist,
)
from mock_excel import MockSingleExcelSheet
import lexedata.importer.excel_matrix as f
from lexedata.exporter.cognates import ExcelWriter
from lexedata.importer.cognates import (
    import_cognates_from_excel,
)
from lexedata import util


@pytest.fixture(
    params=[
        "data/cldf/minimal/cldf-metadata.json",
        "data/cldf/smallmawetiguarani/cldf-metadata.json",
    ]
)
def cldf_wordlist(request):
    return Path(__file__).parent / request.param


@pytest.fixture(
    params=[
        (
            "data/excel/small.xlsx",
            "data/excel/small_cog.xlsx",
            "data/cldf/smallmawetiguarani/cldf-metadata.json",
        ),
        (
            "data/excel/minimal.xlsx",
            "data/excel/minimal_cog.xlsx",
            "data/cldf/minimal/cldf-metadata.json",
        ),
    ]
)
def excel_wordlist(request):
    return (
        Path(__file__).parent / request.param[0],
        Path(__file__).parent / request.param[1],
        empty_copy_of_cldf_wordlist(Path(__file__).parent / request.param[2]),
    )


@pytest.fixture(
    params=[
        copy_to_temp,
        copy_to_temp_no_bib,
        copy_to_temp_bad_bib,
    ]
)
def working_and_nonworking_bibfile(request):
    return request.param


def test_fromexcel_runs(excel_wordlist):
    lexicon, cogsets, (empty_dataset, original) = excel_wordlist
    f.load_dataset(Path(empty_dataset.tablegroup._fname), str(lexicon), str(cogsets))


def test_fromexcel_correct(excel_wordlist):
    lexicon, cogsets, (empty_dataset, original_md) = excel_wordlist
    original = pycldf.Wordlist.from_metadata(original_md)
    # TODO: parameterize original, like the other parameters, over possible
    # test datasets.
    f.load_dataset(Path(empty_dataset.tablegroup._fname), str(lexicon), str(cogsets))
    form_ids_from_excel = {form["ID"] for form in empty_dataset["FormTable"]}
    form_ids_original = {form["ID"] for form in original["FormTable"]}
    cognate_ids_from_excel = {
        cognate["ID"] for cognate in empty_dataset["CognateTable"]
    }
    cognate_ids_original = {form["ID"] for form in original["CognateTable"]}
    assert form_ids_original == form_ids_from_excel, "{:} and {:} don't match.".format(
        empty_dataset["FormTable"]._parent._fname.parent
        / str(empty_dataset["FormTable"].url),
        original["FormTable"]._parent._fname.parent
        / str(empty_dataset["FormTable"].url),
    )
    assert (
        cognate_ids_original == cognate_ids_from_excel
    ), "{:} and {:} don't match.".format(
        empty_dataset["CognateTable"]._parent._fname.parent
        / str(empty_dataset["CognateTable"].url),
        original["CognateTable"]._parent._fname.parent
        / str(empty_dataset["CognateTable"].url),
    )


def test_toexcel_runs(cldf_wordlist, working_and_nonworking_bibfile):
    filled_cldf_wordlist = working_and_nonworking_bibfile(cldf_wordlist)
    writer = ExcelWriter(
        dataset=filled_cldf_wordlist[0],
        database_url=str(filled_cldf_wordlist[1]),
    )
    forms = util.cache_table(filled_cldf_wordlist[0])
    languages = util.cache_table(filled_cldf_wordlist[0], "LanguageTable").values()
    judgements = util.cache_table(filled_cldf_wordlist[0], "CognateTable").values()
    cogsets = util.cache_table(filled_cldf_wordlist[0], "CognatesetTable").values()
    writer.create_excel(
        rows=cogsets, judgements=judgements, forms=forms, languages=languages
    )
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")
    writer.wb.save(filename=out_filename)


def test_roundtrip(cldf_wordlist, working_and_nonworking_bibfile):
    filled_cldf_wordlist = working_and_nonworking_bibfile(cldf_wordlist)
    dataset, target = filled_cldf_wordlist
    c_formReference = dataset["CognateTable", "formReference"].name
    c_cogsetReference = dataset["CognateTable", "cognatesetReference"].name
    old_judgements = {
        (row[c_formReference], row[c_cogsetReference])
        for row in dataset["CognateTable"].iterdicts()
    }
    writer = ExcelWriter(dataset, database_url="https://example.org/lexicon/{:}")
    forms = util.cache_table(filled_cldf_wordlist[0])
    languages = util.cache_table(filled_cldf_wordlist[0], "LanguageTable").values()
    judgements = util.cache_table(filled_cldf_wordlist[0], "CognateTable").values()
    cogsets = util.cache_table(filled_cldf_wordlist[0], "CognatesetTable").values()
    writer.create_excel(
        rows=cogsets, judgements=judgements, forms=forms, languages=languages
    )

    # Reset the existing cognatesets and cognate judgements, to avoid
    # interference with the the data in the Excel file
    dataset["CognateTable"].write([])
    dataset["CognatesetTable"].write([])

    import_cognates_from_excel(writer.ws, dataset)

    new_judgements = {
        (row[c_formReference], row[c_cogsetReference])
        for row in dataset["CognateTable"].iterdicts()
    }

    assert new_judgements == old_judgements


def test_roundtrip_separator_column(cldf_wordlist, working_and_nonworking_bibfile):
    """Test whether a CognatesetTable column with separator survives a roundtrip."""
    dataset, target = working_and_nonworking_bibfile(cldf_wordlist)
    dataset.add_columns("CognatesetTable", "CommaSeparatedTags")
    dataset["CognatesetTable", "CommaSeparatedTags"].separator = ","
    c_id = dataset["CognatesetTable", "id"].name

    write_back = list(dataset["CognatesetTable"])
    tags = []
    for tag, row in zip(
        itertools.cycle(
            [["two", "tags"], ["single-tag"], [], ["tag;containing;other;separator"]]
        ),
        write_back,
    ):
        tags.append((row[c_id], tag))
        row["CommaSeparatedTags"] = tag
    dataset.write(CognatesetTable=write_back)

    writer = ExcelWriter(dataset, database_url="https://example.org/lexicon/{:}")
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")
    forms = util.cache_table(dataset)
    languages = util.cache_table(dataset, "LanguageTable").values()
    judgements = util.cache_table(dataset, "CognateTable").values()
    cogsets = util.cache_table(dataset, "CognatesetTable").values()
    writer.create_excel(
        rows=cogsets, judgements=judgements, forms=forms, languages=languages
    )

    import_cognates_from_excel(writer.ws, dataset)

    reread_tags = [
        (c[c_id], c["CommaSeparatedTags"]) for c in dataset["CognatesetTable"]
    ]
    reread_tags.sort(key=lambda x: x[0])
    tags.sort(key=lambda x: x[0])
    assert reread_tags == tags


def test_cell_comments():
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    excel_filename = Path(__file__).parent / "data/excel/judgement_cell_with_note.xlsx"

    ws = openpyxl.load_workbook(excel_filename).active
    import_cognates_from_excel(ws, dataset)
    cognates = {
        cog["ID"]: {
            k: v for k, v in cog.items() if k in {"Form_ID", "Cognateset", "Comment"}
        }
        for cog in dataset["CognateTable"]
    }
    assert cognates == {
        "autaa_Woman-cogset": {
            "Cognateset": "cogset",
            "Comment": "Comment on judgement",
            "Form_ID": "autaa_Woman",
        }
    }
    cognatesets = {
        cog["ID"]: {k: v for k, v in cog.items()} for cog in dataset["CognatesetTable"]
    }
    assert cognatesets == {
        "cogset": {"Name": "cogset", "Comment": "Cognateset-comment", "ID": "cogset"}
    }


def test_cell_comments_and_comment_column(caplog):
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    excel_filename = Path(__file__).parent / "data/excel/judgement_cell_with_note.xlsx"

    sheet = openpyxl.load_workbook(excel_filename).active
    sheet.insert_cols(2)
    sheet.cell(row=1, column=2, value="Comment")
    sheet.cell(row=2, column=2, value="Comment")

    with caplog.at_level(logging.INFO):
        import_cognates_from_excel(sheet, dataset)

    assert "from the cell comments" in caplog.text

    cognates = {
        cog["ID"]: {
            k: v for k, v in cog.items() if k in {"Form_ID", "Cognateset", "Comment"}
        }
        for cog in dataset["CognateTable"]
    }
    assert cognates == {
        "autaa_Woman-cogset": {
            "Cognateset": "cogset",
            "Comment": "Comment on judgement",
            "Form_ID": "autaa_Woman",
        }
    }
    cognatesets = {
        cog["ID"]: {k: v for k, v in cog.items()} for cog in dataset["CognatesetTable"]
    }
    assert cognatesets == {
        "cogset": {"Name": "cogset", "Comment": "Cognateset-comment", "ID": "cogset"}
    }


def test_cell_comments_export():
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")

    writer = ExcelWriter(dataset, database_url="https://example.org/lexicon/{:}")
    forms = util.cache_table(dataset)
    languages = sorted(
        util.cache_table(dataset, "LanguageTable").values(), key=lambda x: x["name"]
    )
    judgements = util.cache_table(dataset, "CognateTable").values()
    cogsets = util.cache_table(dataset, "CognatesetTable").values()
    writer.create_excel(
        rows=cogsets, judgements=judgements, forms=forms, languages=languages
    )

    for col in writer.ws.iter_cols():
        pass
    assert (
        col[-1].comment and col[-1].comment.content
    ), "Last row of last column should contain a judgement, with a comment attached to it."
    assert (
        col[-1].comment.content == "A judgement comment"
    ), "Comment should match the comment from the cognate table"


def test_excel_messy_row(caplog):
    # Build a dataset with forms F1, F2, F3 in languages L1, L2 and
    # CognateTable columns ID and Status
    dataset = util.fs.new_wordlist(
        FormTable=[
            {"ID": "F1", "Language_ID": "L1", "Form": "f1", "Parameter_ID": "C"},
            {"ID": "F2", "Language_ID": "L2", "Form": "f1", "Parameter_ID": "C"},
            {"ID": "F3", "Language_ID": "L1", "Form": "f1", "Parameter_ID": "C"},
        ],
        LanguageTable=[{"ID": "L1", "Name": "L1"}, {"ID": "L2", "Name": "L2"}],
        ParameterTable=[{"ID": "C"}],
        CognateTable=[],
        CognatesetTable=[],
    )
    # TODO: Ensure FormTable does not need a value
    dataset.add_columns("FormTable", "value")
    dataset["FormTable", "value"].required = False
    dataset.add_columns("CognatesetTable", "Status")
    dataset.add_columns("CognatesetTable", "comment")

    # Construct a sheet with a messy cognateset header
    messy_sheet = MockSingleExcelSheet(
        [
            [
                "CogSet",
                "Status",
                "L1",
                "L2",
            ],
            [
                "S1",
                "valid",
                "F1",
                "F2",
            ],
            [
                "",
                "invalid",
                "F3",
            ],
        ]
    )
    for cell in [(2, 3), (3, 3), (2, 4)]:
        messy_sheet.cell(*cell).hyperlink = "/{:}".format(messy_sheet.cell(*cell).value)

    # Cognate-import this dataset
    with caplog.at_level(logging.INFO):
        import_cognates_from_excel(
            messy_sheet,
            dataset,
        )

    # Check that cognateset S1 contains form F3
    assert ("F3", "S1") in [
        (j["Form_ID"], j["Cognateset_ID"]) for j in dataset["CognateTable"]
    ]

    # Check for warning in caplog
    assert re.search("[Rr]ow 3 .* no cognate ?set .*'Status': 'invalid'", caplog.text)
