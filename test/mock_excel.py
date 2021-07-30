import openpyxl


class MockSingleExcelSheet(openpyxl.worksheet.worksheet.Worksheet):
    def __init__(self, data, title="MockSingleExcelSheet"):
        super().__init__(openpyxl.Workbook(), title)
        for i, dr in enumerate(data, 1):
            for j, d in enumerate(dr, 1):
                if isinstance(d, dict):
                    self.cell(row=i, column=j, value=d["value"]).comment = d["comment"]
                else:
                    self.cell(row=i, column=j, value=d)
