import tempfile
import unicodedata

import openpyxl as op
from lexedata.util.excel import clean_cell_value


# TODO: discuss these multiple asserts
def test_cell_value():
    wb = op.Workbook()
    ws = wb.active
    ws["A2"] = 2
    ws["A3"] = 3.14
    ws["A4"] = "4"
    ws["B1"] = "端ber"
    ws["B2"] = unicodedata.normalize("NFD", "端ber")
    ws["B3"] = "Line\nover\nline"
    _, filename = tempfile.mkstemp(suffix=".xlsx")
    wb.save(filename)
    del wb, ws

    wb = op.load_workbook(filename)
    ws = wb.active

    assert clean_cell_value(ws["A1"]) == ""
    assert clean_cell_value(ws["A2"]) == 2
    assert clean_cell_value(ws["A3"]) == 3.14
    assert clean_cell_value(ws["A4"]) == "4"
    assert clean_cell_value(ws["B1"]) == unicodedata.normalize("NFC", "端ber")
    assert clean_cell_value(ws["B2"]) == unicodedata.normalize("NFC", "端ber")
    assert clean_cell_value(ws["B3"]) == "Line;\tover;\tline"
