# -*- coding: utf-8 -*-
import openpyxl as op
import csv
import re
from collections import defaultdict
from abc import ABC
from math import ceil
import unidecode as uni

from .exceptions import *
from .cellparser import CellParser

#replacing none values with ''
replace_none = lambda x: "" if x == None else x

#lambda function for getting comment of excel cell if comment given
replace_none_comment = lambda x: x.content if x  else ""
comment_getter = lambda x: replace_none_comment(x.comment)


#header for files
header_forms = ["form_id", "language_id", "concept_id",
                "phonemic", "phonetic", "orthography", "source", "form_comment",
                "concept_comment"]

header_concepts = ["concept_id",
                   "set", "english", "english_strict", "spanish", "portuguese", "french", "concept_comment"]

header_languages = ["language_id",
                   "language", "curator", "language_comment"]
header_form_concept = ["concept_id", "list_forms"]


class Cell(ABC):

    #class member functions for bracket checking
    _one_bracket = lambda opening, closing, str, nr: str[0] == opening and str[-1] == closing and \
                                                 (str.count(opening) == str.count(closing) == nr)
    _comment_bracket = lambda str: str.count("(") == str.count(")")

    def __init__(self, values, size):

        #basic test for all cells:
        #len of requiered cell type
        assert len(values) == size, "cell has not required size"
        #not all values empty, none or false
        assert any(values) == True, "cell is entirely empty"

        self._data = tuple(values) #may not be changed anymore

    def __hash__(self):
        return self

    def write(self, out):
        """write to given csv.writer() out"""
        out.writerow(self._data)

    def print_cell(self):
        s = [ele.ljust(10*ceil(len(ele)/10)) for ele in self._data]
        s = "|".join(s)
        print(s)

class LanguageCell(Cell):
    """
       a language element consists of 3 fields:
           (language_id,language,curator, language_comment)
       sharing language_id with a form element; language_comment being the comment of the excel cell
       """
    __language_pattern = re.compile(r"\W+")
    __create_language_id = lambda x: uni.unidecode(LanguageCell.__language_pattern.sub("_", x)).lower()

    def __init__(self, languagecol):
        """
        creates a language cell out of a given languagecol
        :param cell: column containing two cells
        """
        values = [cell.value for cell in languagecol]
        language_id = LanguageCell.__create_language_id(values[0])
        values.insert(0, language_id)
        values.append(comment_getter(languagecol[1]))
        super().__init__(values, size=4)

    def warn(self):
        try:
            if "???" in self._data[1]:
                raise LanguageElementError(self._data)
        except LanguageElementError as err:
            print(err.message)
            input()

class FormCell(Cell):
    """
       a form element consists of 10 fields:
           (form_id,language_id,concept_id,
           phonemic,phonetic,orthography,source,form_comment, #core values
           concept_comment)
       sharing concept_id and concept_comment with a concept element and language_id with a language element
       concept_comment refers to te comment of the cell containing the english meaning
       """


    def __init__(self, language_id, concept_id, cell_comment, form_comment, number, values):
        """
        creates cell out of different parts, checks values for errors, creates id
        :param language_id:
        :param concept_id:
        :param cell_comment:
        :param form_comment:
        :param number: number of form element contained in form cell
        :param values: phonemic, phonetic, orthographic, comment, source
        """
        values = [replace_none(e) for e in values]
        phonemic = values[0]
        if phonemic != "" and phonemic != "No value":
            if not Cell._one_bracket("/", "/", phonemic, 2):
                raise FormCellError(values, "phonemic")
        phonetic = values[1]
        if phonetic != "" and phonetic != "No value":
            if not Cell._one_bracket("[", "]", phonetic, 1):
                raise FormCellError(values, "phonetic")
        ortho = values[2]
        if ortho != "" and ortho != "No value":
            if not Cell._one_bracket("<", ">", ortho, 1):
                raise FormCellError(values, "orthographic")
        comment = values[3]
        if comment != "" and comment != "No value":
            if not Cell._comment_bracket(comment):
                raise FormCellError(values, "comment")

        #subsitute empty source by {1}
        values[4] = "{1}" if values[4] == "" else values[4]


        values.insert(0, concept_id)
        values.insert(0, language_id)
        values.insert(0, "_".join([language_id, concept_id, str(number)]))
        values.append(form_comment)
        values.append(cell_comment)
        super().__init__(values, 10)

class ConceptCell(Cell):
    """
    a concept element consists of 8 fields:
        (concept_id,set,english,english_strict,spanish,portuguese,french,cell_comment)
    sharing concept_id and concept_comment with a form element
    cell_concept refers to te comment of the cell containing the english meaning
    """

    #protected static class variable for creating unique ids, regex-pattern
    _conceptdict = defaultdict(int)
    __id_pattern = re.compile(r"^(.+?)(?:[,\(\@](?:.+)?|)$")


    def __init__(self, conceptrow):
        """
        creates a ConceptCell out of a given excel row
        :param conceptcell: excel row containing a concept entry
        """
        #values of cell
        values = [replace_none(cell.value) for cell in conceptrow]
        #comment of cell
        values.append(comment_getter(conceptrow[1]))
        #create id
        form_id = ConceptCell.create_id(values[1])
        values.insert(0, form_id)
        super().__init__(values, 8) #concept cell has length 8

    @staticmethod
    def create_id(meaning):
        """
        creates unique id out of given english meaning
        :param meaning: str
        :return: unique id (str)
        """
        mymatch = ConceptCell.__id_pattern.match(meaning)
        if mymatch:
            mymatch = mymatch.group(1)
            mymatch = mymatch.replace(" ", "")
            ConceptCell._conceptdict[mymatch] += 1
            mymatch += str(ConceptCell._conceptdict[mymatch])
            return mymatch
        else:
            raise ConceptIdError


def main():
    #path to excel file
    wb = op.load_workbook(filename="Copy of TG_comparative_lexical_online_MASTER.xlsx")
    sheets = wb.sheetnames

    iter_forms = wb[sheets[0]].iter_rows(min_row=3, min_col=7, max_col=44) #iterates over rows with forms

    iter_concept = wb[sheets[0]].iter_rows(min_row=3, max_col=6) #iterates over rows with concepts

    iter_language = wb[sheets[0]].iter_cols(min_row=1, max_row=2, min_col=7, max_col=44)  #iterates over language columns
    lan_ids = [] #language ids are needed for form elements

##########################################
    #create language csv and collect language ids
    with open("languages.csv", "w", encoding="utf8", newline="") as lanout:
        lancsv = csv.writer(lanout, quotechar='"', quoting=csv.QUOTE_MINIMAL)
        lancsv.writerow(header_languages)
        for lan_col in iter_language:
            c = LanguageCell(lan_col)
            c.warn()
            lan_ids.append(c._data[0])
            c.write(lancsv)

#########################################
    #create concept and form elements at the same time (for identical ids)
    with open("forms.csv", "w", encoding="utf8", newline="") as formsout, \
            open("concepts.csv", "w", encoding="utf8", newline="") as conceptsout, \
            open("form_to_concept.csv", "w", encoding="utf8", newline="") as formsout2:

        formscsv = csv.writer(formsout, quotechar='"', quoting=csv.QUOTE_MINIMAL)
        formscsv.writerow(header_forms)

        concsv = csv.writer(conceptsout, quotechar='"', quoting=csv.QUOTE_MINIMAL)
        concsv.writerow(header_concepts)

        formconcsv = csv.writer(formsout2, quotechar='"', quoting=csv.QUOTE_MINIMAL)
        formconcsv.writerow(header_form_concept)

        for row_forms, row_con in zip(iter_forms, iter_concept):
            c_con = ConceptCell(row_con)
            c_con.write(concsv)
            con_comment = comment_getter(row_con[1])
            con_id = c_con._data[0]

            form_ids = [] #collect form_ids for form_to_concept csv
            for f_cell in enumerate(row_forms):
                if f_cell[1].value:
                    try:
                        #csvlines = [] #collect all good form elements
                        this_lan_id = lan_ids[f_cell[0]]
                        f_comment = comment_getter(f_cell[1])
                        for f_ele in enumerate(CellParser(f_cell[1]), start=1):
                            #error is thrown in constructor of FormCell
                            c_form = FormCell(this_lan_id,con_id, con_comment, f_comment, f_ele[0], f_ele[1])
                            form_ids.append(c_form._data[0])

                            #c_form.print_cell()
                            c_form.write(formscsv) #activate and all correct form elements are written out
                            #csvlines.append(c_form)

                        #no CellError til here, all fine, write lines
                        #for cell in csvlines:
                        #    cell.write(formscsv)

                    except CellParsingError as err:
                        print("CellParsingError - somethings quite wrong")
                        print(err.message)
                        input()

                    except TypeError:
                        print("Type Error - None might be a problem")
                        input()

                    except FormCellError as err:
                        pass
                        #prints all error messages
                        #print(f_cell[1].value)
                        #print(err.message)
                        #input()

            #write to form_to_concept.csv
            formconcsv.writerow([con_id, ",".join(form_ids)])
        #print(CellParser._wrongorder)
##########################################################################

if __name__ == "__main__":
    main()
