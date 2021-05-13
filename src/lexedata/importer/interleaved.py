"""Import some Bantu data

This script is a very rough sketch for importing yet another shape of data.
Here, every even column contains cells with forms (some cells containing
multiple forms), while the odd columns contain the associated cognate codes
(generally in the same order).

"""
import re
import csv
import logging
from pathlib import Path
import os

import openpyxl

import lexedata.cli as cli


def import_interleaved(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    logger: logging.Logger = cli.logger,
) -> list:
    comma_or_semicolon = re.compile("[,;]\\W*")

    concepts = []
    for concept_metadata in ws.iter_cols(min_col=1, max_col=1, min_row=2):
        for entry, cogset in zip(concept_metadata[::2], concept_metadata[1::2]):
            try:
                concepts.append(entry.value.strip())
            except AttributeError:
                break

    for language in ws.iter_cols(min_col=2):
        language_name = language[0].value
        for c, (entry, cogset) in enumerate(zip(language[1::2], language[2::2])):
            if not entry.value:
                assert not cogset.value
                continue
            bracket_level = 0
            i = 0
            f = entry.value.strip()
            forms = []
            while i < len(f):
                match = comma_or_semicolon.match(f[i:])
                if f[i] == "(":
                    bracket_level += 1
                    i += 1
                    continue
                elif f[i] == ")":
                    bracket_level -= 1
                    i += 1
                    continue
                elif bracket_level:
                    i += 1
                    continue
                elif match:
                    forms.append(f[:i].strip())
                    i += match.span()[1]
                    f = f[i:]
                    i = 0
                else:
                    i += 1

            forms.append(f.strip())

            if type(cogset.value) == float:
                cogsets = [str(int(cogset.value))]
            else:
                if isinstance(cogset.value, int):
                    cogset = str(cogset.value)
                    print(cogset)
                else:
                    cogset = cogset.value
                cogsets = comma_or_semicolon.split(cogset.strip())

            if len(cogsets) == 1 or len(cogsets) == len(forms):
                True
            else:
                logger.warning(
                    "{:}: Forms ({:}) did not match cognates ({:})".format(
                        entry.coordinate, ", ".join(forms), ", ".join(cogsets)
                    )
                )
            for form, cogset in zip(forms, cogsets + [None]):
                yield [language_name, concepts[c], form, None, cogset]


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", type=Path, help="The Excel file to parse")
    parser.add_argument(
        "--directory",
        type=Path,
        default=Path(os.getcwd()),
        help="Path to directory where forms.csv is created (default: current working directory)",
    )
    cli.add_log_controls(parser)
    args = parser.parse_args()
    logger = cli.setup_logging(args)

    ws = openpyxl.load_workbook(args.excel)

    w = csv.writer(open(Path(args.directory) / "forms.csv", "w", encoding="utf-8"))
    w.writerow(["Language_ID", "Concept_ID", "Form", "Comment", "Cognateset"])

    for sheet in ws.worksheets:
        for row in import_interleaved(sheet, logger=logger):
            w.writerow(row)
