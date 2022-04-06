# -*- coding: utf-8 -*-

import argparse
import logging
import re
import typing as t
from pathlib import Path

import openpyxl
import pycldf

import lexedata.cli as cli
import lexedata.util.excel as cell_parsers
from lexedata.edit.add_status_column import add_status_column_to_table
from lexedata.types import (
    CogSet,
    Concept,
    Form,
    Judgement,
    Language,
    Object,
    R,
    RowObject,
)
from lexedata.util import edit_distance, string_to_id
from lexedata.util.excel import clean_cell_value, get_cell_comment

Ob = t.TypeVar("Ob", bound=Object)

# NOTE: Excel uses 1-based indices, this shows up in a few places in this file.


def cells_are_empty(cells: t.Iterable[openpyxl.cell.Cell]) -> bool:
    return not any([clean_cell_value(cell) for cell in cells])


class DB:
    """An in-memobry cache of a dataset.

    The cache_dataset method is only called in the load_dataset method, but
    also used for finer control by the cognates importer. This means that if
    you would load the CognateParser directly, it doesn't work as the cache is
    empty and the code will throw errors (e.g. when trying to look for
    candidates we get a key error). If you use the CognateParser elsewhere,
    make sure to cache the dataset explicitly, eg. by using DB.from_dataset!

    """

    cache: t.Dict[str, t.Dict[t.Hashable, t.Dict[str, t.Any]]]
    source_ids: t.Set[str]

    def __init__(self, output_dataset: pycldf.Wordlist):
        """Create a new *empty* cache associated with a dataset."""
        self.dataset = output_dataset
        self.cache = {}
        self.source_ids = set()

    @classmethod
    def from_dataset(k, dataset, logger: cli.logging.Logger = cli.logger):
        """Create a (filled) cache from a dataset."""
        ds = k(dataset)
        ds.cache_dataset(logger=logger)
        return ds

    def cache_dataset(self, logger: cli.logging.Logger = cli.logger):
        logger.info("Caching dataset into memory…")
        for table in self.dataset.tables:
            table_type = (
                table.common_props.get("dc:conformsTo", "").rsplit("#", 1)[1]
                or table.url
            )
            (id,) = table.tableSchema.primaryKey
            # Extent may be wrong, but it's usually at least roughly correct
            # and a better indication of the table size than none at all.
            try:
                self.cache[table_type] = {
                    row[id]: row
                    for row in cli.tq(
                        table,
                        task="Cache the dataset",
                        total=table.common_props.get("dc:extent"),
                    )
                }
            except FileNotFoundError:
                self.cache[table_type] = {}

        for source in self.dataset.sources:
            self.source_ids.add(source.id)

    def drop_from_cache(self, table: str):
        self.cache[table] = {}

    def retrieve(self, table_type: str):
        return self.cache[table_type].values()

    def add_source(self, source_id):
        self.source_ids.add(source_id)

    def empty_cache(self):
        self.cache = {
            # TODO: Is there a simpler way to get the list of all tables?
            table.common_props.get("dc:conformsTo", "").rsplit("#", 1)[1]
            or table.url: {}
            for table in self.dataset.tables
        }

    def write_dataset_from_cache(self, tables: t.Optional[t.Iterable[str]] = None):
        if tables is None:
            tables = self.cache.keys()
        for table_type in tables:
            self.dataset[table_type].common_props["dc:extent"] = self.dataset[
                table_type
            ].write(self.retrieve(table_type))
        self.dataset.write_metadata()
        # TODO: Write BIB file, without pycldf
        with self.dataset.bibpath.open("w", encoding="utf-8") as bibfile:
            for source in self.source_ids:
                print("@misc{" + source + ", title={" + source + "} }", file=bibfile)

    def associate(
        self, form_id: str, row: RowObject, comment: t.Optional[str] = None
    ) -> bool:
        form = self.cache["FormTable"][form_id]
        if row.__table__ == "CognatesetTable":
            id = self.dataset["CognatesetTable", "id"].name
            try:
                column = self.dataset["FormTable", "cognatesetReference"]
            except KeyError:
                judgements = self.cache["CognateTable"]
                cognateset = row[self.dataset["CognatesetTable", "id"].name]
                judgement = Judgement(
                    {
                        self.dataset["CognateTable", "id"].name: "{:}-{:}".format(
                            form_id, cognateset
                        ),
                        self.dataset["CognateTable", "formReference"].name: form_id,
                        self.dataset[
                            "CognateTable", "cognatesetReference"
                        ].name: cognateset,
                        self.dataset["CognateTable", "comment"].name: comment or "",
                    }
                )
                self.make_id_unique(judgement)
                judgements[judgement[id]] = judgement
                return True
        elif row.__table__ == "ParameterTable":
            column = self.dataset["FormTable", "parameterReference"]
            id = self.dataset["ParameterTable", "id"].name

        if column.separator is None:
            form[column.name] = row[id]
        else:
            form.setdefault(column.name, []).append(row[id])
        return True

    def insert_into_db(self, object: Object) -> None:
        id = self.dataset[object.__table__, "id"].name
        assert object[id] not in self.cache[object.__table__]
        self.cache[object.__table__][object[id]] = object

    def make_id_unique(self, object: Object) -> str:
        id = self.dataset[object.__table__, "id"].name
        raw_id = object[id]
        i = 0
        while object[id] in self.cache[object.__table__]:
            i += 1
            object[id] = "{:}_{:d}".format(raw_id, i)
        return object[id]

    def find_db_candidates(
        self,
        object: Ob,
        properties_for_match: t.Iterable[str],
        edit_dist_threshold: t.Optional[int] = None,
    ) -> t.Iterable[str]:
        if edit_dist_threshold:

            def match(x, y):
                if (not x and y) or (x and not y):
                    return False
                return edit_distance(x, y) <= edit_dist_threshold

        else:

            def match(x, y):
                return x == y

        return [
            candidate
            for candidate, properties in self.cache[object.__table__].items()
            if all(
                match(properties.get(p), object.get(p)) for p in properties_for_match
            )
        ]

    def commit(self):
        pass


class ExcelParser(t.Generic[R]):
    def __init__(
        self,
        output_dataset: pycldf.Dataset,
        row_type: t.Type[R],
        top: int = 2,
        cellparser: cell_parsers.NaiveCellParser = cell_parsers.CellParser,
        # The following column names should be generated from CLDF terms. This
        # will likely mean that the __init__ method would have to get a
        # slightly different signature, to take that dependency on the output
        # dataset into account.
        row_header: t.List[str] = ["set", "Name", None],
        check_for_match: t.List[str] = ["ID"],
        check_for_row_match: t.List[str] = ["Name"],
        check_for_language_match: t.List[str] = ["Name"],
        fuzzy=0,
    ) -> None:
        self.row_header = row_header
        self.row_type = row_type
        try:
            self.cell_parser = cellparser(output_dataset)
        except TypeError:
            self.cell_parser = cellparser
        self.top = top
        self.left = len(row_header) + 1
        self.check_for_match = check_for_match
        self.check_for_row_match = check_for_row_match
        self.check_for_language_match = check_for_language_match
        self.db = DB(output_dataset)
        self.fuzzy = fuzzy

    def on_language_not_found(
        self, language: t.Dict[str, t.Any], cell_identifier: t.Optional[str] = None
    ) -> bool:
        """Create language"""
        return True

    def on_row_not_found(
        self, row_object: R, cell_identifier: t.Optional[str] = None
    ) -> bool:
        """Create row object"""
        return True

    def on_form_not_found(
        self,
        form: t.Dict[str, t.Any],
        cell_identifier: t.Optional[str] = None,
        language_id: t.Optional[str] = None,
    ) -> bool:
        """Create form"""
        return True

    def language_from_column(self, column: t.List[openpyxl.cell.Cell]) -> Language:
        data = [clean_cell_value(cell) for cell in column[: self.top - 1]]
        comment = get_cell_comment(column[0])
        id = string_to_id(data[0])
        return Language(
            # an id candidate must be provided, which is transformed into a unique id
            ID=id,
            Name=data[0],
            Comment=comment,
        )

    def properties_from_row(self, row: t.List[openpyxl.cell.Cell]) -> t.Optional[R]:
        c_id = self.db.dataset[self.row_type.__table__, "id"].name
        c_comment = self.db.dataset[self.row_type.__table__, "comment"].name
        c_name = self.db.dataset[self.row_type.__table__, "name"].name
        data = [clean_cell_value(cell) for cell in row[: self.left - 1]]
        properties = dict(zip(self.row_header, data))
        # delete all possible None entries coming from row_header
        while None in properties.keys():
            del properties[None]

        # fetch cell comment
        comment = get_cell_comment(row[0])
        properties[c_comment] = comment

        # cldf_name serves as cldf_id candidate
        properties[c_id] = properties[c_name]
        # create new row object

        return self.row_type(properties)

    def parse_all_languages(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> t.Dict[str, str]:
        """Parse all language descriptions in the focal sheet.

        Returns
        =======
        languages: A dictionary mapping columns ("B", "C", "D", …) to language IDs
        """
        languages_by_column: t.Dict[str, str] = {}
        # iterate over language columns
        for lan_col in cli.tq(
            sheet.iter_cols(min_row=1, max_row=self.top - 1, min_col=self.left),
            task="Parse all languages",
            total=sheet.max_column - self.left,
        ):
            c_l_id = self.db.dataset["LanguageTable", "id"].name
            if cells_are_empty(lan_col):
                # Skip empty languages
                continue
            language = self.language_from_column(lan_col)
            candidates = self.db.find_db_candidates(
                language,
                self.check_for_language_match,
            )
            for language_id in candidates:
                break
            else:
                if self.on_language_not_found(language, lan_col[0].coordinate):
                    self.db.insert_into_db(language)
                else:
                    continue
                language_id = language[c_l_id]
            languages_by_column[lan_col[0].column] = language_id

        return languages_by_column

    def parse_cells(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        status_update: t.Optional[str] = None,
    ) -> None:
        languages = self.parse_all_languages(sheet)
        row_object: t.Optional[R] = None
        for row in cli.tq(
            sheet.iter_rows(min_row=self.top),
            task="Parsing cells",
            total=sheet.max_row - self.top,
        ):
            row_header, row_forms = row[: self.left - 1], row[self.left - 1 :]
            # Parse the row header, creating or retrieving the associated row
            # object (i.e. a concept or a cognateset)
            properties = self.properties_from_row(row_header)
            if properties:
                c_r_id = self.db.dataset[properties.__table__, "id"].name
                try:
                    c_r_name = self.db.dataset[properties.__table__, "name"].name
                except KeyError:
                    c_r_name = None
                similar = self.db.find_db_candidates(
                    properties, self.check_for_row_match
                )
                for row_id in similar:
                    properties[c_r_id] = row_id
                    break
                else:
                    if self.on_row_not_found(
                        properties, cell_identifier=row[0].coordinate
                    ):
                        if c_r_id not in properties:
                            properties[c_r_id] = string_to_id(
                                str(properties.get(c_r_name, ""))
                            )
                        self.db.make_id_unique(properties)
                        self.db.insert_into_db(properties)
                    else:
                        continue
                # check the fields of properties are not empty, if so, set row
                # object to properties. This means that if there is no
                # properties object, of if it is empty, the previous row object
                # is re-used. This is intentional.
                if any(properties.values()):
                    row_object = properties

            if row_object is None:
                if any(c.value for c in row_forms):
                    raise AssertionError(
                        "Your first data row didn't have a name. "
                        "Please check your format specification or ensure the first row has a name."
                    )
                else:
                    continue
            # Parse the row, cell by cell
            for cell_with_forms in row_forms:
                try:
                    this_lan = languages[cell_with_forms.column]
                except KeyError:
                    continue

                # Parse the cell, which results (potentially) in multiple forms
                if row_object.__table__ == "FormTable":
                    raise NotImplementedError(
                        "TODO: I am confused why what I'm doing right now ever landed on my agenda, but you seem to have gotten me to attempt it. Please contact the developers and tell them what you did, so they can implement the thing you tried to do properly!"
                    )
                    c_f_form = self.db.dataset[row_object.__table__, "form"].name
                for params in self.cell_parser.parse(
                    cell_with_forms,
                    this_lan,
                    f"{sheet.title}.{cell_with_forms.coordinate}",
                ):
                    if row_object.__table__ == "FormTable":
                        if params[c_f_form] == "?":
                            continue
                        else:
                            self.handle_form(
                                params,
                                row_object,
                                cell_with_forms,
                                this_lan,
                                status_update,
                            )
                    else:
                        self.handle_form(
                            params, row_object, cell_with_forms, this_lan, status_update
                        )
        self.db.commit()

    def handle_form(
        self,
        params,
        row_object: R,
        cell_with_forms,
        this_lan: str,
        status_update: t.Optional[str],
    ):
        form = Form(params)
        c_f_id = self.db.dataset["FormTable", "id"].name
        c_f_language = self.db.dataset["FormTable", "languageReference"].name
        c_f_value = self.db.dataset["FormTable", "value"].name
        c_r_id = self.db.dataset[row_object.__table__, "id"].name

        if c_f_id not in form:
            # create candidate for form[id]
            form[c_f_id] = "{:}_{:}".format(form[c_f_language], row_object[c_r_id])
        candidate_forms = iter(self.db.find_db_candidates(form, self.check_for_match))
        try:
            # if a candidate for form already exists, don't add the form
            form_id = next(candidate_forms)
            self.db.associate(form_id, row_object)
        except StopIteration:
            # no candidates. form is created or not.
            if self.on_form_not_found(
                form, cell_identifier=cell_with_forms, language_id=this_lan
            ):
                form[c_f_id] = "{:}_{:}".format(form[c_f_language], row_object[c_r_id])
                form[c_f_value] = cell_with_forms.value
                # add status update if given
                if status_update:
                    form["Status_Column"] = status_update
                self.db.make_id_unique(form)
                self.db.insert_into_db(form)
                form_id = form[c_f_id]
                self.db.associate(form_id, row_object)


class ExcelCognateParser(ExcelParser[CogSet]):
    def __init__(
        self,
        output_dataset: pycldf.Dataset,
        row_type=CogSet,
        top: int = 2,
        cellparser: cell_parsers.NaiveCellParser = cell_parsers.CellParser,
        row_header=["set", "Name", None],
        check_for_match: t.List[str] = ["Form"],
        check_for_row_match: t.List[str] = ["Name"],
        check_for_language_match: t.List[str] = ["Name"],
    ) -> None:
        super().__init__(
            output_dataset=output_dataset,
            row_type=CogSet,
            top=top,
            cellparser=cellparser,
            check_for_match=check_for_match,
            row_header=row_header,
            check_for_row_match=check_for_row_match,
            check_for_language_match=check_for_language_match,
        )

    def on_language_not_found(
        self,
        language: t.Dict[str, t.Any],
        cell_identifier: t.Optional[str] = None,
    ) -> bool:
        """Should I add a missing object? No, the object missing is an error.

        Raise an exception (ObjectNotFoundWarning) reporting the missing object and cell.

        Raises
        ======
        ObjectNotFoundWarning

        """
        rep = language.get("cldf_name", language.get("cldf_id", repr(language)))
        raise ValueError(
            f"Failed to find object {rep:} in the database. In cell: {cell_identifier:}"
        )

    def on_row_not_found(
        self, row_object: CogSet, cell_identifier: t.Optional[str] = None
    ) -> bool:
        """Create row object"""
        return True

    def on_form_not_found(
        self,
        form: t.Dict[str, t.Any],
        cell_identifier: t.Optional[str] = None,
        language_id: t.Optional[str] = None,
        logger: cli.logging.Logger = cli.logger,
    ) -> bool:
        """Should I add a missing object? No, but inform the user.

        Send a warning (ObjectNotFoundWarning) reporting the missing object and cell.

        Returns
        =======
        False: The object should not be added.

        """
        rep = form.get("cldf_id", repr(form))
        logger.warning(
            f"Unable to find form {rep} in cell {cell_identifier} in the dataset. "
            f"This cognate judgement was skipped. "
            f"Please make sure that the form is present in forms.csv or in the file "
            f"used for the Wordlist importation."
        )
        # Do a fuzzy search
        for row in self.db.find_db_candidates(
            form, self.check_for_match, edit_dist_threshold=4
        ):
            logger.info(f"Did you mean {row} ?")
        return False

    def properties_from_row(
        self, row: t.List[openpyxl.cell.Cell]
    ) -> t.Optional[CogSet]:
        # TODO: get_cell_comment with unicode normalization or not? -> yes, comments also
        c_id = self.db.dataset[self.row_type.__table__, "id"].name
        c_comment = self.db.dataset[self.row_type.__table__, "comment"].name
        c_name = self.db.dataset[self.row_type.__table__, "name"].name
        data = [clean_cell_value(cell) for cell in row[: self.left - 1]]
        properties = dict(zip(self.row_header, data))
        # delete all possible None entries coming from row_header
        while None in properties.keys():
            del properties[None]

        # fetch cell comment
        comment = get_cell_comment(row[0])
        properties[c_comment] = comment

        # cldf_name serves as cldf_id candidate
        properties[c_id] = properties.get(c_id) or properties[c_name]
        # create new row object
        return self.row_type(properties)

    def associate(
        self, form_id: str, row: RowObject, comment: t.Optional[str] = None
    ) -> bool:
        c_id = self.db.dataset[self.row_type.__table__, "id"].name
        assert (
            row.__table__ == "CognatesetTable"
        ), "Expected CognateSet, but got {:}".format(row.__class__)
        row_id = row[c_id]
        judgement = Judgement(
            cldf_id=f"{form_id}-{row_id}",
            cldf_formReference=form_id,
            cldf_cognatesetReference=row_id,
            cldf_comment=comment or "",
        )
        self.db.make_id_unique(judgement)
        return self.db.insert_into_db(judgement)

    def handle_form(
        self,
        params,
        row_object: CogSet,
        cell_with_forms,
        this_lan,
        status_update: t.Optional[str],
    ):
        try:
            if params.__table__ == "CognateTable":
                row_id = row_object[self.db.dataset["CognatesetTable", "id"].name]
                params[
                    self.db.dataset["CognateTable", "cognatesetReference"].name
                ] = row_id
                c_j_id = self.db.dataset["CognateTable", "id"].name
                if c_j_id not in params:
                    form_id = params[
                        self.db.dataset["CognateTable", "formReference"].name
                    ]
                    params[c_j_id] = f"{form_id}-{row_id}"
                    self.db.make_id_unique(params)
                # add status update if given
                if status_update:
                    params["Status_Column"] = status_update
                self.db.insert_into_db(params)
                return
        except AttributeError:
            pass

        # Deal with the more complex case where we are given a form and need
        # to discern what to do with it.
        form = Form(params)
        c_f_id = self.db.dataset["FormTable", "id"].name

        if c_f_id in form:
            self.db.associate(form[c_f_id], row_object)
        else:
            try:
                form_id = next(
                    iter(self.db.find_db_candidates(form, self.check_for_match))
                )
                self.db.associate(form_id, row_object)
            except StopIteration:
                if self.on_form_not_found(
                    form,
                    cell_identifier=cell_with_forms.coordinate,
                    language_id=this_lan,
                ):
                    raise NotImplementedError(
                        "Creating a form is not supported in CognateExcelParser"
                    )


def excel_parser_from_dialect(
    output_dataset: pycldf.Wordlist, dialect: t.NamedTuple, cognate: bool
) -> t.Type[ExcelParser]:
    Row: t.Type[RowObject]
    Parser: t.Type[ExcelParser]
    if cognate:
        Row = CogSet
        Parser = ExcelCognateParser
    else:
        Row = Concept
        Parser = ExcelParser
    top = len(dialect.lang_cell_regexes) + 1
    # prepare cellparser
    row_header = []
    for row_regex in dialect.row_cell_regexes:
        match = re.fullmatch(row_regex, "", re.DOTALL)
        # TODO: when trying to raise a ValueError due to row_regexes not matching with the cell content,
        # I modify one of the regexes so that it does not match with any content of the cell.
        # Thus it doesn't match with '' either. The match object is None, and an AttributeError is raised.
        # What to do about this case?
        if match:
            row_header += list(match.groupdict().keys()) or [None]
        else:
            row_header += [None]
    initialized_cell_parser = getattr(cell_parsers, dialect.cell_parser["name"])(
        output_dataset,
        element_semantics=dialect.cell_parser["cell_parser_semantics"],
        separation_pattern=rf"([{''.join(dialect.cell_parser['form_separator'])}])",
        variant_separator=dialect.cell_parser["variant_separator"],
        add_default_source=dialect.cell_parser.get("add_default_source"),
    )

    class SpecializedExcelParser(Parser):
        def __init__(
            self,
            output_dataset: pycldf.Dataset,
            row_type=Row,
        ) -> None:
            super().__init__(
                output_dataset=output_dataset,
                top=top,
                row_type=Row,
                row_header=row_header,
                cellparser=cell_parsers.CellParser,
                check_for_match=dialect.check_for_match,
                check_for_row_match=dialect.check_for_row_match,
                check_for_language_match=dialect.check_for_language_match,
            )
            self.cell_parser = initialized_cell_parser
            self.db.empty_cache()

        def language_from_column(self, column: t.List[openpyxl.cell.Cell]) -> Language:
            """Parse the row, according to regexes from the metadata.

            Raises
            ======
            ValueError: When the cell cannot be parsed with the specified regex.

            """
            d: t.Dict[str, str] = {}
            for cell, cell_regex, comment_regex in zip(
                column, dialect.lang_cell_regexes, dialect.lang_comment_regexes
            ):
                if cell.value:
                    match = re.fullmatch(cell_regex, cell.value.strip(), re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match "
                            f"for {cell_regex}, but found {cell.value}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v
                if cell.comment:
                    match = re.fullmatch(comment_regex, cell.comment.content, re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match "
                            f"for {comment_regex}, but found {cell.comment.content}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v

            c_l_id = self.db.dataset["LanguageTable", "id"].name
            c_l_name = self.db.dataset["LanguageTable", "name"].name
            if c_l_id not in d:
                d[c_l_id] = string_to_id(d[c_l_name])
            return Language(d)

        def properties_from_row(self, row: t.List[openpyxl.cell.Cell]) -> Row:
            """Parse the row, according to regexes from the metadata.

            Raises
            ======
            ValueError: When the cell cannot be parsed with the specified regex.

            """
            d: t.Dict[str, str] = {}
            for cell, cell_regex, comment_regex in zip(
                row, dialect.row_cell_regexes, dialect.row_comment_regexes
            ):
                if cell.value:
                    match = re.fullmatch(cell_regex, cell.value.strip(), re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match "
                            f"for {cell_regex}, but found {cell.value}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v
                if cell.comment:
                    match = re.fullmatch(comment_regex, cell.comment.content, re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match for "
                            f"{comment_regex}, but found {cell.comment.content}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v

            return Row(d)

    return SpecializedExcelParser


def load_dataset(
    metadata: Path,
    lexicon: t.Optional[str],
    cognate_lexicon: t.Optional[str] = None,
    status_update: t.Optional[str] = None,
    logger: logging.Logger = cli.logger,
):
    # logging.basicConfig(filename="warnings.log")
    dataset = pycldf.Dataset.from_metadata(metadata)
    # load dialect from metadata
    try:
        dialect = argparse.Namespace(
            **dataset.tablegroup.common_props["special:fromexcel"]
        )
    except KeyError:
        dialect = None

    if not lexicon and not cognate_lexicon:
        raise argparse.ArgumentError(
            None,
            "At least one of WORDLIST and COGNATESETS excel files must be specified",
        )
    if lexicon:
        # load dialect from metadata
        if dialect:
            try:
                EP = excel_parser_from_dialect(dataset, dialect, cognate=False)
            except (AttributeError, KeyError) as err:
                field = re.match(r".*?'(.+?)'.+?'(.+?)'$", str(err)).group(2)
                logger.warning(
                    f"User-defined format specification in the json-file was missing the key {field}, "
                    f"falling back to default parser"
                )
                EP = ExcelParser
        else:
            logger.warning(
                "User-defined format specification in the json-file was missing, falling back to default parser"
            )
            EP = ExcelParser
            # The Intermediate Storage, in a in-memory DB (unless specified otherwise)
        # add Status_Column if not existing
        if status_update:
            add_status_column_to_table(dataset=dataset, table_name="FormTable")
        EP = EP(dataset, row_type=Concept)

        EP.db.empty_cache()

        lexicon_wb = openpyxl.load_workbook(lexicon).active
        EP.parse_cells(lexicon_wb, status_update=status_update)
        EP.db.write_dataset_from_cache()

    # load cognate dataset if provided by metadata
    if cognate_lexicon:
        if dialect:
            try:
                ECP = excel_parser_from_dialect(
                    dataset, argparse.Namespace(**dialect.cognates), cognate=True
                )
            except (AttributeError, KeyError) as err:
                field = re.match(r".*?'(.+?)'.+?'(.+?)'$", str(err)).group(2)
                logger.warning(
                    f"User-defined format specification in the json-file was missing the key {field}, "
                    f"falling back to default parser"
                )
                ECP = ExcelCognateParser
        else:
            logger.warning(
                "User-defined format specification in the json-file was missing, falling back to default parser"
            )
            ECP = ExcelCognateParser
        # add Status_Column if not existing
        if status_update:
            add_status_column_to_table(dataset=dataset, table_name="CognateTable")
        ECP = ECP(dataset, row_type=CogSet)
        ECP.db.cache_dataset()
        for sheet in openpyxl.load_workbook(cognate_lexicon).worksheets:
            ECP.parse_cells(sheet, status_update=status_update)
        ECP.db.write_dataset_from_cache()


if __name__ == "__main__":
    import pycldf

    parser = cli.parser(
        __package__ + "." + Path(__file__).stem,
        description="Imports a dataset from an excel file into CLDF. "
        "The import is configured by a special key in the metadata file, check "
        "./test/data/cldf/smallmawetiguarani/Wordlist-metadata.json for examples.",
    )
    parser.add_argument(
        "wordlist",
        nargs="?",
        default=None,
        help="Path to an Excel file containing the dataset",
        metavar="EXCEL",
    )
    parser.add_argument(
        "--cogsets",
        type=Path,
        default=None,
        help="Path to an optional second Excel file containing cogsets and cognate judgements",
        metavar="COGSET_EXCEL",
    )
    parser.add_argument(
        "--status-update",
        type=str,
        default="initial import",
        help="Text written to Status_Column. Set to 'None' for no status update. "
        "(default: initial import)",
    )
    args = parser.parse_args()
    logger = cli.setup_logging(args)

    if args.status_update == "None":
        args.status_update = None
    load_dataset(
        args.metadata, args.wordlist, args.cogsets, args.status_update, logger=logger
    )
