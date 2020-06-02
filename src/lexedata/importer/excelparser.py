# -*- coding: utf-8 -*-
import openpyxl as op


class ExcelParser():

    def __init__(self, sheet: op.worksheet.worksheet.Worksheet,
                 x_offset: int, y_offset: int, values_only=False) -> None:
        x = x_offset
        y = y_offset
        self.sheet = sheet
        self.iter_x_axis = self.sheet.iter_cols(min_row=1, max_row=y, min_col=x)
        self.iter_y_axis = self.sheet.iter_rows(min_row=1+y, max_col=x)  # iterates over rows with concepts
        self.iter_forms = self.sheet.iter_rows(min_row=1+y, min_col=1+x)  # iterates over rows with forms
        self.values_only = values_only

    def yield_x_axis(self) -> tuple[op.cell.cell.Cell]:
        for row in self.iter_x_axis:
            if self.values_only:
                yield (cell.value for cell in row)
            else:
                yield (cell for cell in row)

    def yield_y_axis(self) -> tuple[op.cell.cell.Cell]:
        for row in self.iter_y_axis:
            if self.values_only:
                yield (cell.value for cell in row)
            else:
                yield (cell for cell in row)

    def yield_forms(self) -> tuple[op.cell.cell.Cell]:
        for cell in self.iter_z_axis:
            if self.values_only:
                yield cell.value
            else:
                yield cell


class ExcelParserCognate(ExcelParser):
    pass

