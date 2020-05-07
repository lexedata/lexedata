# -*- coding: utf-8 -*-
import openpyxl as op
import csv
from pathlib import Path
<<<<<<< ours
from lexedata.importer.objects import *
from lexedata.importer.cellparser import *
from lexedata.importer.database import LEXICAL_ORIGIN, COGNATE_ORIGIN, DIR_DATA


def init_lan(dir_path, iter_lan, lan_dict):
    # create iterator over excel cells
    with (dir_path / "language_init.csv").open("w", encoding="utf8", newline="") as lanout:
        header_lan = ["id",
                           "name", "curator", "comment"]
        lanout = csv.DictWriter(lanout, header_lan, extrasaction="ignore", quotechar='"',
                                quoting=csv.QUOTE_MINIMAL)
        lanout.writeheader()
        for language in iter_lan:
            if language[0].value is None:

                l = Language.from_column(language)
                lan_dict[language[0].column_letter] = l.id
                lanout.writerow(l)


def init_con_form(dir_path, con_iter, form_iter, lan_dict, wb):

    with (dir_path / "form_init.csv").open( "w", encoding="utf8", newline="") as formsout, \
            (dir_path / "concept_init.csv").open( "w", encoding="utf8", newline="") as conceptsout:

        header_concepts = ["id",
                           "set", "english", "english_Strict", "spanish", "portuguese", "french"]
        concsv = csv.DictWriter(conceptsout, header_concepts, extrasaction="ignore", quotechar='"',
                                quoting=csv.QUOTE_MINIMAL)
        concsv.writeheader()

        header_forms = ["id", "language_id",
                        "phonemic", "phonetic", "orthographic", "Variants", "form_comment", "source",
                        "procedural_comment", "procedural_comment_concept", "concept_id"]

        formscsv = csv.DictWriter(formsout, header_forms, extrasaction="ignore", quotechar='"',
                                  quoting=csv.QUOTE_MINIMAL)
        formscsv.writeheader()

        for row_forms, row_con in zip(form_iter, con_iter):

            concept_cell = Concept.from_default_excel(row_con)
            concsv.writerow(concept_cell)

            for f_cell in row_forms:
                if f_cell.value:

                    # get corresponding language_id to column
                    this_lan_id = lan_dict[f_cell.column_letter]

                    for f_ele in CellParser(f_cell):

                        form_cell = Form.create_form(f_ele, this_lan_id, f_cell, concept_cell)
                        formscsv.writerow(form_cell)


def initialize_lexical(dir_path, lan_dict,
                       file=LEXICAL_ORIGIN):

    wb = op.load_workbook(filename=file)
    sheets = wb.sheetnames
    wb = wb[sheets[0]]
    iter_forms = wb.iter_rows(min_row=3, min_col=7)  # iterates over rows with forms
    iter_concept = wb.iter_rows(min_row=3, max_col=6)  # iterates over rows with concepts
    iter_lan = wb.iter_cols(min_row=1, max_row=2, min_col=7, max_col=44)

    init_lan(dir_path, iter_lan, lan_dict)
    init_con_form(dir_path, iter_concept, iter_forms, lan_dict, wb)


def cogset_cognate(cogset_iter, cog_iter, lan_dict, wb, cogsetcsv, cogcsv):

    for cogset_row, cog_row in zip(cogset_iter, cog_iter):
        if not cogset_row[1].value:
            continue
        if cogset_row[1].value.isupper():
            cogset = CogSet.from_excel(cogset_row)
            cogsetcsv.writerow(cogset)

            for f_cell in cog_row:
                if f_cell.value:
                    # get corresponding language_id to column
                    this_lan_id = lan_dict[f_cell.column_letter]

                    for f_ele in CogCellParser(f_cell):
                        cog = Cognate.from_excel(f_ele, this_lan_id, f_cell, cogset)
                        cogcsv.writerow(cog)
        # line not to be processed
        else:
            continue


def initialize_cognate(dir_path, lan_dict,
                 file=COGNATE_ORIGIN):

    cogout = (dir_path / "cog_init.csv").open("w", encoding="utf8", newline="")
    cogsetout = (dir_path / "cogset_init.csv").open("w", encoding="utf8", newline="")

    header_cog = ["id", "cog_set_id", "form_id", "language_id",
                  "cognate_comment", "phonemic", "phonetic", "orthographic", "procedural_comment", "source"]
    cogcsv = csv.DictWriter(cogout, header_cog, extrasaction="ignore", quotechar='"',
                            quoting = csv.QUOTE_MINIMAL)
    cogcsv.writeheader()

    header_cogset = ["id", "Set", "Description"]

    cogsetcsv = csv.DictWriter(cogsetout, header_cogset, extrasaction="ignore", quotechar='"',
                               quoting = csv.QUOTE_MINIMAL)
    cogsetcsv.writeheader()
    wb = op.load_workbook(filename=file)
    try:
        for sheet in wb.sheetnames:
            print(sheet+"\n\n")
            ws = wb[sheet]
            iter_cog = ws.iter_rows(min_row=3, min_col=5, max_col=42)  # iterates over rows with forms
            iter_congset = ws.iter_rows(min_row=3, max_col=4)  # iterates over rows with concepts
            cogset_cognate(iter_congset, iter_cog, lan_dict, ws, cogsetcsv, cogcsv)
    except KeyError:
        pass

    cogout.close()
    cogsetout.close()

if __name__ == "__main__":
    lan_dict = dict()
    dir_path = DIR_DATA
    initialize_lexical(dir_path, lan_dict)
    initialize_cognate(dir_path, lan_dict)