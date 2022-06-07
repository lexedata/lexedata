import typing as t
from collections import defaultdict
from pathlib import Path

import attr
import lexedata.cli as cli
import openpyxl
import pycldf
from lexedata.edit.add_status_column import add_status_column_to_table
from lexedata.importer.excel_matrix import DB
from lexedata.types import Form, KeyKeyDict
from lexedata.util import normalize_string, string_to_id, ensure_list
from lexedata.util.excel import clean_cell_value, normalize_header
from tabulate import tabulate

try:
    from typing import Literal
except ImportError:
    from typing_extensions import Literal


@attr.s(auto_attribs=True)
class ImportLanguageReport:
    is_new_language: bool = False
    new: int = 0
    existing: int = 0
    skipped: int = 0
    concepts: int = 0

    def __iadd__(self, other: "ImportLanguageReport"):
        self.is_new_language |= other.is_new_language
        self.skipped += other.skipped
        self.new += other.new
        self.existing += other.existing
        self.concepts += other.concepts
        return self

    def __call__(self, name: str) -> t.Tuple[str, int, int, int, int]:
        return (
            ("(new) " if self.is_new_language else "") + name,
            self.new,
            self.existing,
            self.skipped,
            self.concepts,
        )


def get_headers_from_excel(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> t.Iterable[str]:
    return normalize_header(r for c, r in enumerate(next(sheet.iter_rows(1, 1))))


def import_data_from_sheet(
    sheet,
    sheet_header,
    language_id: str,
    implicit: t.Mapping[Literal["languageReference", "id", "value"], str] = {},
    concept_column: t.Tuple[str, str] = ("Concept_ID", "Concept_ID"),
    skip_if_questionmark: t.Container[str] = set(),
) -> t.Iterable[Form]:
    row_iter = sheet.iter_rows()

    # TODO?: compare header of this sheet to format of given dataset process
    # row. Maybe unnecessary. In any case, do not complain about the unused
    # variable.
    header = next(row_iter)  # noqa: F841

    assert (
        concept_column[1] in sheet_header
    ), f"Could not find concept column {concept_column[1]} in your excel sheet {sheet.title}."

    for row in row_iter:
        data = Form({k: clean_cell_value(cell) for k, cell in zip(sheet_header, row)})
        for s in skip_if_questionmark:
            if data[s] == "?":
                data = {}
        if not any(data.values()):
            continue
        if "value" in implicit:
            data[implicit["value"]] = "\t".join(map(str, data.values()))
        concept_entry = data.pop(concept_column[1])
        data[concept_column[0]] = concept_entry
        if "id" in implicit:
            data[implicit["id"]] = None
        if "languageReference" in implicit:
            data[implicit["languageReference"]] = language_id
        yield data


def read_single_excel_sheet(
    dataset: pycldf.Dataset,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    logger: cli.logging.Logger = cli.logger,
    match_form: t.Optional[t.List[str]] = None,
    entries_to_concepts: t.Mapping[str, str] = KeyKeyDict(),
    concept_column: t.Optional[str] = None,
    language_name_column: t.Optional[str] = None,
    ignore_missing: bool = False,
    ignore_superfluous: bool = False,
    status_update: t.Optional[str] = None,
    missing_concepts: t.Set[str] = set(),
) -> t.Mapping[str, ImportLanguageReport]:
    report: t.Dict[str, ImportLanguageReport] = defaultdict(ImportLanguageReport)

    concept_columns: t.Tuple[str, str]
    if concept_column is None:
        concept_columns = (
            dataset["FormTable", "parameterReference"].name,
            dataset["FormTable", "parameterReference"].name,
        )
    else:
        concept_columns = (
            dataset["FormTable", "parameterReference"].name,
            concept_column,
        )

    db = DB(dataset)
    db.cache_dataset()
    # required cldf fields of a form
    c_f_id = db.dataset["FormTable", "id"].name
    c_f_language = db.dataset["FormTable", "languageReference"].name
    c_f_form = db.dataset["FormTable", "form"].name
    try:
        c_f_value = db.dataset["FormTable", "value"].name
    except KeyError:
        c_f_value = None
        logger.warning(
            "Your metadata file does not specify a #value column (usually called Value) to store the forms as given in the source. Consider adding it to your FormTable."
        )
    c_f_concept = db.dataset["FormTable", "parameterReference"].name
    if not match_form:
        match_form = [c_f_form, c_f_language]
    if not db.dataset["FormTable", c_f_concept].separator:
        logger.warning(
            "Your metadata does not allow polysemous forms. According to your specifications, "
            "identical forms with different concepts will always be considered homophones, not a single "
            "polysemous form. To include polysemous forms, add a separator to your FormTable #parameterReference "
            "in the Metadata.json To find potential polysemies, run lexedata.report.list_homophones."
        )
        match_form.append(c_f_concept)
    else:
        if c_f_concept in match_form:
            logger.info(
                "Matching by concept enabled: To find potential polysemies, run lexedata.report.list_homophones."
            )

    sheet_header = get_headers_from_excel(sheet)
    form_header = list(db.dataset["FormTable"].tableSchema.columndict.keys())

    # These columns don't need to be given, we can infer them from the sheet title and from the other data:
    implicit: t.Dict[Literal["languageReference", "id", "value"], str] = {}
    if c_f_language not in sheet_header:
        implicit["languageReference"] = c_f_language
    if c_f_id not in sheet_header:
        implicit["id"] = c_f_id
    if c_f_value is not None and c_f_value not in sheet_header:
        implicit["value"] = c_f_value

    found_columns = set(sheet_header) - {concept_column} - set(implicit.values())
    expected_columns = set(form_header) - {c_f_concept} - set(implicit.values())
    if language_name_column:
        expected_columns = expected_columns - {c_f_language}
        expected_columns.add(language_name_column)

    if not found_columns >= expected_columns:
        if ignore_missing:
            logger.info(
                f"Your Excel sheet {sheet.title} is missing columns {expected_columns - found_columns}. "
                f"For the newly imported forms, these columns will be left empty in the dataset."
            )
        else:
            raise ValueError(
                f"Your Excel sheet {sheet.title} is missing columns {expected_columns - found_columns}. "
                f"Clean up your data, or use --ignore-missing-columns to import anyway and leave these "
                f"columns empty in the dataset for the newly imported forms."
            )
    if not found_columns <= expected_columns:
        if ignore_superfluous:
            logger.info(
                f"Your Excel sheet {sheet.title} contained unexpected columns "
                f"{found_columns - expected_columns}. These columns will be ignored."
            )
        else:
            raise ValueError(
                f"Your Excel sheet {sheet.title} contained unexpected columns "
                f"{found_columns - expected_columns}. Clean up your data, or use "
                f"--ignore-superfluous-columns to import the data anyway and ignore these columns."
            )
    try:
        # Assume we have a language table
        c_l_name = db.dataset["LanguageTable", "name"].name
        c_l_id = db.dataset["LanguageTable", "id"].name
        language_name_to_language_id = {
            row[c_l_name]: row[c_l_id] for row in db.cache["LanguageTable"].values()
        }
    except pycldf.dataset.SchemaError:
        # Actually, there is no language table.
        language_name_to_language_id = KeyKeyDict()
        logger.info(
            "You have no LanguageTable, so I will have to assume that forms that already exist have the same Language IDs that is already in your FormTable."
        )
    # infer language from sheet data
    if language_name_column:

        def language_name_from_row(row):
            return language_name_to_language_id[
                row[sheet_header.index(language_name_column)].value
            ]

    elif c_f_language in sheet_header:

        def language_name_from_row(row):
            return row[sheet_header.index(c_f_language)].value

    else:

        def language_name_from_row(row):
            return normalize_string(sheet.title)

    row = ""
    for r in sheet.iter_rows(min_row=2):
        row = r
        break
    language_name = language_name_from_row(row)
    if language_name in language_name_to_language_id:
        language_id = language_name_to_language_id[language_name]
        report[language_id].is_new_language = False
    else:
        language_id = language_name
        report[language_id].is_new_language = True
        logger.warning(
            "I am adding forms for a new language %s, but I don't know how to add languages to your LanguageTable. Please ensure to add this language to the LanguageTable manually.",
            language_name,
        )
        logger.info(
            "To add the new language, you may want to add a row with ID %s to the LanguageTable, even if that does not fit the intended ID format, and then fix language IDs using lexedata.edit.simplify_ids --tables LanguageTable",
            language_id,
        )

    # read new data from sheet
    for form in cli.tq(
        import_data_from_sheet(
            sheet,
            sheet_header=sheet_header,
            implicit=implicit,
            language_id=language_id,
            concept_column=concept_columns,
            skip_if_questionmark={c_f_form},
        ),
        task=f"Parsing cells of sheet {sheet.title}",
        total=sheet.max_row,
    ):
        # else, look for candidates, link to existing form or add new form
        for item, value in form.items():
            try:
                sep = db.dataset["FormTable", item].separator
            except KeyError:
                continue
            if sep is None:
                continue
            form[item] = value.split(sep)

        # if concept not in dataset, don't add form
        concept_entries = ensure_list(form[c_f_concept])
        concepts = [
            entries_to_concepts.get(concept_entry) for concept_entry in concept_entries
        ]
        if not any(concepts):
            logger.warning(
                f"The concept(s) {concept_entries} could not be found. Please add them to the concepts.csv file manually. "
                f"The corresponding form {form[c_f_form]} was skipped and not imported."
            )
            missing_concepts.update(concept_entries)
            report[language_id].skipped += 1
            continue
        elif not all(concepts):
            missing = {
                c for c, mapping in zip(concept_entries, concepts) if mapping is None
            }
            missing_concepts.update(missing)
            logger.warning(
                f"The concept(s) {missing} could not be found. Please add them to the concepts.csv file manually. "
                f"The corresponding links to form {form[c_f_form]} were not imported, but they *should* be just added in a later run after you have created the concepts."
            )

        if db.dataset["FormTable", c_f_concept].separator:
            form[c_f_concept] = concepts
        else:
            (form[c_f_concept],) = concepts

        form_candidates = db.find_db_candidates(form, match_form)
        if form_candidates:
            new_concept_added = False
            for form_id in form_candidates:
                try:
                    logger.info(
                        f"Form {form[c_f_value]} was already in dataset. I have not added it again."
                    )
                except KeyError:
                    logger.info(
                        f"Form {form[c_f_form]} '{form[c_f_concept]}' in Language {form[c_f_language]} was already in dataset. I have not added it again."
                    )

                if db.dataset["FormTable", c_f_concept].separator:
                    for new_concept in form[c_f_concept]:
                        if (
                            new_concept
                            not in db.cache["FormTable"][form_id][c_f_concept]
                        ):
                            db.cache["FormTable"][form_id][c_f_concept].append(
                                new_concept
                            )
                            logger.info(
                                f"New form-concept association: Concept {form[c_f_concept]} was added to existing form "
                                f"{form_id}. If this was not intended "
                                f"(because it is a homophonous form, not a polysemy), "
                                f"you need to manually remove that concept from the old form in forms.csv "
                                f"and create a separate new form. If you want to treat identical forms "
                                f"as homophones in general, add  "
                                f"--match-forms={' '.join(match_form)}, "
                                f"{db.dataset['FormTable', 'parameterReference']} "
                                f"when you run this script."
                            )
                            new_concept_added = True
                break

            if new_concept_added:
                report[language_id].concepts += 1
            else:
                report[language_id].existing += 1
        else:
            # we land here after the break and keep adding existing forms to the dataset just with integer in id +1
            form[c_f_language] = language_id
            if "id" in implicit:
                # TODO: check for type of form id column
                form_concept = form[c_f_concept]
                concept_reference = (
                    form_concept[0] if isinstance(form_concept, list) else form_concept
                )
                form[c_f_id] = string_to_id(f"{form[c_f_language]}_{concept_reference}")
            db.make_id_unique(form)
            if status_update:
                form["Status_Column"] = status_update
            db.insert_into_db(form)
            report[language_id].new += 1
    # write to cldf
    db.write_dataset_from_cache()
    return report


def add_single_languages(
    metadata: Path,
    sheets: t.Iterable[openpyxl.worksheet.worksheet.Worksheet],
    match_form: t.Optional[t.List[str]],
    concept_name: t.Optional[str],
    language_name: t.Optional[str],
    ignore_missing: bool,
    ignore_superfluous: bool,
    status_update: t.Optional[str],
    logger: cli.logging.Logger,
    missing_concepts: t.Set[str] = set(),
) -> t.Mapping[str, ImportLanguageReport]:
    if status_update == "None":
        status_update = None
    # initiate dataset from meta
    try:
        dataset = pycldf.Dataset.from_metadata(metadata)
    except FileNotFoundError:
        cli.Exit.FILE_NOT_FOUND(
            "No cldf metadata found, if you have no metadata, "
            "export to csv from your excel file and run lexedata.edit.add_metadata"
        )
    # create concept mapping
    concepts: t.Mapping[str, str]
    try:
        cid = dataset["ParameterTable", "id"].name
        if concept_name is None:
            concepts = {c[cid]: c[cid] for c in dataset["ParameterTable"]}
            concept_column = dataset["FormTable", "parameterReference"].name
        else:
            name = dataset["ParameterTable", "name"].name
            concepts = {c[name]: c[cid] for c in dataset["ParameterTable"]}
            concept_column = concept_name
    except (KeyError, FileNotFoundError) as err:
        if isinstance(err, KeyError):
            logger.warning(
                "Did not find a well-formed ParameterTable. Importing all forms independent of concept"
            )
        elif isinstance(err, FileNotFoundError):
            logger.warning(
                f"Did not find {dataset['ParameterTable'].url.string}. "
                f"Importing all forms independent of concept"
            )
        # set concept_column
        concepts = KeyKeyDict()
        if concept_name:
            concept_column = concept_name
        else:
            concept_column = dataset["FormTable", "parameterReference"].name
    # add Status_Column if not existing and status_update given
    if status_update:
        add_status_column_to_table(dataset=dataset, table_name="FormTable")
    report: t.Dict[str, ImportLanguageReport] = defaultdict(ImportLanguageReport)
    # import all selected sheets
    for sheet in sheets:
        for lang, subreport in read_single_excel_sheet(
            dataset=dataset,
            sheet=sheet,
            logger=logger,
            match_form=match_form,
            entries_to_concepts=concepts,
            concept_column=concept_column,
            language_name_column=language_name,
            ignore_missing=ignore_missing,
            ignore_superfluous=ignore_superfluous,
            status_update=status_update,
            missing_concepts=missing_concepts,
        ).items():
            report[lang] += subreport
    return report


def parser():
    parser = cli.parser(
        __package__ + "." + Path(__file__).stem,
        description="Import forms and associated metadata from an excel file to a cldf dataset.",
    )
    parser.add_argument(
        "excel",
        type=openpyxl.load_workbook,
        help="The Excel file to parse",
        metavar="EXCEL",
    )
    parser.add_argument(
        "--concept-name",
        type=str,
        help="Excel column to interpret as concept names "
        "By default, it is assumed that the #parameterReference column, usually named 'Concept_ID' "
        "or similar, matches the IDs of the concept. Use this "
        "switch if instead of concept IDs you have concept Names in the excel file instead.",
        metavar="CONCEPT-COLUMN",
    )
    parser.add_argument(
        "--language-name",
        type=str,
        help="Excel column to interpret as language names "
        "By default, it is assumed that the #languageReference column, usually named 'Language_ID' "
        "or similar, matches the IDs of the language. If no Language_ID appears in the sheet header,"
        "the language name will be inferred by the sheet title. Use this "
        "switch if instead of language IDs you have language Names in the excel file instead.",
        metavar="LANGUAGE-COLUMN",
    )
    parser.add_argument(
        "--sheets",
        type=str,
        nargs="+",
        metavar="SHEET_NAME",
        help="Sheets to parse. (default: all sheets)",
    )
    parser.add_argument(
        "--match-form",
        "-f",
        type=str,
        nargs="+",
        default=[],
        metavar="COLUMN_NAME",
        help="Forms are considered identical if all columns passed to -f/--match-form are identical",
    )
    parser.add_argument(
        "--ignore-superfluous-columns",
        "-s",
        action="store_true",
        default=False,
        help="Ignore columns in the Excel table which are not in the dataset",
    )
    parser.add_argument(
        "--ignore-missing-columns",
        "-m",
        action="store_true",
        default=False,
        help="Ignore columns missing from the Excel table compared to the dataset",
    )
    parser.add_argument(
        "--exclude-sheet",
        "-x",
        type=str,
        nargs="*",
        default=[],
        metavar="SHEET_NAME",
        help="Sheets not to parse. Does not affect sheets explicitly added using --sheet.",
    )
    parser.add_argument(
        "--status-update",
        type=str,
        default="new import",
        help="Text written to Status_Column. Set to 'None' for no status update. "
        "(default: new import)",
    )
    parser.add_argument(
        "--report",
        action="store_true",
        default=False,
        help="Prints report of newly added forms",
    )
    parser.add_argument(
        "--list-missing-concepts",
        action="store_true",
        default=False,
        help="In the end, list all concepts that were not found.",
    )
    return parser


if __name__ == "__main__":  # pragma: no cover
    args = parser().parse_args()
    logger = cli.setup_logging(args)

    if not args.sheets:
        sheets = [
            sheet for sheet in args.excel if sheet.title not in args.exclude_sheet
        ]
        logger.info("No sheets specified explicitly. Parsing sheets: %s", args.sheets)
    else:
        sheets = [args.excel[s] for s in args.sheets]

    missing_concepts = set()
    report = add_single_languages(
        metadata=args.metadata,
        sheets=sheets,
        match_form=args.match_form,
        concept_name=args.concept_name,
        language_name=args.language_name,
        ignore_missing=args.ignore_missing_columns,
        ignore_superfluous=args.ignore_superfluous_columns,
        status_update=args.status_update,
        logger=logger,
        missing_concepts=missing_concepts,
    )
    if args.report:
        report_data = [report(language) for language, report in report.items()]
        print(
            tabulate(
                report_data,
                headers=[
                    "LanguageID",
                    "New forms",
                    "Existing forms",
                    "Skipped forms",
                    "New concept association",
                ],
                tablefmt="orgtbl",
            )
        )

    if args.list_missing_concepts:
        logger.info(
            "Now outputting the list of missing concepts. Check them for typos (and other minor mismatches) before adding them to the ParameterTable."
        )
        for concept in sorted(missing_concepts):
            print(concept)
