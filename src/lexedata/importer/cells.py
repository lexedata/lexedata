# -*- coding: utf-8 -*-
import openpyxl as op
import csv
import re
from collections import defaultdict
import argparse
import pycldf
import attr

from .exceptions import *
from .cellparser import CellParser
from .database import create_db_session, DatabaseObjectWithUniqueStringID, sa
from .objects import Language, Form, Concept, FormMeaningAssociation, Source

#lambda function for getting comment of excel cell if comment given
replace_none_comment = lambda x: x.content if x  else ""
comment_getter = lambda x: replace_none_comment(x.comment)

# replacing none values with ''
replace_none = lambda x: "" if x == None else x

# lambda function for getting comment of excel cell if comment given
replace_none_comment = lambda x: x.content if x else ""
comment_getter = lambda x: replace_none_comment(x.comment)

# functions for bracket checking
one_bracket = lambda opening, closing, str, nr: str[0] == opening and str[-1] == closing and \
                                                (str.count(opening) == str.count(closing) == nr)
comment_bracket = lambda str: str.count("(") == str.count(")")

def row_to_concept(conceptrow):
    # values of cell
    set, english, english_strict, spanish, portugese, french = [cell.value for cell in conceptrow]
    # comment of cell
    concept_comment = comment_getter(conceptrow[1])
    # create id
    concept_id = Concept.register_new_id(english)
    return Concept(ID=concept_id, set=set, english=english,
                   english_strict=english_strict, spanish=spanish,
                   portuguese=portugese, french=french,
                   concept_comment=concept_comment)



def language_from_column(column):
    name, curator = [cell.value for cell in column]
    id = Language.register_new_id(name)
    return Language(ID=id, name=name, curator=curator, comments="")




def main():
    parser = argparse.ArgumentParser(description="Load a Maweti-Guarani-style dataset into CLDF")
    parser.add_argument(
        "path", nargs="?",
        default="./Copy of TG_comparative_lexical_online_MASTER.xlsx",
        help="Path to an Excel file containing the dataset")
    parser.add_argument(
        "db", nargs="?",
        default="sqlite:///",
        help="Where to store the temp DB")
    parser.add_argument(
        "output", nargs="?",
        default="./",
        help="Directory to create the output CLDF wordlist in")
    parser.add_argument(
        "--debug-level", type=int, default=0,
        help="Debug level: Higher numbers are less forgiving")
    args = parser.parse_args()

    # The Input
    wb = op.load_workbook(filename=args.path)
    sheets = wb.sheetnames

    # The Intermediate Storage, in a in-memory DB
    session = create_db_session(args.db)

    # Start loading the data.
    iter_forms = wb[sheets[0]].iter_rows(min_row=3, min_col=7, max_col=44)  # iterates over rows with forms

    iter_concept = wb[sheets[0]].iter_rows(min_row=3, max_col=6)  # iterates over rows with concepts

    if True:

        # Collect languages
        languages = {}
        # TODO: The following line contains a magic number, can we just iterate until we are done?
        for lan_col in wb[sheets[0]].iter_cols(min_row=1, max_row=2, min_col=7, max_col=44):
            # iterate over language columns
            lan_cell = language_from_column(lan_col)
            session.add(lan_cell)

            # Switch to warnings module or something similar for this
            if args.debug_level >= 2:
                lan_cell.warn()
            elif args.debug_level == 1
                try:
                    lan_cell.warn()
                except LanguageElementError as E:
                    print(E)
                    continue

            languages[lan_cell.name] = lan_cell

        session.commit()

        for row_forms, row_con in zip(iter_forms, iter_concept):

            concept_cell = row_to_concept(row_con)
            session.add(concept_cell)

            for i, f_cell in enumerate(row_forms):
                if f_cell.value:

                    # you can access the cell's column letter and just add 1
                    this_lan_name = wb[sheets[0]][(f_cell.column_letter + "1")].value
                    this_lan_id = languages[this_lan_name].ID
                    f_comment = comment_getter(f_cell)

                    try:

                        for c_form, comment, source_id in CellParser(f_cell, this_lan_id):
                            source_id = Source.string_to_id(source_id)
                            source = session.query(Source).filter(
                                Source.ID == source_id).one_or_none()
                            if source is None:
                                source = Source(
                                    ID=source_id,
                                    genre="misc")
                                session.add(source)

                            c_form.sources.append(source)

                            already_existing = session.query(Form).filter(
                                Form.Language_ID == c_form.Language_ID,
                                Form.phonetic == c_form.phonetic,
                                Form.phonemic == c_form.phonemic,
                                Form.orthographic == c_form.orthographic,
                                Form.sources.contains(c_form.sources[0])).one_or_none()
                            if already_existing is None:
                                session.add(c_form)
                                form = c_form
                            else:
                                form = already_existing
                                existing_variants = form.variants.split(";")
                                new_variants = form.variants.split(";")
                                if set(new_variants) - set(existing_variants):
                                    print(
                                        "{:}: Variants {:} of form {:} were not mentioned earlier. Adding them to the form nonetheless".format(
                                            f_cell.coordinate,
                                            set(new_variants) - set(existing_variants),
                                            c_form,
                                            ))
                                    form.variants = ";".join(set(new_variants) | set(existing_variants))
                                if set(existing_variants) - set(new_variants):
                                    print(
                                        "{:}: Variants {:} of form {:} were not mentioned here.".format(
                                            f_cell.coordinate,
                                            set(existing_variants) - set(new_variants),
                                            c_form,
                                            ))
                                # FIXME: Compare the *set* of variant forms
                            if session.query(FormMeaningAssociation).filter(
                                    FormMeaningAssociation.form==form.ID,
                                    FormMeaningAssociation.concept==concept_cell.ID) is None:
                                session.add(FormMeaningAssociation(
                                    form=form.ID,
                                    concept=concept_cell.ID,
                                    comment=comment,
                                    procedural_comment=f_comment))
                            # Otherwise, the only change necessary will be the
                            # adding of a source, and that just happened.
                            session.commit()

                    except CellParsingError as err:
                        print("CellParsingError - somethings quite wrong")
                        print(err.message)

                    except FormCellError as err:
                        print("Error in cell {:} with content {!r:}: {:}".format(
                            f_cell.coordinate, f_cell.value, err))

    #################################################
    # create cldf
    session.commit()
    dataset = pycldf.Wordlist.from_metadata(
        "Wordlist-metadata.json",
    )

    if args.db.startswith("sqlite:///"):
        db_path = args.db[len("sqlite:///"):]
        if db_path == '':
            db_path = ':memory:'
    db = pycldf.db.Database(dataset, fname=db_path)
    # I have submitted https://github.com/cldf/pycldf/issues/105 concerning the
    # necessity to do this.
    db.tables[2].many_to_many["Concept_IDs"].name = "FormTable_ParameterTable"
    db.to_cldf("from_db/")



if False:
    # For debugging purposes:
    import importlib
    import lexedata.importer.database
    import lexedata.importer.objects
    __package__ = "lexedata.importer"
    import sys
    sys.argv = ["x", "../../../Copy of TG_comparative_lexical_online_MASTER.xlsx", "sqlite:///here.sqlite"]

    importlib.reload(lexedata.importer.database)
    importlib.reload(lexedata.importer.objects)

if __name__ == '__main__':
    main()
