# -*- coding: utf-8 -*-
from collections import defaultdict, OrderedDict
from pathlib import Path
import re

import openpyxl as op
import unidecode as uni

from lexedata.importer.objects import Form, CogSet, Language
from lexedata.importer.database import connect_db, DATABASE_ORIGIN, DIR_DATA
from lexedata.importer.exceptions import CellParsingError

WARNING = "\u26A0"
URL_BASE = r"https://myhost.com"

# ----------- Remark: Indices in excel are always 1-based. -----------


def string_cleaner(string):
    while re.search(r"[^A-z_0-9]", string):
        string = re.sub(r"[^A-z_0-9]", "", string)
    return string


def create_excel(out, db_path=DATABASE_ORIGIN):
    """
    creates excel with:
        columns: cogset(A) tags(B) languages(C-AN)
        rows: all cogsets in database
    :param out: path for created excel
    :param db_path: path to database
    :return:
    """
    # delete existing output file:
    out = Path(out)
    if out.exists():
        out.unlink()

    wb = op.Workbook()
    ws = wb.active

    session = connect_db(db_path)
    languages = session.query(Language).all()

    # Define the columns
    header = ["CogSet", "Tags"]
    lan_dict = dict()
    for col, lan in enumerate(languages, 3):
        # Excel indices are 1-based, not zero-based, so 3 is column C, as
        # intended.
        lan_dict[lan.id] = col
        header.append(lan.name)

    ws.append(header)

    # Iterate over all cognate sets, and prepare the rows.
    # Again, row_index 2 is indeed row 2, because indices are 1-based.
    row_index = 2
    for cogset in session.query(CogSet):

        # create cell for cogset in column A, add comment to excel cell if given description
        cogset_cell = ws.cell(row=row_index, column=1, value=cogset.id)
        # Transfer the cognateset comment to the Excel cell comment.
        if cogset.description != "":
            cogset_cell.comment = op.comments.Comment(cogset.description, "lexicaldata")
        # create cell for tag in column B
        tag_cell = ws.cell(row=row_index, column=2, value=cogset.set)
        # TODO should there be cogsets with no jdugements or is this just a temporal state?
        row_index = create_formcells_for_cogset(cogset, ws, row_index, lan_dict)

        # just for debugging
        #v = input()
        #if v == "s":
        #    break
    wb.save(filename=out)


def create_formcells_for_cogset(cogset, ws, row_index, lan_dict):
    """
    writes all forms for given cogset to excel
    modifying row_index to fit maximum of forms given for a language for this cogset
    returns modified row_index
    :param cogset:
    :param ws:
    :param row_index:
    :param lan_dict:
    :return:
    """
    # skip this row, if no judgements given
    if not cogset.judgements:
        row_index += 1
        return row_index

    # get sorted version of judgements for this cogset, language with maximum of forms first
    row_dict = defaultdict(list)
    for judgement in cogset.judgements:
        row_dict[judgement.language_id].append(judgement)
    ordered_dict = OrderedDict(sorted(row_dict.items(), key=lambda t: len(t[1]), reverse=True))
    # maximum of rows to be added
    maximum_cogset = len(ordered_dict[next(iter(ordered_dict.keys()))])
    for i in range(maximum_cogset):
        this_row = row_index + i
        for k, v in list(ordered_dict.items()):
            this_judgement = v.pop(0)
            # if it was last form, remove this language from dict
            if len(v) == 0:
                del ordered_dict[k]
            # create cell for this judgement
            create_formcell(this_judgement, ws, this_row, lan_dict[k])
    # increase row_index and return
    row_index += maximum_cogset

    return row_index


def create_formcell(judgement, ws, row, col):
    "Creates formcell for judgment in ws at row, col. With Return None"
    form = judgement.form
    cell_value = form_to_cell_value(form)
    form_cell = ws.cell(row=row, column=col, value=cell_value)
    if judgement.procedural_comment != "":
        comment = judgement.procedural_comment
        form_cell.comment = op.comments.Comment(comment, "lexicaldata")
    my_formid = string_cleaner(uni.unidecode(form.id))  # no illegal characters in URL
    link = "{}/{}".format(URL_BASE, my_formid)
    print(link)
    form_cell.hyperlink = link


def form_to_cell_value(form):
    "returns cell value of formcell: best transcription +  all translations"
    transcription = get_best_transcription(form)
    mywarning = ""
    translations = []

    # iterate over corresponding concepts
    for form_to_concept in form.toconcepts:
        if form_to_concept.procedural_comment != "":
            mywarning = WARNING
        translations.append(form_to_concept.concept.english)
    translations = " ".join(translations)

    return " ".join([transcription, translations, mywarning])


def get_best_transcription(form):
    if form.phonemic != "":
        return form.phonemic
    elif form.phonetic != "":
        return form.phonetic
    elif form.orthographic != "":
        return form.orthographic
    else:
        CellParsingError("empty values ''", form.id)


if __name__ == "__main__":
    out = DIR_DATA / "output.xlsx"
    create_excel(out)
