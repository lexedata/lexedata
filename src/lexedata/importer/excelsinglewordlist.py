import logging
import typing as t
from pathlib import Path
from collections import defaultdict
from tabulate import tabulate

import openpyxl
import pycldf

from lexedata.util import string_to_id, clean_cell_value, normalize_header, normalize_string
from lexedata.importer.fromexcel import DB
from lexedata.types import Form
from lexedata.enrich.add_status_column import add_status_column_to_table

try:
    from typing import Literal
except ImportError:
    from typing_extensions import Literal

logger = logging.getLogger(__file__)


class KeyKeyDict:
    def __getitem__(self, key):
        return key


def get_headers_from_excel(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> t.Iterable[str]:
    return normalize_header(r for c, r in enumerate(next(sheet.iter_rows(1, 1))))


def import_data_from_sheet(
    sheet,
    sheet_header,
    implicit: t.Mapping[Literal["languageReference", "id", "value"], str] = {},
    entries_to_concepts: t.Mapping[str, str] = KeyKeyDict(),
    concept_column: t.Tuple[str, str] = ("Concept_ID", "Concept_ID"),
) -> t.Iterable[Form]:
    row_iter = sheet.iter_rows()

    # TODO?: compare header of this sheet to format of given data set process
    # row. Maybe unnecessary. In any case, do not complain about the unused
    # variable.
    header = next(row_iter)  # noqa: F841

    assert (
        concept_column[1] in sheet_header
    ), f"Could not find concept column {concept_column[0]} in your excel sheet {sheet.title}."

    for row in row_iter:
        data = Form({k: clean_cell_value(cell) for k, cell in zip(sheet_header, row)})
        if "value" in implicit:
            data[implicit["value"]] = "\t".join(map(str, data.values()))
        concept_entry = data.pop(concept_column[1])
        data[concept_column[0]] = concept_entry
        # TODO: I moved this warning to read_single_excel_sheet in order to count the skipped forms
        # try:
        #     concept_entry = data.pop(concept_column[1])
        #     data[concept_column[0]] = entries_to_concepts[concept_entry]
        # except KeyError:
        #     logger.warning(
        #          f"Concept {concept_entry} was not found. Please add it to the concepts table manually. The corresponding form was ignored and not added to the dataset."
        #     )
        #     breakpoint()
        #     data[concept_column[0]] = concept_entry
        #    continue
        if "id" in implicit:
            data[implicit["id"]] = None
        if "languageReference" in implicit:
            data[implicit["languageReference"]] = normalize_string(sheet.title)
        yield data


def read_single_excel_sheet(
    dataset: pycldf.Dataset,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    match_form: t.Optional[t.List[str]] = None,
    entries_to_concepts: t.Mapping[str, str] = KeyKeyDict(),
    concept_column: t.Optional[str] = None,
    ignore_missing: bool = False,
    ignore_superfluous: bool = False,
    status_update: t.Optional[str] = None,
    report: t.Optional[t.Dict] = None,
):
    concept_columns: t.Tuple[str, str]
    if concept_column is None:
        concept_columns = (
            dataset["FormTable", "parameterReference"].name,
            dataset["FormTable", "parameterReference"].name,
        )
    else:
        concept_columns = (
            dataset["FormTable", "parameterReference"].name,
            concept_column,
        )
    db = DB(dataset)
    db.cache_dataset()
    # required cldf fields of a form
    c_f_id = db.dataset["FormTable", "id"].name
    c_f_language = db.dataset["FormTable", "languageReference"].name
    c_f_form = db.dataset["FormTable", "form"].name
    c_f_value = db.dataset["FormTable", "value"].name
    c_f_concept = db.dataset["FormTable", "parameterReference"].name
    if not match_form:
        match_form = [c_f_form]
        # TODO: adding the language ID as a default parameter for matching seems a bad idea to me.
        # The candidate form gets the sheet title as language id, in most cases this is not identical to the actual language id
        #match_form = [c_f_form, c_f_language]
    if not db.dataset["FormTable", c_f_concept].separator:
        match_form.append(c_f_concept)

    sheet_header = get_headers_from_excel(sheet)
    form_header = list(db.dataset["FormTable"].tableSchema.columndict.keys())

    # These columns don't need to be given, we can infer them from the sheet title and from the other data:
    implicit: t.Dict[Literal["languageReference", "id", "value"], str] = {}
    if c_f_language not in sheet_header:
        implicit["languageReference"] = c_f_language
    if c_f_id not in sheet_header:
        implicit["id"] = c_f_id
    if c_f_value not in sheet_header:
        implicit["value"] = c_f_value

    found_columns = set(sheet_header) - {concept_column} - set(implicit.values())
    expected_columns = set(form_header) - {c_f_concept} - set(implicit.values())
    if not found_columns >= expected_columns:
        message = f"Your Excel sheet {sheet.title} is missing columns {expected_columns - found_columns}."
        if ignore_missing:
            logger.warning(message)
        else:
            raise ValueError(message)
    if not found_columns <= expected_columns:
        message = f"Your Excel sheet {sheet.title} contained unexpected columns {found_columns - expected_columns}."
        if ignore_superfluous:
            logger.warning(message)
        else:
            raise ValueError(message)
    # check if language exist
    c_l_name = db.dataset["LanguageTable", "name"].name
    c_l_id = db.dataset["LanguageTable", "id"].name
    language_name_to_language_id = {
        row[c_l_name]: row[c_l_id] for row in db.cache["LanguageTable"].values()
    }
    language_name = sheet.title
    if language_name in language_name_to_language_id:
        language_id = language_name_to_language_id[language_name]
        lan_stat = 1
    else:
        language_id = language_name
        lan_stat = 0
    if isinstance(report, dict):
        report[language_id]["lan_stat"] = lan_stat
        report[language_id]["new"]
        report[language_id]["existing"]
        report[language_id]["skipped"]
        report[language_id]["concepts"]
    # read new data from sheet
    for form in import_data_from_sheet(
        sheet,
        sheet_header=sheet_header,
        implicit=implicit,
        entries_to_concepts=entries_to_concepts,
        concept_column=concept_columns,
    ):
        # if concept not in dataset, don't add form
        try:
            concept_entry = form[c_f_concept]
            entries_to_concepts[concept_entry]
        except KeyError:
            logger.warning(
                f"Concept {concept_entry} was not found. Please add it to the concepts table manually. The corresponding form was ignored and not added to the dataset."
            )
            if report:
                report[language_id]["skipped"] += 1
            continue
        # else, look for candidates, link to existing form or add new form
        for item, value in form.items():
            try:
                sep = db.dataset["FormTable", item].separator
            except KeyError:
                continue
            if sep is None:
                continue
            form[item] = value.split(sep)
        form_candidates = db.find_db_candidates(form, match_form)
        if form_candidates: # indented the for block so only if no form_candidates are found the new form is created
            new_concept_added = False
            for form_id in form_candidates:
                logger.info(f"Form {form[c_f_value]} was already in data set.")

                if db.dataset["FormTable", c_f_concept].separator:
                    # TODO: @Gereon, I think we should discuss these lines quickly
                    # most of the time we check if a concepts exists with entries_to_concept, but here we use the db.chache
                    for new_concept in form[c_f_concept]:
                        if new_concept not in db.cache["FormTable"][form_id][c_f_concept]:
                            db.cache["FormTable"][form_id][c_f_concept].append(new_concept)
                            logger.info(
                                f"Existing form {form_id} was added to concept {form[c_f_concept]}. "
                                f"If this was not intended (because it was a homophonous form, not a polysemy), "
                                f"you need to manually remove that concept "
                                f"from the old form and create a separate new form."
                            )
                            new_concept_added = True

                break
            if new_concept_added:
                if report:
                    report[language_id]["concepts"] += 1
            else:
                if report:
                    report[language_id]["existing"] += 1
        else:
            # we land here after the break and keep adding existing forms to the dataset just with integer in id +1
            form[c_f_language] = language_id
            if "id" in implicit:
                # TODO: check for type of form id column
                form_concept = form[c_f_concept]
                concept_reference = (
                    form_concept[0] if isinstance(form_concept, list) else form_concept
                )
                form[c_f_id] = string_to_id(f"{form[c_f_language]}_{concept_reference}")
            db.make_id_unique(form)
            if status_update:
                form["Status_Column"] = status_update
            db.insert_into_db(form)
            if report:
                report[language_id]["new"] += 1
    # write to cldf
    db.write_dataset_from_cache()


def add_single_languages(
    metadata: Path,
    excel: str,
    sheet: t.Optional[t.List[str]],
    match_form: t.Optional[t.List[str]],
    concept_name: t.Optional[str],
    ignore_missing: bool,
    ignore_superfluous: bool,
    exclude_sheet,
    verbose: bool,
    status_update: t.Optional[str],
    report: bool,
) -> None:
    if status_update == "None":
        status_update = None
    if verbose:
        logging.basicConfig(level=logging.INFO)
    if not sheet:
        sheets = [sheet for sheet in excel.sheetnames if sheet not in exclude_sheet]
        logging.warning("No sheets specified. Parsing sheets: %s", sheet)
    # initiate data set from meta data or csv depending on command line arguments
    if metadata:
        if metadata.name == "forms.csv":
            dataset = pycldf.Dataset.from_data(metadata)
        else:
            dataset = pycldf.Dataset.from_metadata(metadata)

    try:
        cid = dataset["ParameterTable", "id"].name
        if concept_name is None:
            concepts = {c[cid]: c[cid] for c in dataset["ParameterTable"]}
            concept_column = dataset["FormTable", "parameterReference"].name
        else:
            name = dataset["ParameterTable", "name"].name
            concepts = {c[name]: c[cid] for c in dataset["ParameterTable"]}
            concept_column = concept_name
    except KeyError:
        concepts = KeyKeyDict()
        concept_column = dataset["FormTable", "parameterReference"].name
    # add Status_Column if not existing and status_update given
    if status_update:
        add_status_column_to_table(dataset=dataset, table_name="FormTable")
    if report:
        report = defaultdict(lambda: defaultdict(int))
    else:
        report = None
    # import all selected sheets
    for sheet in sheets:
        read_single_excel_sheet(
            dataset=dataset,
            sheet=excel[sheet],
            match_form=match_form,
            entries_to_concepts=concepts,
            concept_column=concept_column,
            ignore_missing=ignore_missing,
            ignore_superfluous=ignore_superfluous,
            status_update=status_update,
            report=report,
        )
    if report:
        report_data = [
            [
                ("(new)" if bool(values["lan_stat"]) else "") + language,
                values["new"],
                values["existing"],
                values["skipped"],
                values["concepts"],
            ]
            for language, values in report.items()
        ]
        print(
            tabulate(
                report_data,
                headers=[
                    "LanguageID",
                    "New forms",
                    "Existing forms",
                    "Skipped forms",
                    "New concept reference",
                ],
                tablefmt="orgtbl",
            )
        )


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Add forms from Excel file to dataset")
    parser.add_argument(
        "excel", type=openpyxl.load_workbook, help="The Excel file to parse"
    )
    parser.add_argument(
        "--metadata",
        type=Path,
        default="Wordlist-metadata.json",
        help="Path to the JSON metadata file describing the dataset (default: ./Wordlist-metadata.json)",
    )
    parser.add_argument(
        "--concept-name",
        type=str,
        help="Column to interpret as concept names "
        "Default: assume the #parameterReference column, usually named 'Concept_ID' "
        "or similar, matches the IDs of the concept. Use this "
        "switch if you have concept Names in the wordlist instead.",
    )
    parser.add_argument(
        "--sheet", type=str, action="append", help="Sheets to parse. (default: all)"
    )
    parser.add_argument(
        "--match-form",
        type=str,
        nargs="*",
        default=[],
        help="Columns to match forms by",
    )
    parser.add_argument(
        "--ignore-superfluous-excel-columns",
        "->",
        action="store_true",
        default=False,
        help="Ignore columns in the Excel table which are not in the dataset",
    )
    parser.add_argument(
        "--ignore-missing-excel-columns",
        "-<",
        action="store_true",
        default=False,
        help="Ignore columns missing from the Excel table compared to the dataset",
    )
    parser.add_argument(
        "--exclude-sheet",
        "-x",
        type=str,
        nargs="*",
        default=[],
        help="Sheets not to parse",
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        default=False,
        help="Report existing forms",
    )
    parser.add_argument(
        "--status-update",
        type=str,
        default="new import",
        help="Text written to Status_Column. Set to 'None' for no status update. "
        "(default: new import)",
    )
    parser.add_argument(
        "--report",
        action="store_true",
        default=False,
        help="Prints report of newly added forms",
    )
    args = parser.parse_args()
    add_single_languages(
        metadata=args.metadata,
        excel=args.excel,
        sheet=args.sheet,
        match_form=args.match_form,
        concept_name=args.concept_name,
        ignore_missing=args.ignore_missing_excel_columns,
        ignore_superfluous=args.ignore_superfluous_excel_columns,
        exclude_sheet=args.exclude_sheet,
        verbose=args.verbose,
        status_update=args.status_update,
        report=args.report,
    )
