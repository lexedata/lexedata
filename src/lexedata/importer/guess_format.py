import os
import re
import shutil
import argparse
import typing as t
import unicodedata
from pathlib import Path
import platform

import csvw
import pycldf
import openpyxl

if platform.system() != "Windows":
    import readline
else:
    from pyreadline import Readline

    readline = Readline()

if os.name == "nt":

    def clear():
        os.system("cls")


else:

    def clear():
        return
        os.system("clear")


def select_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    sheets = wb.sheetnames
    if len(sheets) == 1:
        return wb.active
    for i, name in enumerate(sheets, 1):
        print(f"{i: 3d}) {name:}")
    sheet = input("Pick a sheet to work with (default: 0) >")

    for i, name in enumerate(sheets, 1):
        if str(i) == sheet.strip() or name.strip().startswith(sheet.strip()):
            return wb[name]
    else:
        print("That was incomprehensible. Press Ctrl+C to exit.")
        return select_sheet(wb)


def to_width(s: str, target_width: int):
    """Pad or trim a string to target width

    Width is in terms of monospaced spacing glyphs, not in terms of characters,
    to be used in tabular console views.

    >>> to_width("t", 3)
    't  '
    >>> to_width("test", 3)
    'te…'
    >>> to_width("txt", 3)
    'txt'
    >>> to_width("dolﬁns", 5)
    'dolf…'
    >>> to_width("error", 0)
    Traceback (most recent call last):
    ...
    ValueError: Target width must be positive

    """
    if target_width < 1:
        raise ValueError("Target width must be positive")
    s = unicodedata.normalize("NFKD", s)
    if width(s) <= target_width:
        return s + " " * (target_width - width(s))
    while width(s) > target_width - 1:
        s = s[:-1]
    if width(s) < target_width - 1:
        return s + "…" + " " * (target_width - width(s) - 1)
    else:
        return s + "…"


def width(value: t.Optional[str]) -> int:
    """Compute the decomposed width of a string

    Width is in terms of monospaced spacing glyphs, not in terms of characters,
    to be used in tabular console views. This function applies the NFKD first,
    so ligatures will be unpacked.

    >>> width("t")
    1
    >>> width("ê")
    1
    >>> width("ﬁ")
    2
    >>> width("eǩe")
    3
    >>> width("全角")
    4
    """
    s: str = unicodedata.normalize("NFKD", value or "")
    # There might be CJK characters of double width, or things like that?
    # That's why we use the sum construction, which can be extended to such
    # cases.
    return sum(
        (unicodedata.category(c) not in {"Mn", "Cf", "Cc"})
        + (unicodedata.east_asian_width(c) in {"W", "F"})
        for c in s
    )


WHITESPACE = r"\s*"


def to_regex(pattern: str) -> re.Pattern:
    """Convert a pattern to a regular expression with capturing groups.

    >>> l = to_regex("Language")
    >>> l.fullmatch("Maweti-Guarani ").groupdict()
    {'language': 'Maweti-Guarani'}

    >>> n = to_regex("Name (Abbreviation): Curator_Initials")
    >>> n.fullmatch("German (deu): MBO").groupdict()
    {'name': 'German', 'abbreviation': 'deu', 'curator_initials': 'MBO'}
    >>> n.fullmatch("German  (deu):MBO   ").groupdict()
    {'name': 'German', 'abbreviation': 'deu', 'curator_initials': 'MBO'}
    """
    elements = re.compile(r"(\w+)").split(pattern)
    for i in range(0, len(elements), 2):
        elements[i] = WHITESPACE.join(
            re.escape(s) for s in [""] + elements[i].split() + [""]
        )
        if elements[i] == WHITESPACE * 2:
            elements[i] = WHITESPACE
    for i in range(1, len(elements), 2):
        this = elements[i].lower()
        elements[i] = f"(?P<{this:}>.*?)"
    return re.compile("".join(elements), re.IGNORECASE)


def pretty(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    col_widths: t.List[int],
    top: int,
    left: int,
    tty_rows: int = 20,
) -> None:
    for r, row in enumerate(
        ws.iter_rows(max_row=min(tty_rows - 2, ws.max_row), max_col=len(col_widths))
    ):
        c1, c2 = col_widths[: left - 1], col_widths[left - 1 :]
        if r + 1 == top:
            print(
                "┿".join(["━" * c for c in c1]),
                "┿".join(["━" * c for c in c2]),
                sep="╋",
            )
        print(
            "│".join(to_width(c.value or "", l) for c, l in zip(row, c1)),
            "│".join(to_width(c.value or "", l) for c, l in zip(row[left - 1 :], c2)),
            sep="┃",
        )


CELL = re.compile(
    r"""
    ((target\W)?cell\W)?  # Maybe the user thought that was part of the syntax
    \W*                   # There may be leading whitespace
    \$?                   # The cell reference may be absolute
    (?P<col>[a-z]{1,3})   # The column has at most three letters
    \$?                   # The cell reference may be absolute
    (?P<row>[0-9]+)       # The row number
    \W*                   # There may be trailing whitespace
    """,
    re.VERBOSE | re.IGNORECASE,
)


def pretty_print_example_cells(example: t.List[openpyxl.cell.Cell]) -> None:
    i = 0
    for cell in example:
        value = cell.value or ""
        cell_content = "[{:}]".format("\n     ".join(value.split("\n")))
        i += 1
        print(f"{i:2d}:", cell_content)
        cell_comment = cell.comment.content if cell.comment else ""
        cell_comment = "".join(
            (to_width(c, 15) if width(c) > 15 else c)
            for c in SEP_C_RE.split(cell_comment)
        )
        cell_comment = "({:})".format(
            "\n     ".join(e for e in cell_comment.split("\n"))
        )
        i += 1
        print(f"{i:2d}:", cell_comment)


def understand(c: str, top: int, left: int) -> t.Optional[t.Tuple[int, int]]:
    """Parse user input
    >>> understand("cell c2", 3, 2)
    (2, 3)
    >>> understand("$G$3", 3, 2)
    (3, 7)
    >>> understand("d2", 3, 2)
    (2, 4)
    >>> understand("d", 3, 2)
    (4, 2)
    >>> understand("down", 3, 2)
    (4, 2)
    >>> understand("u l u", 3, 2)
    (1, 1)
    >>> understand("[u]p", 3, 2)
    (2, 2)
    """
    if re.fullmatch(r"\W*", c):
        return None
    elif CELL.fullmatch(c):
        # User entered target cell coordinates
        groups = CELL.fullmatch(c).groupdict()
        return openpyxl.utils.cell.coordinate_to_tuple(
            groups["col"].upper() + groups["row"]
        )
    elif re.fullmatch("[^a-z]*u[^a-z]*p?[^a-z]*", c):
        return (top - 1, left)
    elif re.fullmatch("[^a-z]*l[^a-z]*(eft)?[^a-z]*", c):
        return (top, left - 1)
    elif re.fullmatch("[^a-z]*d[^a-z]*(own)?[^a-z]*", c):
        return (top + 1, left)
    elif re.fullmatch("[^a-z]*r[^a-z]*(ight)?[^a-z]*", c):
        return (top, left + 1)
    elif re.fullmatch("[^a-z]*([uldr][^a-z]*)+", c):
        return (top + c.count("d") - c.count("u"), left + c.count("r") - c.count("l"))
    elif not re.fullmatch("[^a-z]*[h].*", c, re.IGNORECASE):
        print("Command not understood.")
    print(
        """Enter (ie. type and then press 'Enter') one of the following:
  • The Excel coordinates (eg. 'C2' or 'G3') of the top left cell of your data.
  • A single direction string ('up', 'left', 'down', or 'right') to shift the
    boundary between headers and data by one cell in that direction.
  • A sequence of one or more direction abbreviations ('u', 'drr') to shift
    the boundary by that.
  • 'H' or 'help' or anything starting with 'h' to display this help.
  • Nothing, just press 'Enter', to accept the current boundary."""
    )
    return understand(input(">"), top, left)


FAV_SEP = "[,:;–]|\n|--"
SEP_RE = re.compile(FAV_SEP)
SEP_C_RE = re.compile(f"({FAV_SEP:})")


def most_content(
    cells: t.Iterable[t.List[openpyxl.cell.Cell]],
) -> t.List[openpyxl.cell.Cell]:
    mcl = 0
    for col in cells:
        content_length = sum(
            len(cell.value or "")
            + ((len(SEP_RE.split(cell.comment.content)) * 8 + 8) if cell.comment else 0)
            for cell in col
        )
        if content_length > mcl:
            long_col = col
            mcl = content_length
    return long_col


LANGUAGE_PROPERTIES = {
    "ID": "http://cldf.clld.org/v1.0/terms.rdf#id",
    "Name": "http://cldf.clld.org/v1.0/terms.rdf#name",
    "Macroarea": "http://cldf.clld.org/v1.0/terms.rdf#macroarea",
    "Latitude": "http://cldf.clld.org/v1.0/terms.rdf#latitude",
    "Longitude": "http://cldf.clld.org/v1.0/terms.rdf#longitude",
    "Glottocode": "http://cldf.clld.org/v1.0/terms.rdf#glottocode",
    "ISO639P3code": "http://cldf.clld.org/v1.0/terms.rdf#iso639P3code",
    "Comment": "http://cldf.clld.org/v1.0/terms.rdf#comment",
}

CONCEPT_PROPERTIES = {
    "Name": "http://cldf.clld.org/v1.0/terms.rdf#name",
    "Gloss": "http://cldf.clld.org/v1.0/terms.rdf#gloss",
    "Description": "http://cldf.clld.org/v1.0/terms.rdf#description",
    "Part_Of_Speech": "http://cldf.clld.org/v1.0/terms.rdf#partOfSpeech",
    "Comment": "http://cldf.clld.org/v1.0/terms.rdf#comment",
}

COGNATESET_PROPERTIES = {
    "ID": "http://cldf.clld.org/v1.0/terms.rdf#id",
    "Description": "http://cldf.clld.org/v1.0/terms.rdf#description",
    "Source": "http://cldf.clld.org/v1.0/terms.rdf#source",
    "Comment": "http://cldf.clld.org/v1.0/terms.rdf#comment",
    "Source_Form_ID": "http://cldf.clld.org/v1.0/terms.rdf#sourceFormReference",
}

CLDF = {}
CLDF.update(LANGUAGE_PROPERTIES)
CLDF.update(CONCEPT_PROPERTIES)
CLDF.update(COGNATESET_PROPERTIES)


def completer(properties: t.Iterable[str] = CLDF):
    def complete(text, state):
        for cmd in properties:
            if cmd.lower().startswith(text.lower()):
                if not state:
                    return cmd
                else:
                    state -= 1

    return complete


# TODO: write cell_parser_semantics to metadata.json
def create_parsers(
    i: int,
    mapping: t.Mapping[str, str] = CLDF,
    # For testing, one might want to overload the input() fn with a mock,
    # eg. a dictionary's .get()
    input: t.Callable[[str], str] = input,
) -> t.Tuple[t.List[re.Pattern], t.List[re.Pattern]]:
    cell_parsers: t.List[re.Pattern] = []
    for j in range(1, i * 2 - 1):
        pattern = input(f"{j:2d} > ")
        if not pattern.strip():
            pattern = ".*"
        # FIXME: This is a bad heuristic to distinguish regexes from raw
        # patterns.
        if "(?P<" in pattern or "*" in pattern:
            # Pattern is a regex
            pattern = pattern.lower().replace("(?p", "(?P")
        else:
            # Pattern is a naïve string pattern
            pattern = to_regex(pattern).pattern
        # FIXME: Compile, unwrap, re-compile seems cumbersome, but how else
        # can we rename those groups to the CLDF names?
        for readable, cldf in mapping.items():
            name = "cldf_" + cldf.split("#")[-1]
            r = readable.lower()
            pattern = pattern.replace(f"(?P<{r:}>", f"(?P<{name:}>")
            pattern = pattern.replace(f"(?P={r:})", f"(?P={name:})")
        regex = re.compile(pattern, re.IGNORECASE)
        cell_parsers.append(regex)
    return [x for x in cell_parsers[::2]], [x for x in cell_parsers[1::2]]


def add_table_with_columns(
    table: str, column_names: t.Set[str], data: pycldf.Dataset
) -> None:
    """Add a table with the given columns to the dataset.

    If such a table already exists, only add the columns that do not exist
    yet.

    """
    delete = True
    try:
        data[table]
        delete = False
    except KeyError:
        data.add_component(table)
    columns = data[table].tableSchema.columns
    for c in range(len(columns) - 1, -1, -1):
        column = columns[c]
        expected_name = "cldf_{}".format(column.propertyUrl.uri.split("#")[-1].lower())
        if expected_name not in column_names and delete:
            del columns[c]
        else:
            column_names.remove(expected_name)
    for column_name in column_names:
        data.add_columns(
            table, column_name.replace("cldf_", "http://cldf.clld.org/v1.0/terms.rdf#")
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a custom lexical dataset parser and dataset metadata for a particular dataset"
    )
    parser.add_argument("excel", type=Path, help="The Excel file to inspect")
    parser.add_argument(
        "--metadata",
        type=Path,
        default="Wordlist-metadata.json",
        help="Path to the JSON metadata file describing the dataset (default: ./Wordlist-metadata.json)",
    )
    args = parser.parse_args()

    # STEP 1: Select a sheet
    # ======================
    wb = openpyxl.load_workbook(filename=args.excel)

    ws = select_sheet(wb)

    tty_columns, tty_rows = shutil.get_terminal_size()
    running = 0
    col_widths: t.List[int] = []
    for col in ws.iter_cols():
        w = min(15, max(len(cell.value or "") for cell in col))
        running += w + len("│")
        col_widths.append(w)
        if running >= tty_columns:
            overshoot = running - tty_columns
            col_widths[-1] -= overshoot
            break

    # STEP 2: Select the boundary between headers and data
    # ====================================================
    if ws.freeze_panes:
        top, left = openpyxl.utils.cell.coordinate_to_tuple(ws.freeze_panes)
    # TODO: There can be more guessers here
    else:
        # Fall back to one column and one row of headers
        top, left = (2, 2)

    while True:
        clear()
        pretty(ws, col_widths, top, left, tty_rows)
        command = input(
            "Enter a target cell or a direction (up/down/left/right) to move the data boundary. Enter [h]elp for more info. Press [Enter] to confirm. >"
        )
        tl = understand(command, top, left)
        if tl is None:
            break
        top, left = tl

    # STEP 3: Select the data type
    # ============================
    ms = input("Does this describe a [M]eaning/Concept or a cognate[s]et? > ")
    if re.search(r"\w", ms) == "s":
        cognateset = True
    else:
        cognateset = False

    data = pycldf.Wordlist(csvw.TableGroup(fname=args.json))
    # TODO: Should this be necessary? Check with @xrotwang
    data.properties["dc:conformsTo"] = "http://cldf.clld.org/v1.0/terms.rdf#Wordlist"

    # STEP 3: Determine how languages are mapped to columns
    # =====================================================
    example_language = most_content(ws.iter_cols(min_col=left, max_row=top - 1))
    clear()
    print(
        "Your most verbose language metadata, with all potential cells and comments, looks like this:"
    )
    pretty_print_example_cells(example_language)
    print(
        "What should these items be mapped to? Enter column names and separators like 'Name (Code)'."
    )
    print("For help: http://git.io/lexedata-g")

    readline.set_completer(completer(LANGUAGE_PROPERTIES))
    readline.parse_and_bind("tab: complete")
    lang_cell_regexes, lang_comment_regexes = create_parsers(top)

    lang_column_names: t.Set[str] = {"cldf_id"}
    for r in lang_cell_regexes:
        lang_column_names.update(r.groupindex)
    for r in lang_comment_regexes:
        lang_column_names.update(r.groupindex)

    add_table_with_columns("LanguageTable", lang_column_names, data)
    data.write(LanguageTable=[])

    example_row = most_content(ws.iter_rows(max_col=left - 1, min_row=top))
    clear()
    print(
        "Your most verbose row metadata, with all potential cells and comments, looks like this:"
    )
    pretty_print_example_cells(example_row)
    print(
        "What should these items be mapped to? Enter column names and separators like 'Name (Code)'."
    )
    print("For help: http://git.io/lexedata-g")

    readline.set_completer(
        completer(COGNATESET_PROPERTIES if cognateset else CONCEPT_PROPERTIES)
    )
    readline.parse_and_bind("tab: complete")
    row_cell_regexes, row_comment_regexes = create_parsers(left)

    column_names: t.Set[str] = set("cldf_id")
    for r in row_cell_regexes:
        column_names.update(r.groupindex)
    for r in row_comment_regexes:
        column_names.update(r.groupindex)

    add_table_with_columns(
        "CognatesetTable" if cognateset else "ParameterTable", column_names, data
    )
    data.write(**{"CognatesetTable" if cognateset else "ParameterTable": []})

    # For debugging:
    return locals()


if __name__ == "__main__":
    globals().update(main())
