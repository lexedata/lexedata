"""Import data in the "interleaved" format from an Excel spreadsheet.

Here, every even row contains cells with forms (cells may contain
multiple forms), while every odd row contains the associated cognate codes (a one-to-one relationship between forms and codes is expected).
Forms and cognate codes are separated by commas (",") and semi-colons (";"). Any other information existing in the cell will be parsed as
part of the form or the cognate code.
"""
import csv
import itertools
import logging
import os
import re
import typing as t
from pathlib import Path

import openpyxl

from lexedata import cli, types, util
from lexedata.util.excel import clean_cell_value


def import_interleaved(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    logger: logging.Logger = cli.logger,
    ids: t.Optional[t.Set[types.Cognateset_ID]] = None,
) -> t.Iterable[
    t.Tuple[
        types.Form_ID,
        types.Language_ID,
        types.Parameter_ID,
        t.Optional[str],  # Form
        t.Optional[str],  # Comment
        types.Cognateset_ID,
    ]
]:
    if ids is None:
        ids = set()

    comma_or_semicolon = re.compile("[,;]\\W*")

    concepts = []
    for concept_metadata in ws.iter_cols(min_col=1, max_col=1, min_row=2):
        for entry, cogset in zip(concept_metadata[::2], concept_metadata[1::2]):
            try:
                concepts.append(clean_cell_value(entry))
            except AttributeError:
                break

    if ws.max_row % 2 == 0:
        logger.warning(
            f"The sheet {ws.title} contained an even number of rows in total "
            f"({ws.max_row} including headers). "
            f"Given that all other rows come in pairs (forms+cognatesets), "
            f"I may have a problem with the format. "
            f"Are there multiple header rows, "
            f"or rows that look empty but are not after the end of the word list?"
        )

    for language in cli.tq(
        ws.iter_cols(min_col=2), task="Parsing cells", total=ws.max_column
    ):
        language_name = clean_cell_value(language[0])
        for c, (entry, cogset) in enumerate(zip(language[1::2], language[2::2])):
            if not entry.value:
                if cogset.value:
                    logger.warning(
                        f"Cell {entry.coordinate} was empty, but cognatesets {cogset.value} were given in {cogset.coordinate}."
                    )
                continue
            if not cogset.value and entry.value:
                logger.warning(
                    f"Cell {entry.coordinate} is a form cell, but is not followed by a cognateset. "
                    f"There was an even number of header rows which needs to be fixed."
                )
            bracket_level = 0
            i = 0
            f = clean_cell_value(entry)
            forms = []

            try:
                len(f)
            except TypeError:
                cli.Exit.INVALID_INPUT(
                    "I expected one or more forms (so, text) in cell {}, but found {}. Do you have more than one header row?".format(
                        entry.coordinate, f
                    )
                )

            while i < len(f):
                match = comma_or_semicolon.match(f[i:])
                if f[i] == "(":
                    bracket_level += 1
                    i += 1
                    continue
                elif f[i] == ")":
                    bracket_level -= 1
                    i += 1
                    continue
                elif bracket_level:
                    i += 1
                    continue
                elif match:
                    forms.append(f[:i].strip())
                    i += match.span()[1]
                    f = f[i:]
                    i = 0
                else:
                    i += 1

            forms.append(f.strip())

            if isinstance(clean_cell_value(cogset), int):
                cogsets = [str(clean_cell_value(cogset))]
            else:
                cogset = clean_cell_value(cogset)
                cogsets = comma_or_semicolon.split(cogset.strip())

            if len(cogsets) == len(forms):
                pass
            elif len(cogsets) == 1:
                logger.warning(
                    "%s: Multiple forms (%s) did not match single cognateset (%s), using that cognateset for each form.",
                    entry.coordinate,
                    ", ".join(forms),
                    cogsets[0],
                )
                cogsets = [cogsets[0] for _ in forms]
            else:
                logger.warning(
                    "%s: Forms (%s) did not match cognates (%s), adding NA values to the shorter one.",
                    entry.coordinate,
                    ", ".join(forms),
                    ", ".join(cogsets),
                )
            for form, cogset in itertools.zip_longest(forms, cogsets):
                if form == "?" or form is None or not form.strip():
                    form = None
                if cogset == "?" or cogset is None or not cogset.strip():
                    cogset = None
                if form is None and cogset is None:
                    # I'm not sure how this could happen, but if it happens,
                    # this is what should be done about it.
                    continue
                base_id = util.string_to_id(f"{language_name}_{concepts[c]}")
                id: types.Cognateset_ID = base_id
                synonym = 1
                while id in ids:
                    synonym += 1
                    id = f"{base_id}_s{synonym:d}"
                yield (id, language_name, concepts[c], form, None, cogset)
                ids.add(id)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        prog=f"python -m {__package__}.{Path(__file__).stem}", description=__doc__
    )
    parser.add_argument(
        "excel", type=Path, help="The Excel file to parse", metavar="EXCEL"
    )
    parser.add_argument(
        "--sheets",
        metavar="SHEET",
        nargs="+",
        default=[],
        help="Excel sheet name(s) to import (default: all sheets)",
    )
    parser.add_argument(
        "--directory",
        type=Path,
        default=Path(os.getcwd()),
        help="Path to directory where forms.csv is to be created (default: current working directory)",
    )
    cli.add_log_controls(parser)
    args = parser.parse_args()
    logger = cli.setup_logging(args)

    ws = openpyxl.load_workbook(args.excel)

    w = csv.writer(
        open(Path(args.directory) / "forms.csv", "w", newline="", encoding="utf-8")
    )
    w.writerow(
        ["ID", "Language_ID", "Parameter_ID", "Form", "Comment", "Cognateset_ID"]
    )

    if not args.sheets:
        args.sheets = [sheet for sheet in ws.sheetnames]

    ids: t.Set[str] = set()
    for sheetname in args.sheets:
        sheet = ws[sheetname]
        for row in import_interleaved(sheet, logger=logger, ids=ids):
            w.writerow(row)
