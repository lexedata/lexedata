# -*- coding: utf-8 -*-
from pathlib import Path

from sqlalchemy import or_, and_
import openpyxl as op

from lexedata.importer.objects import Language, Concept, Form, FormToConcept, Cognate, CogSet, CognateJudgement, \
    DatabaseObjectWithUniqueStringID, create_db_session
from lexedata.importer.database import LEXICAL_ORIGIN, COGNATE_ORIGIN, DATABASE_ORIGIN, connect_db
from lexedata.importer.cellparser import CellParser, CogCellParser
from lexedata.importer.exceptions import *


def create_db(db_path=DATABASE_ORIGIN, lexical=LEXICAL_ORIGIN, cogset_file=COGNATE_ORIGIN, echo=False):
    # check for existing resources and database
    if not lexical.exists():
        print("Necessary data not found: {}".format(lexical))
        print("exit")
        exit()
    if not cogset_file.exists():
        print("Necessary data not found: {}".format(cogset))
        print("exit")
        exit()

    if db_path.exists():
        answer = input("Do you want do delete the existing data base?y/n?")
        if answer == "y":
            db_path.unlink()
        else:
            print("exit")
            exit()


    session = create_db_session(location=db_path, echo=echo)

    # add languages
    lan_dict = insert_languages(session, source=lexical)
    print("--- Languages successfully inserted ---")

    # add concepts and forms
    wb = op.load_workbook(filename=lexical)
    sheets = wb.sheetnames
    wb = wb[sheets[0]]
    for row_forms, row_con in zip(wb.iter_rows(min_row=3, min_col=7), wb.iter_rows(min_row=3, max_col=6)):
        concept = Concept.from_default_excel(row_con)
        insert_concepts(session, concept)

        for form in yield_forms(row_forms, concept, lan_dict):
            insert_forms_and_forms_to_concepts(session, form)
    print("--- Concepts and Forms successfully inserted ---")

    # add cogsets and cognatejudgements
    wb = op.load_workbook(filename=cogset_file)
    for sheet in ['Numbers, Body Parts, Food, Anim', 'Kinship, Colors, Time, Nature', 'Tools, Adj, Adv', 'Verbs']:
        ws = wb[sheet]
        for cogset_row, cog_row in zip(ws.iter_rows(min_row=3, max_col=4), ws.iter_rows(min_row=3, min_col=5)):
            # ignore empty rows
            if not cogset_row[1].value:
                continue
            if cogset_row[1].value.isupper():
                cogset = CogSet.from_excel(cogset_row)
                insert_congsets(session, cogset)

                for cognate in yield_cognates(cog_row, cogset, lan_dict):
                    # within the insert function the db is queried for forms corresponding forms
                    # the actually inserted element links form, cognatejudgement and cogset
                    insert_cognates(session, cognate)
            # line not to be processed
            else:
                continue
    print("--- Cognate sets and cognate judgement successfully inserted ---")
    session.close()


# just for debugging
def inspect_cognates(cogset_file=COGNATE_ORIGIN):
    wb = op.load_workbook(filename=LEXICAL_ORIGIN)
    sheets = wb.sheetnames
    wb = wb[sheets[0]]
    language_iter = wb.iter_cols(min_row=1, max_row=2, min_col=7)
    lan_dict = dict()
    for language_cell in language_iter:
        if language_cell[0].value is None:
            continue
        else:
            language = Language.from_column(language_cell)
            lan_dict[language_cell[0].column_letter] = language.id
    #session = connect_db(read_only=False)
    #session.query(CogSet).delete()
    #session.query(CognateJudgement).delete()
    #session.commit()
    # add cogsets and cognatejudgements
    wb = op.load_workbook(filename=cogset_file)
    for sheet in ['Numbers, Body Parts, Food, Anim', 'Kinship, Colors, Time, Nature', 'Tools, Adj, Adv', 'Verbs']:
        ws = wb[sheet]
        try:
            for cogset_row, cog_row in zip(ws.iter_rows(min_row=3, max_col=4), ws.iter_rows(min_row=3, min_col=5)):
                # ignore empty rows
                if not cogset_row[1].value:
                    continue
                if cogset_row[1].value.isupper():
                    cogset = CogSet.from_excel(cogset_row)
                    #insert_congsets(session, cogset)


                    for cognate in yield_cognates(cog_row, cogset, lan_dict):
                        # within the insert function the db is queried for forms corresponding forms
                        # the actually inserted element links form, cognatejudgement and cogset
                        #insert_cognates(session, cognate)
                        print(cognate)
                        input()
                # line not to be processed
                else:
                    continue
        except AlreadyExistsError as err:
            print(err)

    print("--- Cognate sets and cognate judgement successfully inserted ---")
    #session.close()


def insert_languages(session, source=LEXICAL_ORIGIN, return_dictionary=True):
    "Reads languages from excel and inserts into db. optionally return dict: column_letter : language.id"
    # create iterator over excel cells
    wb = op.load_workbook(filename=source)
    sheets = wb.sheetnames
    wb = wb[sheets[0]]
    language_iter = wb.iter_cols(min_row=1, max_row=2, min_col=7)
    if return_dictionary:
        lan_dict = dict()
    for language_cell in language_iter:
        if language_cell[0].value is None:
            continue
        else:
            language = Language.from_column(language_cell)
            session.add(language)
            if return_dictionary:
                # add language id to lan dict:
                lan_dict[language_cell[0].column_letter] = language.id
    session.commit()
    if return_dictionary:
        return lan_dict


def insert_concepts(session, my_concept):
    try:
        exists = session.query(Concept).filter(Concept.id == my_concept.id).scalar()
        if exists is not None:
            raise AlreadyExistsError(my_concept)
        session.add(my_concept)
        session.commit()
    except AlreadyExistsError as err:
        print(err)


def insert_congsets(session, my_cogset):
    exists = session.query(CogSet).filter(CogSet.id == my_cogset.id).scalar()
    if exists is not None:
        raise AlreadyExistsError(my_cogset)
    session.add(my_cogset)
    session.commit()


def insert_forms_and_forms_to_concepts(session, form):
    """
    :param session: db session
    :param form: element of Class Form
    form is only inserted if it does not exist in the data base
    a form_to_concept is created, linking concept and form
        or in case of existing form: linking concept and existing form
    :return: None
    """
    form_to_concept = FormToConcept.from_form(form)
    # check if existing, and if so add variants to variants
    # only when exact match
    already_existing = session.query(Form).filter(
        Form.language_id == form.language_id,
        Form.phonetic == form.phonetic,
        Form.phonemic == form.phonemic,
        Form.orthographic == form.orthographic).one_or_none()
    if already_existing is None:
        session.add(form)
        session.add(form_to_concept)
    else:
        existing_form = already_existing
        existing_variants = existing_form.variants.split(";")
        new_variants = form.variants.split(";")
        if set(new_variants) - set(existing_variants):
            print(
                "Variants {:} of form {:} were not mentioned earlier. Adding them to the form nonetheless".format(
                    set(new_variants) - set(existing_variants),
                    form))
            form.variants = ";".join(set(new_variants) | set(existing_variants))
        if set(existing_variants) - set(new_variants):
            print(set(existing_variants) - set(new_variants))
            # print("{:}: Variants {:} of form {:} were not mentioned here.".format(
            #        set(existing_variants) - set(new_variants), existing_form))
        # link form_to_concept element to existing form
        form_to_concept.form_id = existing_form.id

    session.add(form_to_concept)

    session.commit()


def query_forms_to_cognate(cog, session, myfilters):
    # just select same language
    myquery = session.query(Form).filter(Form.language_id == cog.language_id)

    # select with just one matching transcription, i.e. using or_
    one_transcription = myquery.filter(or_(getattr(Form, attribute) == value for attribute, value in myfilters.items()))

    len_one = one_transcription.scalar()
    # if just one match or no match so far, exit here
    if len_one == 1 or len_one == 0:
        return one_transcription.all()

    # match for the source
    one_transcription_source = one_transcription.filter(Form.source == cog.source)

    # if just one match with source, exit here; if no match with source, return without source
    len_source = one_transcription_source.scalar()
    if len_source == 1:
        return one_transcription_source.all()
    elif len_source == 0:
        return one_transcription.all()

    # match all given transcriptions
    myquery = one_transcription_source.filter(and_(getattr(Form, attribute) == value for attribute, value in myfilters.items())).all()

    # if no exact match, return match with source
    if len(myquery) == 0:
        return one_transcription_source.all()
    else:
        return myquery


def insert_cognates(session, cog):
    # create dict with not empty attributes
    myfilters = {}
    if cog.phonemic != "":
        myfilters["phonemic"] = cog.phonemic
    if cog.phonetic != "":
        myfilters["phonetic"] = cog.phonetic
    if cog.orthographic != "":
        myfilters["orthographic"] = cog.orthographic

    res = query_forms_to_cognate(cog, session, myfilters)
    try:
        if compare_cognates(cog, res, myfilters):
            judgement = CognateJudgement.from_cognate_and_form(cog, res[0])
            session.add(judgement)
            session.commit()

    except MultipleMatchError as err:
        print(err)
        #
        input()
    except NoSourceMatchError as err:
        print(err)
        #input()
    except PartialMatchError as err:
        print(err)
        #input()
    except MatchingError as err:
        print(err)
        #input()


def compare_cognates(cog, form, myfilters):
    "returns True if one exact match is found, raises an error otherwise"
    if len(form) == 0:
        raise MatchingError(cog, "")
    elif len(form) > 1:
        raise MultipleMatchError(cog, form)

    elif cog.source != form[0].source:
        raise NoSourceMatchError(cog, form)

    elif not all(getattr(form[0], attribute) == getattr(cog, attribute) for attribute in myfilters):
        raise PartialMatchError(cog, form)
    else:
        return True


def yield_forms(row, concept, lan_dict):
    for f_cell in row:
        if f_cell.value:

            # get corresponding language_id to column
            this_lan_id = lan_dict[f_cell.column_letter]
            for f_ele in CellParser(f_cell):
                yield Form.create_form(f_ele, this_lan_id, f_cell, concept)


def yield_cognates(row, cogset, lan_dict):
    for f_cell in row:
        if f_cell.value:
            # get corresponding language_id to column
            this_lan_id = lan_dict[f_cell.column_letter]
            print(f_cell.value)
            for f_ele in CogCellParser(f_cell):
                print(f_ele)
                yield Cognate.from_excel(f_ele, this_lan_id, f_cell, cogset)


def test():
    db_path = DIR_DATA / "lexedata.db"
    db_path = "sqlite:///" + str(db_path)
    db_path = db_path.replace("\\", "\\\\")  # can this cause problems on IOS?
    session = create_db_session(location=db_path, echo=False)
    insert_cognates(dir_path, session)
    session.close()


def show_empty_forms():
    dir_path = Path.cwd() / "initial_data"
    db_path = dir_path / "lexedata.db"
    db_path = "sqlite:///" + str(db_path)
    db_path = db_path.replace("\\", "\\\\")  # can this cause problems on IOS?
    session = create_db_session(location=db_path, echo=False)
    return session.query(Form).filter(
        Form.phonetic == "",
        Form.phonemic == "",
        Form.orthographic == "").all()


if __name__ == "__main__":
    inspect_cognates()


