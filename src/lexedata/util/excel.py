# -*- coding: utf-8 -*-
"""Various helper functions for Excel file parsing

"""

import re
import typing as t
import unicodedata
import datetime

import openpyxl as op
import pycldf

import lexedata.cli as cli
from lexedata.types import Form, Judgement
from lexedata.util import string_to_id


def clean_cell_value(cell: op.cell.cell.Cell, logger=cli.logger):
    """Return the value of an Excel cell in a useful format and normalized."""
    if cell.value is None:
        return ""
    if type(cell.value) == float:
        if cell.value == int(cell.value):
            return int(cell.value)
        return cell.value
    elif type(cell.value) == int:
        return cell.value
    elif type(cell.value) == datetime.datetime:  # pragma: no cover
        logger.warning(
            "Encountered Date/Time value %s in cell %s.", cell.value, cell.coordinate
        )
        cell.value = str(cell.value)
    try:
        v = unicodedata.normalize("NFC", (cell.value or "").strip())
        return v.replace("\n", ";\t")
    except TypeError:
        return str(v)


def get_cell_comment(cell: op.cell.Cell) -> str:
    """Get the comment of a cell.

    Get the normalized comment of a cell: Guaranteed to be a string (empty if
    no comment), with lines joined by spaces instead and all 'lexedata' author
    annotations stripped.


    >>> from openpyxl.comments import Comment
    >>> wb = op.Workbook()
    >>> ws = wb.active
    >>> ws["A1"].comment = Comment('''This comment
    ... contains a linebreak and a signature.
    ...   -lexedata.exporter''',
    ... 'lexedata')
    >>> get_cell_comment(ws["A1"])
    'This comment contains a linebreak and a signature.'
    >>> get_cell_comment(ws["A2"])
    ''
    """
    raw_comment = cell.comment.text.strip() if cell.comment else ""
    lines = [
        line for line in raw_comment.split("\n") if line.strip() != "-lexedata.exporter"
    ]
    return " ".join(lines)


def normalize_header(row: t.Iterable[op.cell.Cell]) -> t.Iterable[str]:
    header = [unicodedata.normalize("NFC", (n.value or "").strip()) for n in row]
    return header


def check_brackets(string, bracket_pairs):
    """Check whether all brackets match.

    This function can check the matching of simple bracket pairs, like this:

    >>> b = {"(": ")", "[": "]", "{": "}"}
    >>> check_brackets("([])", b)
    True
    >>> check_brackets("([]])", b)
    False
    >>> check_brackets("([[])", b)
    False
    >>> check_brackets("This (but [not] this)", b)
    True

    But it can also deal with multi-character matches

    >>> b = {"(": ")", "begin": "end"}
    >>> check_brackets("begin (__ (!) xxx) end", b)
    True
    >>> check_brackets("begin (__ (!) end) xxx", b)
    False

    This includes multi-character matches where some pair is a subset of
    another pair. Here the order of the pairs in the dictionary is important –
    longer pairs must be defined first.

    >>> b = {":::": ":::", ":": ":"}
    >>> check_brackets("::: :::", b)
    True
    >>> check_brackets("::::::", b)
    True
    >>> check_brackets("::::", b)
    False
    >>> check_brackets(":: ::", b)
    True

    In combination, these features allow for natural escape sequences:

    >>> b = {"!(": "", "!)": "", "(": ")", "[": "]"}
    >>> check_brackets("(text)", b)
    True
    >>> check_brackets("(text", b)
    False
    >>> check_brackets("text)", b)
    False
    >>> check_brackets("(te[xt)]", b)
    False
    >>> check_brackets("!(text", b)
    True
    >>> check_brackets("text!)", b)
    True
    >>> check_brackets("!(te[xt!)]", b)
    True
    """
    waiting_for = []
    i = 0
    while i < len(string):
        if waiting_for and string[i:].startswith(waiting_for[0]):
            i += len(waiting_for.pop(0))
        else:
            for q, p in bracket_pairs.items():
                if string[i:].startswith(q):
                    waiting_for.insert(0, p)
                    i += len(q)
                    break
                elif p and string[i:].startswith(p):
                    return False
            else:
                i += 1
    return not any(waiting_for)


def components_in_brackets(form_string, bracket_pairs):
    """Find all elements delimited by complete pairs of matching brackets.

    >>> b = {"!/": "", "(": ")", "[": "]", "{": "}", "/": "/"}
    >>> components_in_brackets("/aha/ (exclam. !/ int., also /ah/)",b)
    ['', '/aha/', ' ', '(exclam. !/ int., also /ah/)', '']

    Recovery from mismatched delimiters early in the string is difficult. The
    following example is still waiting for the first '/' to be closed by the
    end of the string.

    >>> components_in_brackets("/aha (exclam. !/ int., also /ah/)",b)
    ['', '/aha (exclam. !/ int., also /ah/)']

    """
    elements = []

    i = 0
    remainder = form_string
    waiting_for = []
    while i < len(remainder):
        if waiting_for and remainder[i:].startswith(waiting_for[0]):
            i += len(waiting_for.pop(0))
            if not any(waiting_for):
                elements.append(remainder[:i])
                remainder = remainder[i:]
                i = 0
        else:
            for q, p in bracket_pairs.items():
                if remainder[i:].startswith(q):
                    if not any(waiting_for):
                        elements.append(remainder[:i])
                        remainder = remainder[i:]
                        i = 0
                    waiting_for.insert(0, p)
                    i += len(q)
                    break
                # elif p and remainder[i:].startswith(p):
                #     # TODO: @Geroen: do we need this warning? I think this case is better handled in parse_form...
                #     logger.info(
                #         f"{context:}In form {form_string}: Encountered mismatched closing delimiter {p}. "
                #         f"This could be a bigger problem in the cell, so the form was not imported."
                #    )
            else:
                i += 1

    return elements + [remainder]


class NaiveCellParser:
    c: t.Dict[str, str]

    def __init__(self, dataset: pycldf.Dataset):
        self.separators = {}
        self.c = {}
        self.cc(short="value", long=("FormTable", "value"), dataset=dataset)
        self.cc(short="form", long=("FormTable", "form"), dataset=dataset)
        self.cc(short="lang", long=("FormTable", "languageReference"), dataset=dataset)

    def cc(self, short, dataset, long=None):
        """Cache the name of a column, or complain if it doesn't exist"""
        if long is None:
            long = ("FormTable", short)
        try:
            self.c[short] = dataset[long].name
            self.separators[self.c[short]] = dataset[long].separator
        except KeyError:
            raise ValueError(
                "Your metadata json file and your cell parser don’t match: "
                f"Your cell parser {self.__class__.__name__} expects a #{long[1]} column (usually named '{long[1]}') "
                f"in FormTable, but your metadata defines no such column."
            )

    def separate(self, values: str, context: str = "") -> t.Iterable[str]:
        """Separate different form descriptions in one string.

        Separate forms separated by comma.
        """
        return values.split(",")

    def parse_form(
        self, form_string: str, language_id: str, cell_identifier: str = ""
    ) -> t.Optional[Form]:
        return Form(
            {
                self.c["value"]: form_string,
                self.c["form"]: form_string.strip(),
                self.c["lang"]: language_id,
            }
        )

    def parse(
        self,
        cell: op.cell.Cell,
        language_id: str,
        cell_identifier: str = "",
        logger: cli.logging.Logger = cli.logger,
    ) -> t.Iterable[Form]:
        """Return form properties for every form in the cell"""
        # cell_identifier format: sheet.cell_coordinate
        cell_identifier = "{}: ".format(cell_identifier) if cell_identifier else ""

        text = clean_cell_value(cell)
        if not text:
            return []

        for element in self.separate(
            text, context=cell_identifier and f"{cell_identifier}: "
        ):
            try:
                form = self.parse_form(element, language_id, cell_identifier)
            except (KeyError, ValueError) as e:
                logger.warning(f"Could not import form {element}: {e}")
                continue
            if form:
                yield form


class CellParser(NaiveCellParser):
    def __init__(
        self,
        dataset: pycldf.Dataset,
        element_semantics: t.Iterable[t.Tuple[str, str, str, bool]] = [
            # ("[", "]", "phonetic", True),
            ("<", ">", "form", True),
            # ("/", "/", "phonemic", True),
            ("(", ")", "comment", False),
            ("{", "}", "source", False),
        ],
        separation_pattern: str = r"([;,])",
        variant_separator: t.Optional[t.List[str]] = ["~", "%"],
        add_default_source: t.Optional[str] = "1",
        logger: cli.logging.Logger = cli.logger,
    ):
        super().__init__(dataset)

        # Colums implied by element semantics
        self.bracket_pairs = {start: end for start, end, _, _ in element_semantics}
        self.element_semantics = {
            start: (term, transcription)
            for start, _, term, transcription in element_semantics
        }
        for start, end, term, transcription in element_semantics:
            # Ensure that all terms required by the element semantics are fields we can write to.
            self.cc(short=term, long=("FormTable", term), dataset=dataset)
            if ("FormTable", "source") in dataset and dataset[
                "FormTable", term
            ].name == dataset["FormTable", "source"].name:
                self.source_delimiter = (start, end)
            if ("FormTable", "comment") in dataset and dataset[
                "FormTable", term
            ].name == dataset["FormTable", "comment"].name:
                self.comment_delimiter = (start, end)
        assert self.transcriptions, (
            "Your metadata json file and your cell parser don’t match: Your cell parser "
            f"{self.__class__.__name__} expects to work with transcriptions "
            "(at least one of 'orthographic', 'phonemic', and 'phonetic') to derive a #form "
            "in #FormTable, but your metadata defines no such column."
        )
        self.leftovers = self.element_semantics.pop("", None)
        if self.bracket_pairs.pop("", ""):
            logger.warning(
                "Your ‘delimiter pair’ for forms outside delimiters should be the empty string for start and end, but your end string was not empty. I will ignore the end string you specificed."
            )

        # Colums necessary for word list
        self.cc(short="source", long=("FormTable", "source"), dataset=dataset)
        self.cc(short="comment", long=("FormTable", "comment"), dataset=dataset)
        # Pretend that the comment field has a separator, but keep track of the original one.
        self.comment_separator = self.separators[self.c["comment"]]
        self.separators[self.c["comment"]] = self.separators[self.c["comment"]] or "\t"

        try:
            # As long as there is no CLDF term #variants, this will either be
            # 'variants' or raise a KeyError. However, it is a transparent
            # re-use of an otherwise established idiom in this module, so we
            # use this minor overhead.
            self.cc(short="variants", long=("FormTable", "variants"), dataset=dataset)
            # Assert that #variants has a separator?
        except ValueError:
            logger.warning(
                "No 'variants' column found for FormTable in Wordlist-metadata.json. "
                "Form variants will be added to #comment."
            )
            self.c["variants"] = self.c["comment"]

        # Other class attributes
        self.separation_pattern = separation_pattern
        self.variant_separator = variant_separator
        self.add_default_source = add_default_source

    def source_from_source_string(
        self,
        source_string: str,
        language_id: t.Optional[str],
        delimiters: t.Tuple[str, str] = ("{", "}"),
        logger: cli.logging.Logger = cli.logger,
    ) -> str:
        """Parse a string referencing a language-specific source"""
        if source_string.startswith(delimiters[0]) and source_string.endswith(
            delimiters[1]
        ):
            source_string = source_string[len(delimiters[0]) : -len(delimiters[1])]
        elif source_string.startswith(delimiters[0]):
            logger.warning(
                f"In source {source_string}: Closing bracket '{delimiters[1]}' is missing."
            )
            source_string = source_string[len(delimiters[0]) :]

        context: t.Optional[str]
        if ":" in source_string:
            source_string, context = source_string.split(":", maxsplit=1)
            source_string = source_string.strip()
            context = context.strip()
            context = context.replace("]", "")
        else:
            context = None

        if language_id is None:
            source_id = string_to_id(source_string)
        else:
            source_id = string_to_id(f"{language_id:}_s{source_string:}")

        if context:
            return f"{source_id}[{context}]"
        else:
            return source_id

    @property
    def transcriptions(self):
        try:
            return self._transcriptions
        except AttributeError:
            self._transcriptions = [
                self.c[k] for k, t in self.element_semantics.values() if t
            ]
            return self._transcriptions

    def separate(
        self,
        values: str,
        context: str = "",
        logger: cli.logging.Logger = cli.logger,
    ) -> t.Iterable[str]:
        """Separate different form descriptions in one string.

        Separate forms separated by comma or semicolon, unless the comma or
        semicolon occurs within a set of matching component delimiters (eg.
        brackets)

        If the brackets don't match, the whole remainder string is passed on,
        so that the form parser can try to recover as much as possible or throw
        an exception.
        """
        raw_split = re.split(self.separation_pattern, values)
        if len(raw_split) <= 1:
            for form in raw_split:
                yield form
            return

        while len(raw_split) > 1:
            if check_brackets(raw_split[0], self.bracket_pairs):
                form = raw_split.pop(0).strip()
                if form:
                    yield form
                raw_split.pop(0)
            else:
                raw_split[:2] = ["".join(raw_split[:2])]
        if not check_brackets(raw_split[0], self.bracket_pairs):
            logger.warning(
                f"{context:}In values {values:}: "
                "Encountered mismatched closing delimiters. Please check that the "
                "separation of the cell into multiple entries, for different forms, was correct."
            )

        form = raw_split.pop(0).strip()
        if form:
            yield form
        assert not raw_split

    def parse_form(
        self,
        form_string: str,
        language_id: str,
        cell_identifier: str = "",
        logger: cli.logging.Logger = cli.logger,
    ) -> t.Optional[Form]:
        """Create a dictionary of columns from a form description.

        Extract each value (transcriptions, comments, sources etc.) from a
        string describing a single form.
        """
        # not required fields
        c_comment = self.c.get("comment")
        c_variants = self.c.get("variants", c_comment)

        # if string is only whitespaces, there is no form.
        if not form_string.strip():
            return None

        properties: t.Dict[str, t.Any] = {
            self.c["lang"]: language_id,
            self.c["value"]: form_string,
        }

        # Semantics: 'None' for no variant expected, any string for the
        # decorator that introduces variant forms. Currently we expect '~' and
        # '%', see below.
        expect_variant: t.Optional[str] = None
        # Iterate over the delimiter-separated elements of the form.
        for element in components_in_brackets(form_string, self.bracket_pairs):
            element = element.strip()

            if not element:
                continue

            # If the element has mismatched brackets (tends to happen only for
            # the last element, because a mismatched opening bracket means we
            # are still waiting for the closing one), warn.
            if not check_brackets(element, self.bracket_pairs):
                try:
                    delimiter = self.bracket_pairs[element[0]]
                except KeyError:
                    delimiter = element[0]
                raise ValueError(
                    f"{cell_identifier}In form {form_string}: Element {element} had mismatching delimiters "
                    f"{delimiter}. This could be a bigger problem in the cell, "
                    f"so the form was not imported."
                )
            # Check what kind of element we have.
            for start, (term, transcription) in self.element_semantics.items():
                if element.startswith(start):
                    field = self.c[term]
                    delimiters = start, self.bracket_pairs[start]
                    break
            else:
                # The only thing we expect outside delimiters is the variant
                # separators, '~' and '%'. That is, unless we have some
                # 'leftovers' specification.
                if self.variant_separator and element in self.variant_separator:
                    expect_variant = element
                    continue
                elif self.leftovers:
                    # Check whether the stuff outside ends with a variant
                    # separator, because that is problematic.
                    maybe_sep = [
                        v for v in self.variant_separator if element.endswith(v)
                    ]
                    if maybe_sep:
                        element = element[: -len(maybe_sep[0])].strip()
                        # XXX Can we do better? We'd need to keep the variant separator around for *next* loop iteration
                        logger.warning(
                            f"{cell_identifier}In form {form_string}: Element {element} ended with a variant separator, which is not reflected in the output."
                        )
                    elif any([v in element for v in self.variant_separator]):
                        logger.info(
                            f"{cell_identifier}In form {form_string}: Element {element} contained variant separator, but you also specified that elements outside delimiters should be treated as {term}. Please check your output."
                        )

                    # If there are delimiters for this kind of element, use
                    # those instead.
                    for start, semantics in self.element_semantics.items():
                        if self.leftovers == semantics:
                            delimiters = start, self.bracket_pairs[start]
                            element = f"{delimiters[0]}{element}{delimiters[1]}"
                            break
                    else:
                        delimiters = ("", "")
                    term, transcription = self.leftovers
                    field = self.c[term]
                else:
                    logger.warning(
                        f"{cell_identifier}In form {form_string}: Element {element} could not be parsed, ignored"
                    )
                    continue

            # If we encounter a field for the first time, we add it to the
            # dictionary. If repeatedly, to the variants, with a decorator that
            # shows how expected the variant was. This drops sources and
            # comments in variants, if more than one source or comment is
            # provided – We clean this up in self.postprocess_form

            if self.separators[field]:
                element = element[len(delimiters[0]) :]
                if element.endswith(delimiters[1]):
                    element = element[: -len(delimiters[1])]
                properties.setdefault(field, []).append(element)
            elif field in properties:
                if (
                    not expect_variant
                    and field != c_comment
                    and not self.separators[field]
                ):
                    logger.warning(
                        f"{cell_identifier}In form {form_string}: Element {element} was an unexpected variant for {field}"
                    )
                    properties.setdefault(c_variants, []).append(
                        (expect_variant or "") + element
                    )
                else:
                    properties.setdefault(c_variants, []).append(
                        (expect_variant or "") + element
                    )
            else:
                if expect_variant:
                    logger.warning(
                        f"{cell_identifier}In form {form_string}: Element {element} was supposed to be a variant, but there is no earlier {field}"
                    )
                element = element[len(delimiters[0]) :]
                if element.endswith(delimiters[1]):
                    element = element[: -len(delimiters[1])]
                properties[field] = element

            expect_variant = None

        self.postprocess_form(properties, language_id)
        return Form(properties)

    def create_cldf_form(self, properties: t.Dict[str, t.Any]) -> t.Optional[str]:
        """
        Return first transcription out of properties as a candidate for cldf_form.
        Order of transcriptions corresponds to order of cell_parser_semantics as provided in the metadata.
        """
        for candidate in self.transcriptions:
            if candidate in properties:
                return properties[candidate]
        return None

    def postprocess_form(
        self,
        properties: t.Dict[str, t.Any],
        language_id: str,
        source_delimiters: t.Tuple[str, str] = ("{", "}"),
    ) -> None:
        """Modify the form in-place

        Fix some properties of the form. This is the place to add default
        sources, cut of delimiters, split unmarked variants, etc.

        """
        # remove delimiters from comment
        if self.comment_separator is None:
            properties[self.c["comment"]] = self.separators[self.c["comment"]].join(
                properties.get(self.c["comment"], [])
            )

        source = properties.pop(self.c["source"], [])
        if self.add_default_source and not source:
            source = [self.add_default_source]
        # if source is already a set with source, don't change anything
        properties[self.c["source"]] = {
            self.source_from_source_string(
                s, language_id, delimiters=self.source_delimiter
            )
            for s in source
        }

        # add form to properties
        properties.setdefault(self.c["form"], self.create_cldf_form(properties))


def alignment_from_braces(text, start=0):
    """Convert a brace-delimited morpheme description into slices and alignments.

    The "-" character is used as the alignment gap character, so it does not
    count towards the segment slices.

    If opening or closing brackets are missing, the slice goes until the end of the form.

    >>> alignment_from_braces("t{e x t")
    ([(2, 4)], ['e', 'x', 't'])
    >>> alignment_from_braces("t e x}t")
    ([(1, 3)], ['t', 'e', 'x'])
    >>> alignment_from_braces("t e x t")
    ([(1, 4)], ['t', 'e', 'x', 't'])
    """
    # TODO: Should we warn/error instead?
    try:
        before, remainder = text.split("{", 1)
    except ValueError:
        before, remainder = "", text
    try:
        content, remainder = remainder.split("}", 1)
    except ValueError:
        content, remainder = remainder, ""
    content = content.strip()
    i = len(before.strip())
    j = len([s for s in content.split() if s != "-"])
    slice = (start + i + 1, start + i + j)
    if "{" in remainder:
        slices, alignment = alignment_from_braces(remainder, start + i + j)
        slices.insert(0, slice)
        return slices, content.split() + alignment
    else:
        return [slice], content.split()


class CellParserHyperlink(NaiveCellParser):
    def __init__(self, dataset: pycldf.Dataset, extractor: re.Pattern):
        super().__init__(dataset=dataset)
        self.extractor = extractor
        self.cc(short="c_id", long=("CognateTable", "formReference"), dataset=dataset)
        try:
            self.c["c_comment"] = dataset["CognateTable", "comment"].name
        except KeyError:
            pass
        try:
            self.c["c_segments"] = dataset["CognateTable", "segmentSlice"].name
            self.c["c_alignment"] = dataset["CognateTable", "alignment"].name
        except KeyError:
            pass

    def parse(
        self,
        cell: op.cell.Cell,
        language_id: str,
        cell_identifier: str = "",
        logger: cli.logging.Logger = cli.logger,
    ) -> t.Iterable[Judgement]:
        try:
            url = cell.hyperlink.target
            text = clean_cell_value(cell)
            comment = get_cell_comment(cell)
            if "{" not in text:
                slice, alignment = alignment_from_braces("{" + text + "}")
            else:
                slice, alignment = alignment_from_braces(text)
            try:
                form_id = self.extractor.search(url)["ID"]
            except (TypeError, IndexError):
                logger.error(
                    f"Could not extract group ID from URL {url} using regular expression {self.extractor.pattern}"
                )
                cli.Exit.INVALID_ID()
            properties = {
                self.c["c_id"]: form_id,
                self.c.get("c_segments"): ["{:}:{:}".format(i, j) for i, j in slice],
                self.c.get("c_alignment"): alignment,
                self.c.get("c_comment"): comment,
            }
            properties.pop(None, None)
            yield Judgement(properties)

        except AttributeError:
            pass


class MawetiCellParser(CellParser):
    def __init__(
        self,
        dataset: pycldf.Dataset,
        element_semantics: t.Iterable[t.Tuple[str, str, str, bool]],
        separation_pattern: str,
        variant_separator: list,
        add_default_source: t.Optional[str],
    ):
        super(MawetiCellParser, self).__init__(
            dataset,
            element_semantics=element_semantics,
            separation_pattern=separation_pattern,
            variant_separator=variant_separator,
            add_default_source=add_default_source,
        )

        # TODO: We might not have this
        self.cc("procedural_comment", dataset=dataset)

    def postprocess_form(
        self,
        properties: t.Dict[str, t.Any],
        language_id: str,
    ) -> None:
        """
        Post processing specific to the Maweti dataset
        """

        # catch procedural comments (e.g. NPC: ...) in #comment and add to
        # corresponding procedural comment.
        for i, subcomment in enumerate(properties.get(self.c["comment"], [])[::-1], 1):
            if re.match(r"^[A-Z]{2,}:", subcomment):
                properties.setdefault(self.c["procedural_comment"], []).insert(
                    0, subcomment
                )
                del properties[self.c["comment"]][-i]

        # TODO: Currently "..." lands in the forms, with empty other entries
        # (and non-empty source). This is not too bad for now, how should it
        # be?

        # Split transcriptions of form that contain '%' or '~', drop the
        # variant in the #variants column. if variants not in properties, add
        # default list
        for key in self.transcriptions:
            try:
                property_value = properties[key]
            except KeyError:
                continue
            # if any separator is in this value, split value. add first as key and rest to variants.
            for separator in self.variant_separator:
                if separator in property_value:
                    # split string with variants
                    # add first transcription as transcription, rest to variants
                    # ensure correct opening and closing for transcription and variants
                    values = property_value.split(separator)
                    first_value = values.pop(0)
                    first_value = first_value.strip()
                    opening = [
                        start
                        for start, (
                            term,
                            transcription,
                        ) in self.element_semantics.items()
                        if self.c[term] == key
                    ][0]
                    closing = self.bracket_pairs[opening]
                    properties[key] = first_value
                    for value in values:
                        while value.startswith(" "):
                            value = value.lstrip(" ")
                        while value.endswith(" "):
                            value = value.rstrip(" ")
                        if not value[0] == opening:
                            value = opening + value
                        if not value[-1] == closing:
                            value = value + closing
                        properties.setdefault(self.c["variants"], []).append(
                            separator + value
                        )

        properties.setdefault(self.c["form"], self.create_cldf_form(properties))
        super().postprocess_form(properties, language_id)


class MawetiCognateCellParser(MawetiCellParser):
    def parse_form(self, values, language, cell_identifier: str = ""):
        if values.isupper():
            return None
        else:
            return super().parse_form(values, language, cell_identifier)
