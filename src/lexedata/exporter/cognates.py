# -*- coding: utf-8 -*-
import abc
import re
import typing as t
import urllib.parse
from pathlib import Path

import openpyxl as op
import pycldf

from lexedata import cli, types, util
from lexedata.edit.add_singleton_cognatesets import create_singletons
from lexedata.util import parse_segment_slices

WARNING = "\u26A0"

# ----------- Remark: Indices in excel are always 1-based. -----------

# TODO: Make comments on Languages, Cognatesets, and Judgements appear as notes
# in Excel.

# Type aliases, for clarity
CognatesetID = str


class BaseExcelWriter:
    """Class logic for matrix-shaped Excel export."""

    row_table: str
    header: t.List[t.Tuple[str, str]]

    def __init__(
        self,
        dataset: pycldf.Dataset,
        database_url: t.Optional[str] = None,
        logger: cli.logging.Logger = cli.logger,
    ):
        self.set_header(dataset)
        self.separators = {
            util.cldf_property(c.propertyUrl) or c.name: c.separator
            for c in dataset[self.row_table].tableSchema.columns
            if c.separator
        }

        self.URL_BASE = database_url

        self.wb = op.Workbook()
        self.ws: op.worksheet.worksheet.Worksheet = self.wb.active

        self.logger = logger

    def create_excel(
        self,
        rows: t.Iterable[types.RowObject],
        languages,
        judgements: t.Iterable[types.Judgement],
        forms,
        size_sort: bool = False,
    ) -> None:
        """Convert the initial CLDF into an Excel cognate view

        The Excel file has columns "CogSet", one column each mirroring the
        other cognateset metadata, and then one column per language.

        The rows contain cognate data. If a language has multiple reflexes in
        the same cognateset, these appear in different cells, one below the
        other.

        """
        # Define the columns, i.e. languages and write to excel
        self.lan_dict: t.Dict[str, int] = {}
        excel_header = [name for cldf, name in self.header]

        for col, lan in cli.tq(
            enumerate(languages, len(excel_header) + 1),
            task="Writing languages to excel header",
            total=len(languages),
        ):
            # TODO: This should be based on the foreign key relation of
            # languageReference
            self.lan_dict[lan["id"]] = col
            excel_header.append(lan.get("name", lan["id"]))
        self.ws.append(excel_header)

        # Again, row_index 2 is indeed row 2, row 1 is header
        row_index = 1 + 1

        forms_by_row = self.collect_forms_by_row(judgements, rows)

        # iterate over all rows
        for row in cli.tq(
            rows,
            task="Writing rows to Excel",
            total=len(rows),
            logger=self.logger,
        ):
            # possibly a row can appear without any forms. Unlikely, but just
            # ignore those.
            if row["id"] not in forms_by_row:
                continue
            # write all forms of this cognateset to excel
            new_row_index = self.create_formcells(
                [
                    (forms[f], m)
                    for f, metadata in forms_by_row[row["id"]].items()
                    for m in metadata
                ],
                row_index,
            )
            # write rows for cognatesets, now that we know how many rows there
            # are. (TODO: Maybe we could format all but the first row a lot
            # weaker, so the groups stand out better?)
            for r in range(row_index, new_row_index):
                self.write_row_header(row, r)

            row_index = new_row_index

    def create_formcells(
        self,
        row_forms: t.Iterable[types.Form],
        row_index: int,
    ) -> int:
        """Writes all forms for given cognate set to Excel.

        Take all forms for a given cognate set as given by the database, create
        a hyperlink cell for each form, and write those into rows starting at
        row_index.

        Return the row number of the first empty row after this cognate set,
        which can then be filled by the following cognate set.

        """
        # Read the forms from the database and group them by language
        forms = t.DefaultDict[int, t.List[types.Form]](list)
        for form, metadata in row_forms:
            forms[self.lan_dict[form["languageReference"]]].append((form, metadata))

        if not forms:
            return row_index + 1

        # maximum of rows to be added
        maximum_size = max([len(c) for c in forms.values()])
        for column, cells in forms.items():
            for row, entry in enumerate(cells, row_index):
                self.create_formcell(entry, column, row)
        # increase row_index and return
        row_index += maximum_size

        return row_index

    def create_formcell(self, form: types.Form, column: int, row: int) -> None:
        """Fill the given cell with the form's data.

        In the cell described by ws, column, row, dump the data for the form:
        Write into the the form data, and supply a comment from the judgement
        if there is one.

        """
        form, metadata = form
        cell_value = self.form_to_cell_value(form)
        form_cell = self.ws.cell(row=row, column=column, value=cell_value)
        comment = metadata.get("comment")
        if comment:
            form_cell.comment = op.comments.Comment(comment, __package__)
        if self.URL_BASE:
            link = self.URL_BASE.format(urllib.parse.quote(form["id"]))
            form_cell.hyperlink = link

    @abc.abstractmethod
    def form_to_cell_value(self, form: types.Form):
        "Format a form into text for an Excel cell value"

    def collect_forms_by_row(
        self,
        judgements: t.Iterable[types.Judgement],
        rows: t.Iterable[types.Row_ID],
    ) -> t.Mapping[
        types.Cognateset_ID, t.Mapping[types.Form_ID, t.Sequence[types.Judgement]]
    ]:
        "Collect forms by row object (ie. concept or cognate set)"
        all_forms: t.MutableMapping[
            types.Cognateset_ID, t.Mapping[types.Form_ID, t.List[types.Judgement]]
        ] = t.DefaultDict(lambda: t.DefaultDict(list))
        for judgement in judgements:
            all_forms[judgement["cognatesetReference"]][
                judgement["formReference"]
            ].append(judgement)
        return all_forms

    @abc.abstractmethod
    def write_row_header(self, row_object: types.RowObject, row_index: int):
        """Write a row header

        Write into the first few columns of the row `row_index` of self.ws the
        metadata of a row, eg. concept ID and gloss or cognateset ID,
        cognateset name and status.

        """

    @abc.abstractmethod
    def set_header(
        self,
        dataset: types.Wordlist[
            types.Language_ID,
            types.Form_ID,
            types.Parameter_ID,
            types.Cognate_ID,
            types.Cognateset_ID,
        ],
    ):
        "Define the header for the first few columns"


class ExcelWriter(BaseExcelWriter):
    """Class logic for cognateset Excel export."""

    row_table = "CognatesetTable"

    def __init__(
        self,
        dataset: pycldf.Dataset,
        database_url: t.Optional[str] = None,
        singleton_cognate: bool = False,
        singleton_status: t.Optional[str] = None,
        logger: cli.logging.Logger = cli.logger,
    ):
        super().__init__(dataset=dataset, database_url=database_url, logger=logger)
        # assert that all required tables are present in Dataset
        try:
            for _ in dataset["CognatesetTable"]:
                break
        except (KeyError, FileNotFoundError):
            cli.Exit.INVALID_DATASET(
                "This script presupposes a separate CognatesetTable. Call `lexedata.edit.add_table CognatesetTable` to automatically add one."
            )
        try:
            for _ in dataset["CognateTable"]:
                break
        except (KeyError, FileNotFoundError):
            cli.Exit.NO_COGNATETABLE(
                "This script presupposes a separate CognateTable. Call `lexedata.edit.add_cognate_table` to automatically add one."
            )
        self.singleton = singleton_cognate
        self.singleton_status = singleton_status
        if self.singleton_status is not None:
            if ("Status_Column", "Status_Column") not in self.header:
                self.logger.warning(
                    f"You requested that I set the status of new singleton cognate sets to {self.singleton_status}, but your CognatesetTable has no Status_Column to write it to. If you want a Status "
                )

    def write_row_header(self, cogset, row_number: int):
        for col, (db_name, header) in enumerate(self.header, 1):
            # db_name is '' when add_central_concepts is activated
            # and there is no concept column in cognateset table
            # else read value from cognateset table
            if header == "Central_Concept" and db_name == "":
                # this is the concept associated to the first cognate in this cognateset
                raise NotImplementedError(
                    "You expect central conceps in your cognate set table, but you don't have any central concepts stored with your cognate sets"
                )
            try:
                value = self.separators[db_name].join([str(v) for v in cogset[db_name]])
            except KeyError:
                # No separator
                value = cogset.get(db_name, "")
            cell = self.ws.cell(row=row_number, column=col, value=value)
            # Transfer the cognateset comment to the first Excel cell.
            if col == 1 and cogset.get("comment"):
                cell.comment = op.comments.Comment(
                    re.sub(f"-?{__package__}", "", cogset["comment"]).strip(),
                    "lexedata.exporter",
                )

    def set_header(
        self,
        dataset: types.Wordlist[
            types.Language_ID,
            types.Form_ID,
            types.Parameter_ID,
            types.Cognate_ID,
            types.Cognateset_ID,
        ],
    ):
        c_id = dataset["CognatesetTable", "id"].name
        try:
            c_comment = dataset["CognatesetTable", "comment"].name
        except (KeyError):
            c_comment = None
        self.header = []
        for column in dataset["CognatesetTable"].tableSchema.columns:
            if column.name == c_id:
                self.header.insert(0, ("id", "CogSet"))
            elif column.name == c_comment:
                continue
            else:
                property = util.cldf_property(column.propertyUrl) or column.name
                self.header.append((property, column.name))

    def form_to_cell_value(self, form: types.Form) -> str:
        """Build a string describing the form itself

        Provide the best transcription and all translations of the form strung
        together.

        >>> ds = util.fs.new_wordlist(FormTable=[], CognatesetTable=[], CognateTable=[])
        >>> E = ExcelWriter(dataset=ds)
        >>> E.form_to_cell_value({"form": "f", "parameterReference": "c"})
        'f ‘c’'
        >>> E.form_to_cell_value(
        ...   {"form": "f", "parameterReference": "c", "formComment": "Not empty"})
        'f ‘c’ ⚠'
        >>> E.form_to_cell_value(
        ...   {"form": "fo", "parameterReference": "c", "segments": ["f", "o"]})
        '{ f o } ‘c’'
        >>> E.form_to_cell_value(
        ...   {"form": "fo",
        ...    "parameterReference": "c",
        ...    "segments": ["f", "o"],
        ...    "segmentSlice": ["1:1"]})
        '{ f }o ‘c’'

        TODO: This function should at some point support alignments, so that
        the following call will return '{ - f - }o ‘c’' instead.

        >>> E.form_to_cell_value(
        ...   {"form": "fo",
        ...    "parameterReference": "c",
        ...    "segments": ["f", "o"],
        ...    "segmentSlice": ["1:1"],
        ...    "alignment": ["", "f", ""]})
        '{ f }o ‘c’'

        """
        segments = form.get("segments")
        if not segments:
            transcription = form["form"]
        else:
            transcription = ""
            # TODO: use CLDF property instead of column name
            included_segments: t.Iterable[int]
            try:
                included_segments = set(
                    parse_segment_slices(form["segmentSlice"], enforce_ordered=True)
                )
            except TypeError:
                self.logger.warning(
                    "In judgement %s, for form %s, there was no segment slice. I will use the whole form.",
                    form["cognateReference"],
                    form["id"],
                )
                included_segments = range(len(form["segments"]))
            except KeyError:
                included_segments = range(len(form["segments"]))
            except ValueError:
                # What if segments overlap or cross? Overlap shouldn't happen,
                # but we don't check here. Crossing might happen, but this
                # serialization cannot reflect it, so we enforce order,
                # expecting that an error message here will be more useful than
                # silently messing with data. If the check fails, we take the
                # whole segment and warn.
                self.logger.warning(
                    "In judgement %s, for form %s, segment slice %s is invalid. I will use the whole form.",
                    form["cognateReference"],
                    form["id"],
                    ",".join(form["segmentSlice"]),
                )
                included_segments = range(len(form["segments"]))

            included = False
            for i, s in enumerate(segments):
                if included and i not in included_segments:
                    transcription += " }" + s
                    included = False
                elif not included and i in included_segments:
                    transcription += "{ " + s
                    included = True
                elif i in included_segments:
                    transcription += " " + s
                else:
                    transcription += s
            if included:
                transcription += " }"

            transcription = transcription.strip()
        translations = []

        suffix = ""
        try:
            if form.get("formComment"):
                suffix = f" {WARNING:}"
        except (KeyError):
            pass

        # corresponding concepts
        # (multiple concepts) and others (single concept)
        if isinstance(form["parameterReference"], list):
            for f in form["parameterReference"]:
                translations.append(f)
        else:
            translations.append(form["parameterReference"])
        return "{:} ‘{:}’{:}".format(transcription, ", ".join(translations), suffix)


def properties_as_key(data, columns):
    mapping = {
        column.name: util.cldf_property(column.propertyUrl)
        for column in columns
        if util.cldf_property(column.propertyUrl)
    }
    for s in data:
        for name, property in mapping.items():
            s[property] = s.pop(name, None)


def sort_cognatesets(
    cogsets: t.List[types.CogSet],
    judgements: t.Sequence[types.Judgement] = [],
    sort_column: t.Optional[str] = None,
    size: bool = True,
) -> None:
    """Sort cognatesets by a given column, and optionally by size."""
    # Sort first by size, then by the specified column, so that if both
    # happen, the cognatesets are globally sorted by the specified column
    # and within one group by size.
    if size:
        cogsets.sort(
            key=lambda x: len(
                [j for j in judgements if j["cognatesetReference"] == x["id"]]
            ),
            reverse=True,
        )

    if sort_column:
        cogsets.sort(key=lambda c: c[sort_column])


def parser():
    parser = cli.parser(
        __package__ + "." + Path(__file__).stem,
        description="Create an Excel cognate view from a CLDF dataset",
    )
    parser.add_argument(
        "excel",
        type=Path,
        help="File path for the generated cognate excel file.",
    )
    parser.add_argument(
        "--size-sort",
        action="store_true",
        default=False,
        help="List the biggest cognatesets first (within a group, if another sort order is specified by --sort-cognatesets-by)",
    )
    parser.add_argument(
        "--sort-languages-by",
        help="The name of a column in the LanguageTable to sort languages by in the output",
    )
    parser.add_argument(
        "--sort-cognatesets-by",
        help="The name of a column in the CognatesetTable to sort cognates by in the output",
    )
    parser.add_argument(
        "--url-template",
        type=str,
        default="https://example.org/lexicon/{:}",
        help="A template string for URLs pointing to individual forms. For example, to"
        " point to lexibank, you would use https://lexibank.clld.org/values/{:}."
        " (default: https://example.org/lexicon/{:})",
    )
    parser.add_argument(
        "--add-singletons-with-status",
        default=None,
        metavar="MESSAGE",
        help="Include in the output all forms that don't belong to a cognateset. For each form, a singleton cognateset is created, and its status column (if there is one) is set to MESSAGE.",
    )
    parser.add_argument(
        "--add-singletons",
        action="store_const",
        const="automatic singleton",
        help="Short for `--add-singletons-with-status='automatic singleton'`",
        dest="add_singletons_with_status",
    )
    parser.add_argument(
        "--by-segment",
        default=False,
        action="store_true",
        help="If adding singletons: Instead of creating singleton cognate sets only for forms that are not cognate coded at all, make sure every contiguous set of segments in every form is in a cognate set.",
    )
    return parser


def cogsets_and_judgements(
    dataset,
    status: t.Optional[str],
    by_segment=True,
    logger: cli.logging.Logger = cli.logger,
):
    if status is not None:
        cogsets, judgements = create_singletons(
            dataset,
            status=status,
            by_segment=by_segment,
            logger=logger,
        )
        properties_as_key(cogsets, dataset["CognatesetTable"].tableSchema.columns)
        properties_as_key(judgements, dataset["CognateTable"].tableSchema.columns)
    else:
        cogsets = util.cache_table(dataset, "CognatesetTable").values()
        judgements = util.cache_table(dataset, "CognateTable").values()

    return list(cogsets), list(judgements)


if __name__ == "__main__":  # pragma: no cover
    args = parser().parse_args()
    logger = cli.setup_logging(args)

    dataset = pycldf.Wordlist.from_metadata(args.metadata)
    try:
        cogsets = list(dataset["CognatesetTable"])
    except (KeyError):
        cli.Exit.INVALID_DATASET(
            "Dataset has no explicit CognatesetTable. Add one using `lexedata.edit.add_table CognatesetTable`."
        )

    E = ExcelWriter(
        dataset,
        database_url=args.url_template,
        logger=logger,
    )

    cogsets, judgements = cogsets_and_judgements(
        dataset, args.add_singletons_with_status, args.by_segment, logger
    )
    if args.sort_cognatesets_by:
        try:
            cogset_order = (
                util.cldf_property(
                    dataset["CognatesetTable", args.sort_cognatesets_by].propertyUrl
                )
                or dataset["CognatesetTable", args.sort_cognatesets_by].name
            )
        except (KeyError):
            cli.Exit.INVALID_COLUMN_NAME(
                f"No column '{args.sort_cognatesets_by}' in your CognatesetTable."
            )
    else:
        cogset_order = None
    sort_cognatesets(cogsets, judgements, cogset_order, size=args.size_sort)

    # TODO: wrap the following two blocks into a
    # get_sorted_languages() -> t.OrderedDict[languageReference, Column Header/Titel/Name]
    # function
    languages = list(util.cache_table(dataset, "LanguageTable").values())
    if args.sort_languages_by:
        c_sort = (
            util.cldf_property(
                dataset["LanguageTable", args.sort_languages_by].propertyUrl
            )
            or dataset["LanguageTable", args.sort_languages_by].name
        )
        languages.sort(key=lambda x: x[c_sort], reverse=False)

    forms = util.cache_table(dataset)

    E.create_excel(
        size_sort=args.size_sort,
        languages=languages,
        rows=cogsets,
        judgements=judgements,
        forms=forms,
    )
    E.wb.save(
        filename=args.excel,
    )
