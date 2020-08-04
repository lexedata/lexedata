# -*- coding: utf-8 -*-
import typing as t
import urllib.parse
from pathlib import Path
from collections import OrderedDict, defaultdict

import pycldf
import sqlalchemy
import openpyxl as op
import sqlalchemy
import sqlalchemy.ext.automap
from sqlalchemy.ext.automap import automap_base

import lexedata.cldf.db as db

WARNING = "\u26A0"

Form = t.TypeVar("Form", bound=sqlalchemy.ext.automap.AutomapBase)
CogSet = t.TypeVar("CogSet", bound=sqlalchemy.ext.automap.AutomapBase)

# ----------- Remark: Indices in excel are always 1-based. -----------

# TODO: ExcelWrite still uses SQLAlchemy, which we have nearly completely taken
# out. Rewrite those bits and get it to work – It's probably worth thinking
# about doing this without touching the database at all, and just work on the
# iterators like dataset[FormTable'], because here we don't need the smart
# lookup capabilities of the database to match similar forms.

class ExcelWriter():
    """Class logic for cognateset Excel export."""
    header = [("cldf_id", "CogSet")]  # Add columns here for other datasets.

    def __init__(self, dataset: pycldf.Dataset, database: str,
                 url_base: t.Optional[str] = None, **kwargs):
        super().__init__(dataset, fname=database, **kwargs)

        self.URL_BASE = "https://example.org/{:s}"
        if url_base:
            self.URL_BASE = url_base
        engine = sqlalchemy.create_engine(f"sqlite:///{database:}")
        self.session = sqlalchemy.orm.Session(engine)

    def create_excel(
            self,
            out: Path) -> None:
        """Convert the initial CLDF into an Excel cognate view

        The Excel file has columns "CogSet", "Tags", and then one column per
        language.

        The rows contain cognate data. If a language has multiple reflexes in
        the same cognateset, these appear in different cells, one below the
        other.

        Parameters
        ==========
        out: The path of the Excel file to be written.

        """
        # TODO: Check whether openpyxl.worksheet._write_only.WriteOnlyWorksheet
        # will be useful (maybe it's faster or prevents us from some mistakes)?
        # https://openpyxl.readthedocs.io/en/stable/optimized.html#write-only-mode
        wb = op.Workbook()
        ws: op.worksheet.worksheet.Worksheet = wb.active

        # Define the columns
        self.lan_dict: t.Dict[str, int] = {}
        excel_header = [name for cldf, name in self.header]
        for col, lan in enumerate(
                self.session.query(self.Language).all(),
                len(excel_header) + 1):
            # len(header) + 1 refers to column 3
            self.lan_dict[lan.cldf_id] = col
            excel_header.append(lan.cldf_name)

        ws.append(excel_header)

        # Iterate over all cognate sets, and prepare the rows.
        # Again, row_index 2 is indeed row 2
        row_index = 1 + 1
        for cogset in self.session.query(self.CogSet):
            # Put the cognateset's tags in column B.
            for col, (db_name, header) in enumerate(self.header, 1):
                cell = ws.cell(row=row_index, column=col,
                               value=getattr(cogset, db_name))
                # Transfer the cognateset comment to the first Excel cell.
                if col == 1 and cogset.cldf_comment:
                    cell.comment = op.comments.Comment(
                        cogset.cldf_comment, __package__)

            new_row_index = self.create_formcells_for_cogset(
                cogset, ws, row_index, self.lan_dict)
            row_index = new_row_index
        wb.save(filename=out)

    def create_formcells_for_cogset(
            self,
            cogset: CogSet,
            ws: op.worksheet.worksheet.Worksheet,
            row_index: int,
            # FIXME: That's not just a str, it's a language_id, but String()
            # alias Language.id.type is not a Type, according to `typing`.
            lan_dict: t.Dict[str, int]) -> int:
        """Writes all forms for given cognate set to Excel.

        Take all forms for a given cognate set as given by the database, create
        a hyperlink cell for each form, and write those into rows starting at
        row_index.

        Return the row number of the first empty row after this cognate set,
        which can then be filled by the following cognate set.

        """
        # Read the forms from the database and group them by language
        forms = t.DefaultDict[int, t.List[Form]](list)
        for judgement in cogset.cognates:
            form: Form = judgement.form
            forms[self.lan_dict[form.language.cldf_id]].append(judgement)

        if not forms:
            return row_index + 1

        # maximum of rows to be added
        maximum_cogset = max([len(c) for c in forms.values()])
        for column, cells in forms.items():
            for row, judgement in enumerate(cells, row_index):
                self.create_formcell(judgement, ws, column, row)
        # increase row_index and return
        row_index += maximum_cogset

        return row_index

    def create_formcell(
            self,
            judgement,
            ws: op.worksheet.worksheet.Worksheet,
            column: int,
            row: int) -> None:
        """Fill the given cell with the form's data.

        In the cell described by ws, column, row, dump the data for the form:
        Write into the the form data, and supply a comment from the judgement
        if there is one.

        """
        cell_value = self.form_to_cell_value(judgement.form)
        form_cell = ws.cell(row=row, column=column, value=cell_value)
        comment = getattr(judgement, "cldf_comment", None)
        if comment:
            form_cell.comment = op.comments.Comment(comment, __package__)
        link = self.URL_BASE.format(urllib.parse.quote(judgement.form.cldf_id))
        form_cell.hyperlink = link

    def form_to_cell_value(self, form: Form) -> str:
        """Build a string describing the form itself

        Provide the best transcription and all translations of the form strung
        together.

        """

        transcription = self.get_best_transcription(form)
        translations = []

        suffix = ""
        if form.cldf_comment:
            suffix = f" {WARNING:}"

        # iterate over corresponding concepts
        try:
            for concept in form.parameters:
                translations.append(concept.cldf_name)
        except AttributeError:
            translations.append(form.parameter.cldf_name)

        return "{:} ‘{:}’{:}".format(
            transcription, ", ".join(translations), suffix)

    #def get_best_transcription(self, form):
    #    if form.phonemic:
    #        return form.phonemic
    #    elif form.phonetic:
    #        return form.phonetic
    #    elif form.orthographic:
    #        return form.orthographic
    #    else:
    #        ValueError(f"Form {form:} has no transcriptions.")

    def get_best_transcription(self, form):
        return form.cldf_form


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Create an Excel cognate view from a CLDF dataset")
    parser.add_argument("metadata", help="Path to metadata file for dataset input")
    parser.add_argument("database", help="Path to database")
    parser.add_argument("excel", help="Excel output file path")
    args = parser.parse_args()
    E = ExcelWriter(pycldf.Dataset.from_metadata(args.metadata), args.database)
    E.create_excel(args.excel)
